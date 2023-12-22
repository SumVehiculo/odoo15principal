# -*- coding: utf-8 -*-
from odoo.tools.misc import DEFAULT_SERVER_DATETIME_FORMAT
import time
import odoo.addons.decimal_precision as dp
from openerp.osv import osv
import base64
from odoo import models, fields, api
import codecs
from datetime import timedelta

values = {}

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')

import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
values = {}
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell


from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from reportlab.lib.colors import magenta, red , black , blue, gray, Color, HexColor
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table
from reportlab.lib.units import  cm,mm
from reportlab.lib.utils import simpleSplit
import decimal


class stock_location(models.Model):
	_inherit = 'stock.location'

	l10n_pe_edi_branch_code = fields.Char('Código Establecimiento Anexo')
	
	
def border(ws,texto):
	cell = WriteOnlyCell(ws, value=texto)
	cell.font = Font(name='Arial',size=13,bold=True)
	cell.border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000'))
	return cell


def number_format_costo(ws,numero):
    cell = WriteOnlyCell(ws, value=numero)
    cell.number_format = "0.000000"
    return cell

def number_format_cantidad(ws,numero):
    cell = WriteOnlyCell(ws, value=numero)
    cell.number_format = "0.00"
    return cell

def border_sub(ws,texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Arial',size=12,bold=True)
    cell.border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000'))
    return cell


class make_kardex_valorado_formato_sunat_v12(models.TransientModel):
	_name = "make.kardex.valorado.formato.sunat.v12"
	_description = "make kardex valorado formato sunat"

	fini= fields.Date('Fecha inicial',required=True)
	ffin= fields.Date('Fecha final',required=True)
	products_ids_it=fields.Many2many('product.product','rel_wiz_kardex_valorado_formato_sunat_v12','product_id','kardex_id')
	location_ids=fields.Many2many('stock.location','rel_kardex_location_valorado_formato_sunat_v12','location_id','kardex_id','Ubicacion',required=True)
	allproducts=fields.Boolean('Todos los productos',default=True)
	destino = fields.Selection([('csv','CSV')],'Destino')
	check_fecha = fields.Boolean('Editar Fecha')
	alllocations = fields.Boolean('Todos los almacenes',default=True)

	fecha_ini_mod = fields.Date('Fecha Inicial')
	fecha_fin_mod = fields.Date('Fecha Final')
	analizador = fields.Boolean('Analizador')


	@api.onchange('fecha_ini_mod')
	def onchange_fecha_ini_mod(self):
		self.fini = self.fecha_ini_mod


	@api.onchange('fecha_fin_mod')
	def onchange_fecha_fin_mod(self):
		self.ffin = self.fecha_fin_mod


	@api.model
	def default_get(self, fields):
		res = super(make_kardex_valorado_formato_sunat_v12, self).default_get(fields)
		import datetime
		fecha_hoy = str(datetime.datetime.now())[:10]
		fecha_inicial = fecha_hoy[:4] + '-01-01'
		res.update({'fecha_ini_mod':fecha_inicial})
		res.update({'fecha_fin_mod':fecha_hoy})
		res.update({'fini':fecha_inicial})
		res.update({'ffin':fecha_hoy})

		#locat_ids = self.pool.get('stock.location').search(cr, uid, [('usage','in',('internal','inventory','transit','procurement','production'))])
		locat_ids = self.env['stock.location'].search([('usage','in',('internal','inventory','transit','procurement','production'))])
		locat_ids = [elemt.id for elemt in locat_ids]
		res.update({'location_ids':[(6,0,locat_ids)]})
		return res


	def reportpdf_sunat(self):
		obj_move = False
		self.reporteador(obj_move)
		
		mod_obj = self.env['ir.model.data']
		act_obj = self.env['ir.actions.act_window']
		import os

		direccion = self.env['account.main.parameter'].search([('company_id','=',self.env.company.id)])[0].dir_create_file

		f = open(direccion+'KardexSunat.pdf', 'rb')

		return self.env['popup.it'].get_file('KardexSunat.pdf',base64.encodestring(b''.join(f.readlines())))


	def x_aument(self,a):
		a[0] = a[0]+1

	def reporteador(self,obj_move):

		#pdfmetrics.registerFont(TTFont('Calibri', 'Calibri.ttf'))
		#pdfmetrics.registerFont(TTFont('Calibri-Bold', 'CalibriBold.ttf'))

		width ,height  = A4  # 595 , 842
		wReal = height- 30
		hReal = width - 40

		direccion = self.env['account.main.parameter'].search([('company_id','=',self.env.company.id)])[0].dir_create_file
		c = canvas.Canvas( direccion + "KardexSunat.pdf", pagesize= (842,595) )
		inicio = 0
		pos_inicial = hReal
		libro = None
		voucher = None
		total = 0
		debeTotal = 0
		haberTotal = 0
		pagina = 1
		textPos = 0
		


		cad = ""

		s_prod = [-1,-1,-1]
		s_loca = [-1,-1,-1]
		if self.alllocations == True:
			locat_ids = self.env['stock.location'].search( [('usage','in',('internal','inventory','transit','procurement','production'))] )
			lst_locations = locat_ids.ids
		else:
			lst_locations = self.location_ids.ids
		lst_products  = self.products_ids_it.ids
		productos='{'
		almacenes='{'
		date_ini=self.fini
		date_fin=self.ffin
		if self.allproducts:
			lst_products = self.env['product.product'].with_context(active_test=False).search([]).ids

		else:
			lst_products = self.products_ids_it.ids

		if len(lst_products) == 0:
			raise osv.except_osv('Alerta','No existen productos seleccionados')

		for producto in lst_products:
			productos=productos+str(producto)+','
			s_prod.append(producto)
		productos=productos[:-1]+'}'
		for location in lst_locations:
			almacenes=almacenes+str(location)+','
			s_loca.append(location)
		almacenes=almacenes[:-1]+'}'
		# raise osv.except_osv('Alertafis',[almacenes,productos])


		t = self.env['make.kardex.valorado'].create({
				'fini':date_ini,
				'ffin':date_fin,
				'fecha_ini_mod':date_ini,
				'fecha_fin_mod':date_fin,
				'allproducts':True,
				'alllocations':True,
				'moneda':'pen',
				})

		t.with_context({'res_model_it':'make.kardex.valorado.formato.sunat','id_it':self.id}).guardado()


		self.env.cr.execute("""
			 select

				fecha_albaran as "Fecha Alb.",	
				fecha as "Fecha",
				type_doc as "T. Doc.",
				serial as "Serie",
				nro as "Nro. Documento",
				numdoc_cuadre as "Nro. Documento",
				doc_partner as "Nro Doc. Partner",
				name as "Proveedor",							
				operation_type as "Tipo de operacion",				 
				name_template as "Producto",
				unidad as "Unidad",			
				ingreso as "Ingreso Fisico",
				round(debit,6) as "Ingreso Valorado.",
				salida as "Salida Fisico",
				round(credit,6) as "Salida Valorada",
				saldof as "Saldo Fisico",
				round(saldov,6) as "Saldo valorado",
				round(cadquiere,6) as "Costo adquisicion",
				round(cprom,6) as "Costo promedio",
					origen as "Origen",
					destino as "Destino",
				almacen AS "Almacen",
				default_code as "Cod Producto",
				null::varchar "guia remision",
				product_id

			from get_kardex_v("""+ str(date_ini).replace('-','') + "," + str(date_fin).replace('-','') + ",'" + productos + """'::INT[], '""" + almacenes + """'::INT[], """ +str(self.env.company.id)+ """)

		""")

		ingreso1= 0
		ingreso2= 0
		salida1= 0
		salida2= 0

		product_index = False
		almacen_index = False
		total_1 = 0
		total_2 = 0
		total_3 = 0 
		total_4 = 0
		total_5 = 0
		primero = True
		for line in self.env.cr.fetchall():
			if line[21] != almacen_index or line[9] != product_index:
				primero = True
				if product_index != False:

					c.drawString( 216 ,pos_inicial,self.particionar_text( 'TOTAL',35) )

					# c.drawRightString( 366+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_1) )))
					c.drawRightString( 466+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_1) )))
					# c.drawRightString( 466+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_2) )))

					# c.drawRightString( 516+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_3) )))
					c.drawRightString( 616+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_3) )))
					# c.drawRightString( 616+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_4) )))
					# c.drawRightString( 766+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_5) )))

					total_1 = 0
					total_2 = 0
					total_3 = 0 
					total_4 = 0
					total_5 = 0
					c.showPage()
					pos_inicial = hReal

				c.setFont("Helvetica-Bold", 10)
				c.setFillColor(black)
				c.drawCentredString((wReal/2)+20,pos_inicial, 'FORMATO 12.1: "REGISTRO DEL INVENTARIO PERMANENTE EN UNIDADES FÍSICAS - DETALLE DEL INVENTARIO PERMANENTE EN UNIDADES FÍSICAS"')

				c.setFont("Helvetica-Bold", 8)

				c.drawString( 10,pos_inicial-20, 'PERIODO')
				c.drawString( 10,pos_inicial-32, 'RUC')
				c.drawString( 10,pos_inicial-44, 'APELLIDO Y NOMBRES, DEMONINACIÓN o RAZÓN SOCIAL')
				c.drawString( 10,pos_inicial-56, 'ESTABLECIMIENTO')
				c.drawString( 10,pos_inicial-68, 'CÓDIGO DE EXISTENCIA')
				c.drawString( 10,pos_inicial-80, 'TIPO')
				c.drawString( 10,pos_inicial-92, 'DESCRIPCIÓN')
				c.drawString( 10,pos_inicial-104, 'CÓDIGO DE LA UNIDAD DE MEDIDA')
				c.drawString( 10,pos_inicial-116, 'MÉTODO DE EVALUACIÓN')


				c.setFont("Helvetica", 8)		
				c.drawString( 270,pos_inicial-20, str(date_ini) + ' - ' + str(date_fin))
				c.drawString( 270,pos_inicial-32, self.env.company.partner_id.vat or '' )
				c.drawString( 270,pos_inicial-44, self.env.company.partner_id.name or '')
				c.drawString( 270,pos_inicial-56, line[21] or '')
				c.drawString( 270,pos_inicial-68, line[22] or '')
				producto_obj = self.env['product.product'].browse(line[24])
				c.drawString( 270,pos_inicial-80, producto_obj.categ_id.existence_type_id.code or '')
				c.drawString( 270,pos_inicial-92, line[9] or '')
				c.drawString( 270,pos_inicial-104,'')#producto_obj.uom_id.einvoice_06.code or '')
				c.drawString( 270,pos_inicial-116, 'Costo Promedio')

				pos_inicial = pos_inicial - 128
				almacen_index = line[21]
				product_index = line[9]

				c.setFont("Helvetica-Bold", 8)

				c.drawString( 10 ,pos_inicial, "DOCUMENTO DE TRANSLADO," )
				c.drawString( 10 ,pos_inicial-12, "COMPROBANTE DE PAGO," )
				c.drawString( 10 ,pos_inicial-24, "DOCUMENTO INTERIOR O SIMILAR" )

				c.line( 7, pos_inicial-72-3, 816+2 ,pos_inicial-72-3)
				c.line( 7, pos_inicial+9, 816+2 ,pos_inicial+9)

				c.line( 7, pos_inicial-72-3, 7 ,pos_inicial+9)
				c.line( 816+2, pos_inicial-72-3, 816+2 ,pos_inicial+9)

				c.line( 10+156-2, pos_inicial-72-3, 10+156-2 ,pos_inicial+9)
				c.line( 216-2, pos_inicial-72-3, 216-2 ,pos_inicial+9)
				c.line( 366-2, pos_inicial-72-3, 366-2 ,pos_inicial+9) #LINEA IZQ ENTRADAS

				c.line( 7, pos_inicial-24-3, 10+156-2 ,pos_inicial-24-3) # SALDO FINAL


				c.line( 10+36-2, pos_inicial-24-3, 10+36-2 ,pos_inicial-72-3)
				c.line( 10+66-2, pos_inicial-24-3, 10+66-2 ,pos_inicial-72-3)
				c.line( 10+96-2, pos_inicial-24-3, 10+96-2 ,pos_inicial-72-3)


				c.line( 366-2, pos_inicial-3, 816+2 ,pos_inicial-3)# SALDO FINAL

				# c.line( 416-2, pos_inicial-3, 416-2 ,pos_inicial-72-3)
				# c.line( 466-2, pos_inicial-3, 466-2 ,pos_inicial-72-3)
				c.line( 516-2, pos_inicial+9, 516-2 ,pos_inicial-72-3)

				# c.line( 566-2, pos_inicial-3, 566-2 ,pos_inicial-72-3)
				# c.line( 616-2, pos_inicial-3, 616-2 ,pos_inicial-72-3)
				c.line( 666-2, pos_inicial+9, 666-2 ,pos_inicial-72-3)

				# c.line( 716-2, pos_inicial-3, 716-2 ,pos_inicial-72-3)
				# c.line( 766-2, pos_inicial-3, 766-2 ,pos_inicial-72-3)


				c.drawString( 10 ,pos_inicial-52, "FECHA" )
				c.drawString( 10 + 36 ,pos_inicial-52, "TIPO" )
				c.drawString( 10 + 66 ,pos_inicial-52, "SERIE" )
				c.drawString( 10 + 96 ,pos_inicial-52, "NUMERO" )


				c.drawString( 10 + 156 ,pos_inicial, "TIPO" )
				c.drawString( 10 + 156 ,pos_inicial-12, "DE" )
				c.drawString( 10 + 156 ,pos_inicial-24, "OPERA" )
				c.drawString( 10 + 156 ,pos_inicial-36, "CIÓN" )
				c.drawString( 10 + 156 ,pos_inicial-48, "(Tabla" )
				c.drawString( 10 + 156 ,pos_inicial-60, "12)" )


				c.drawString( 216 ,pos_inicial-52, "CLIENTE / PROVEEDOR" )

				c.drawString( 420 ,pos_inicial-52, "CANTIDAD" )
				c.drawString( 416 ,pos_inicial, "ENTRADAS" )
				# c.drawString( 416 ,pos_inicial-52, "C.U." )
				# c.drawString( 466 ,pos_inicial-52, "COSTO" )


				c.drawString( 570 ,pos_inicial-52, "CANTIDAD" )
				c.drawString( 566 ,pos_inicial, "SALIDAS" )
				# c.drawString( 566 ,pos_inicial-52, "C.U." )
				# c.drawString( 616 ,pos_inicial-52, "COSTO" )


				c.drawString( 720 ,pos_inicial-52, "CANTIDAD" )
				c.drawString( 716 ,pos_inicial, "SALDO FINAL" )
				# c.drawString( 716 ,pos_inicial-52, "C.U." )
				# c.drawString( 766 ,pos_inicial-52, "COSTO" )


				pos_inicial = pos_inicial -	72

			if primero:

				c.setFont("Helvetica", 6)	
				c.drawString( 10 ,pos_inicial, self.particionar_text( str(line[1]) if line[1] else '',36) )
				c.drawString( 10+36 ,pos_inicial, self.particionar_text( line[2] if line[2] else '',30) )
				c.drawString( 10+66 ,pos_inicial,self.particionar_text( line[3] if line[3] else '',30) )
				c.drawString( 10+96 ,pos_inicial,self.particionar_text( line[4] if line[4] else '',35) )

				c.drawString( 10+156 ,pos_inicial,self.particionar_text(  '16',35) )
				c.drawString( 216 ,pos_inicial,self.particionar_text(  'Saldo Inicial',35) )

				cantidadx =((line[15] if line[15] else 0) - (line[11] if line[11] else 0) +   (line[13] if line[13] else 0) )
				montox = ((line[16] if line[16] else 0) - (line[12] if line[12] else 0) +   (line[14] if line[14] else 0) )
				# c.drawRightString( 366+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[11] if line[11] else 0) )))
				# c.drawRightString( 416+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % ( (line[12] if line[12] else 0)/(line[11] if line[11] else 1)     ) )))
				# c.drawRightString( 466+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[12] if line[12] else 0) )))
				c.drawRightString( 466+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % cantidadx )))

				# c.drawRightString( 516+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[13] if line[13] else 0) )))
				# c.drawRightString( 566+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % ( (line[14] if line[14] else 0)/(line[13] if line[13] else 1)     ) )))
				# c.drawRightString( 616+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[14] if line[14] else 0) )))
				c.drawRightString( 616+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (0) )))

				# c.drawRightString( 666+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[15] if line[15] else 0) )))
				# c.drawRightString( 716+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % ( (line[16] if line[16] else 0)/(line[15] if line[15] else 1)     ) )))
				# c.drawRightString( 766+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[16] if line[16] else 0) )))
				c.drawRightString( 766+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (cantidadx) )))

				total_1 += cantidadx
				total_2 += 0
				total_3 += 0
				total_4 += 0
				total_5 += 0
				
				c.line( 7, pos_inicial-3, 816+2 ,pos_inicial-3)
				c.line( 7, pos_inicial+9, 816+2 ,pos_inicial+9)

				c.line( 7, pos_inicial-3, 7 ,pos_inicial+9)
				c.line( 816+2, pos_inicial-3, 816+2 ,pos_inicial+9)

				c.line( 10+36-2, pos_inicial-3, 10+36-2 ,pos_inicial+9)
				c.line( 10+66-2, pos_inicial-3, 10+66-2 ,pos_inicial+9)
				c.line( 10+96-2, pos_inicial-3, 10+96-2 ,pos_inicial+9)
				c.line( 10+156-2, pos_inicial-3, 10+156-2 ,pos_inicial+9)
				c.line( 216-2, pos_inicial-3, 216-2 ,pos_inicial+9)
				c.line( 366-2, pos_inicial-3, 366-2 ,pos_inicial+9)
				# c.line( 416-2, pos_inicial-3, 416-2 ,pos_inicial+9)
				# c.line( 466-2, pos_inicial-3, 466-2 ,pos_inicial+9)
				c.line( 516-2, pos_inicial-3, 516-2 ,pos_inicial+9)
				# c.line( 566-2, pos_inicial-3, 566-2 ,pos_inicial+9)
				# c.line( 616-2, pos_inicial-3, 616-2 ,pos_inicial+9)
				c.line( 666-2, pos_inicial-3, 666-2 ,pos_inicial+9)
				# c.line( 716-2, pos_inicial-3, 716-2 ,pos_inicial+9)
				# c.line( 766-2, pos_inicial-3, 766-2 ,pos_inicial+9)

				pagina, pos_inicial = self.verify_linea(c,wReal,hReal,pos_inicial,12,pagina,2,obj_move)
				primero = False

			c.setFont("Helvetica", 6)	
			c.drawString( 10 ,pos_inicial, self.particionar_text( str(line[1]) if line[1] else '',36) )
			c.drawString( 10+36 ,pos_inicial, self.particionar_text( line[2] if line[2] else '',30) )
			c.drawString( 10+66 ,pos_inicial,self.particionar_text( line[3] if line[3] else '',30) )
			c.drawString( 10+96 ,pos_inicial,self.particionar_text( line[4] if line[4] else '',35) )

			c.drawString( 10+156 ,pos_inicial,self.particionar_text( line[8] if line[8] else '',35) )
			c.drawString( 216 ,pos_inicial,self.particionar_text( line[7] if line[7] else '',35) )

			# c.drawRightString( 366+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[11] if line[11] else 0) )))
			# c.drawRightString( 416+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % ( (line[12] if line[12] else 0)/(line[11] if line[11] else 1)     ) )))
			# c.drawRightString( 466+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[12] if line[12] else 0) )))
			c.drawRightString( 466+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[11] if line[11] else 0) )))

			# c.drawRightString( 516+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[13] if line[13] else 0) )))
			# c.drawRightString( 566+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % ( (line[14] if line[14] else 0)/(line[13] if line[13] else 1)     ) )))
			# c.drawRightString( 616+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[14] if line[14] else 0) )))
			c.drawRightString( 616+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[13] if line[13] else 0) )))

			# c.drawRightString( 666+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[15] if line[15] else 0) )))
			# c.drawRightString( 716+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % ( (line[16] if line[16] else 0)/(line[15] if line[15] else 1)     ) )))
			# c.drawRightString( 766+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[16] if line[16] else 0) )))
			c.drawRightString( 766+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[15] if line[15] else 0) )))

			total_1 += line[11] if line[11] else 0
			total_2 += line[12] if line[12] else 0
			total_3 += line[13] if line[13] else 0
			total_4 += line[14] if line[14] else 0
			total_5 += line[15] if line[15] else 0
			
			c.line( 7, pos_inicial-3, 816+2 ,pos_inicial-3)
			c.line( 7, pos_inicial+9, 816+2 ,pos_inicial+9)

			c.line( 7, pos_inicial-3, 7 ,pos_inicial+9)
			c.line( 816+2, pos_inicial-3, 816+2 ,pos_inicial+9)

			c.line( 10+36-2, pos_inicial-3, 10+36-2 ,pos_inicial+9)
			c.line( 10+66-2, pos_inicial-3, 10+66-2 ,pos_inicial+9)
			c.line( 10+96-2, pos_inicial-3, 10+96-2 ,pos_inicial+9)
			c.line( 10+156-2, pos_inicial-3, 10+156-2 ,pos_inicial+9)
			c.line( 216-2, pos_inicial-3, 216-2 ,pos_inicial+9)
			c.line( 366-2, pos_inicial-3, 366-2 ,pos_inicial+9)
			# c.line( 416-2, pos_inicial-3, 416-2 ,pos_inicial+9)
			# c.line( 466-2, pos_inicial-3, 466-2 ,pos_inicial+9)
			c.line( 516-2, pos_inicial-3, 516-2 ,pos_inicial+9)
			# c.line( 566-2, pos_inicial-3, 566-2 ,pos_inicial+9)
			# c.line( 616-2, pos_inicial-3, 616-2 ,pos_inicial+9)
			c.line( 666-2, pos_inicial-3, 666-2 ,pos_inicial+9)
			# c.line( 716-2, pos_inicial-3, 716-2 ,pos_inicial+9)
			# c.line( 766-2, pos_inicial-3, 766-2 ,pos_inicial+9)

			pagina, pos_inicial = self.verify_linea(c,wReal,hReal,pos_inicial,12,pagina,2,obj_move)
		if product_index != False:

					c.drawString( 216 ,pos_inicial,self.particionar_text( 'TOTAL',35) )

					# c.drawRightString( 366+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_1) )))
					c.drawRightString( 466+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_1) )))
					# c.drawRightString( 466+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_2) )))

					# c.drawRightString( 516+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_3) )))
					c.drawRightString( 616+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_3) )))
					# c.drawRightString( 616+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_4) )))
					# c.drawRightString( 766+50-2 ,pos_inicial, '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_5) )))

					total_1 = 0
					total_2 = 0
					total_3 = 0 
					total_4 = 0
					total_5 = 0
					c.showPage()
					pos_inicial = hReal

		c.save()


	def particionar_text(self,c,tam):
		tet = ""
		for i in range(len(c)):
			tet += c[i]
			lines = simpleSplit(tet,'Helvetica',8,tam)
			if len(lines)>1:
				return tet[:-1]
		return tet

	def verify_linea(self,c,wReal,hReal,posactual,valor,pagina,titulo,obj_move):
		if posactual <40:
			c.showPage()
			c.setFont("Helvetica-Bold", 8)
			#c.drawCentredString(300,25,'Pág. ' + str(pagina+1))
			return pagina+1,hReal-83
		else:
			return pagina,posactual-valor

	def reportxls_sunat(self):
		cad = ""

		s_prod = [-1,-1,-1]
		s_loca = [-1,-1,-1]
		if self.alllocations == True:
			locat_ids = self.env['stock.location'].search( [('usage','in',('internal','inventory','transit','procurement','production'))] )
			lst_locations = locat_ids.ids
		else:
			lst_locations = self.location_ids.ids
		lst_products  = self.products_ids_it.ids
		productos='{'
		almacenes='{'
		date_ini=self.fini
		date_fin=self.ffin
		if self.allproducts:
			lst_products = self.env['product.product'].with_context(active_test=False).search([]).ids

		else:
			lst_products = self.products_ids_it.ids

		if len(lst_products) == 0:
			raise osv.except_osv('Alerta','No existen productos seleccionados')

		for producto in lst_products:
			productos=productos+str(producto)+','
			s_prod.append(producto)
		productos=productos[:-1]+'}'
		for location in lst_locations:
			almacenes=almacenes+str(location)+','
			s_loca.append(location)
		almacenes=almacenes[:-1]+'}'
		# raise osv.except_osv('Alertafis',[almacenes,productos])


		t = self.env['make.kardex.valorado'].create({
				'fini':date_ini,
				'ffin':date_fin,
				'fecha_ini_mod':date_ini,
				'fecha_fin_mod':date_fin,
				'allproducts':True,
				'alllocations':True,
				'moneda':'pen',
				})

		t.with_context({'res_model_it':'make.kardex.valorado.formato.sunat','id_it':self.id}).guardado()


		import io
		output = io.BytesIO()

		workbook = Workbook(write_only=True)
		ws = workbook.create_sheet("Kardex Formato Sunat")
		x= 10
		tam_col = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
	
 
		ws.column_dimensions['A'].width = 15
		ws.column_dimensions['B'].width = 15
		ws.column_dimensions['E'].width = 15
		ws.column_dimensions['F'].width = 15
		ws.column_dimensions['G'].width = 15
		ws.column_dimensions['H'].width = 15
		ws.column_dimensions['J'].width = 15
		ws.column_dimensions['K'].width = 15
		ws.column_dimensions['L'].width = 15
		ws.column_dimensions['M'].width = 15
		ws.column_dimensions['N'].width = 15
		ws.column_dimensions['O'].width = 15
		ws.column_dimensions['P'].width = 15
		ws.column_dimensions['Q'].width = 15
		ws.column_dimensions['R'].width = 15
		ws.column_dimensions['S'].width = 15
		ws.column_dimensions['T'].width = 15
		ws.column_dimensions['U'].width = 15
		ws.column_dimensions['V'].width = 15
		ws.column_dimensions['W'].width = 15

    

		self.env.cr.execute("""
			 select

				fecha_albaran as "Fecha Alb.",	
				fecha as "Fecha",
				type_doc as "T. Doc.",
				serial as "Serie",
				nro as "Nro. Documento",
				numdoc_cuadre as "Nro. Documento",
				doc_partner as "Nro Doc. Partner",
				name as "Proveedor",							
				operation_type as "Tipo de operacion",				 
				name_template as "Producto",
				unidad as "Unidad",			
				ingreso as "Ingreso Fisico",
				round(debit,6) as "Ingreso Valorado.",
				salida as "Salida Fisico",
				round(credit,6) as "Salida Valorada",
				saldof as "Saldo Fisico",
				round(saldov,6) as "Saldo valorado",
				round(cadquiere,6) as "Costo adquisicion",
				round(cprom,6) as "Costo promedio",
					origen as "Origen",
					destino as "Destino",
				almacen AS "Almacen",
				default_code as "Cod Producto",
				null::varchar "guia remision",
				product_id

			from get_kardex_v("""+ str(date_ini).replace('-','') + "," + str(date_fin).replace('-','') + ",'" + productos + """'::INT[], '""" + almacenes + """'::INT[], """ +str(self.env.company.id)+ """)

		""")

		ingreso1= 0
		ingreso2= 0
		salida1= 0
		salida2= 0

		product_index = False
		almacen_index = False
		total_1 = 0
		total_2 = 0
		total_3 = 0 
		total_4 = 0
		total_5 = 0
		primero = True
		for line in self.env.cr.fetchall():
			if line[21] != almacen_index or line[9] != product_index:
				primero = True
				if product_index != False:

					ws.append(["","","","","","TOTAL", '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_1) )) , '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_3) ))])
					
					total_1 = 0
					total_2 = 0
					total_3 = 0 
					total_4 = 0
					total_5 = 0
					
					ws.append([""])
					ws.append([""])


				ws.append([""])
				ws.append([""])
				cell = WriteOnlyCell(ws, value='FORMATO 12.1: "REGISTRO DEL INVENTARIO PERMANENTE EN UNIDADES FÍSICAS - DETALLE DEL INVENTARIO PERMANENTE EN UNIDADES FÍSICAS"')
				cell.font = Font(name='Bahnschrift Light SemiCondensed',size=33,bold=True)


				cell_dat1 = WriteOnlyCell(ws, value="PERIODO")
				cell_dat1.font = Font(name='Arial',size=10,bold=True)
				cell_dat2 = WriteOnlyCell(ws, value="RUC")
				cell_dat2.font = Font(name='Arial',size=10,bold=True)
				cell_dat3 = WriteOnlyCell(ws, value="APELLIDO Y NOMBRES, DEMONINACIÓN o RAZÓN SOCIAL")
				cell_dat3.font = Font(name='Arial',size=10,bold=True)
				cell_dat4 = WriteOnlyCell(ws, value="ESTABLECIMIENTO")
				cell_dat4.font = Font(name='Arial',size=10,bold=True)
				cell_dat5 = WriteOnlyCell(ws, value="CÓDIGO DE EXISTENCIA")
				cell_dat5.font = Font(name='Arial',size=10,bold=True)
				cell_dat6 = WriteOnlyCell(ws, value="TIPO")
				cell_dat6.font = Font(name='Arial',size=10,bold=True)
				cell_dat7 = WriteOnlyCell(ws, value="DESCRIPCIÓN")
				cell_dat7.font = Font(name='Arial',size=10,bold=True)
				cell_dat8 = WriteOnlyCell(ws, value="CÓDIGO DE LA UNIDAD DE MEDIDA")
				cell_dat8.font = Font(name='Arial',size=10,bold=True)
				cell_dat9 = WriteOnlyCell(ws, value="MÉTODO DE EVALUACIÓN")
				cell_dat9.font = Font(name='Arial',size=10,bold=True)




				ws.append(["","","","",cell])

				ws.append([""])


				ws.append([cell_dat1,str(date_ini) + ' - ' + str(date_fin)])
				ws.append([cell_dat2,self.env.company.partner_id.vat or ''])
				ws.append([cell_dat3,self.env.company.partner_id.name or ''])
				ws.append([cell_dat4,line[21] or ''])
				ws.append([cell_dat5,line[22] or ''])
				producto_obj = self.env['product.product'].browse(line[24])
				ws.append([cell_dat6,producto_obj.categ_id.existence_type_id.code or ''])
				ws.append([cell_dat7,line[9] or ''])
				ws.append([cell_dat8,''])
				ws.append([cell_dat9,'Costo Promedio'])

				almacen_index = line[21]
				product_index = line[9]
				ws.append([""])
				ws.append([""])
				ws.append([""])

					
				import datetime		


				linea = [border(ws,u"Fecha"),border(ws,u"Tipo"),border(ws,u"Serie"),border(ws,u"Numero"),border(ws,u"Tipo de operación"),border(ws,u"Cliente / Proveedor"),border(ws,u"Entrada Cantidad"),border(ws,u"Salida Cantidad"),border(ws,u"Saldo Final Cantidad")]

				ws.append(linea)


			if primero:
				cantidadx =((line[15] if line[15] else 0) - (line[11] if line[11] else 0) +   (line[13] if line[13] else 0) )
				
				total_1 += cantidadx
				total_2 += 0
				total_3 += 0
				total_4 += 0
				total_4 += 0

				ws.append([str(line[1]) if line[1] else '',
				str(line[2]) if line[2] else '',
				str(line[3]) if line[3] else '',
				str(line[4]) if line[4] else '',
				'16',
				'Saldo Inicial',
				'{:,.2f}'.format(decimal.Decimal ("%0.2f" % (cantidadx) )),
				  
					  '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (0) )),
					    
						    '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (cantidadx) )),
							  ])
				primero = False

			ws.append([str(line[1]) if line[1] else '',
			str(line[2]) if line[2] else '',
			str(line[3]) if line[3] else '',
			str(line[4]) if line[4] else '',
			str(line[8]) if line[8] else '',
			str(line[7]) if line[7] else '',
			'{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[11] if line[11] else 0) )),
			  
				  '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[13] if line[13] else 0) )),
				    
					    '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (line[15] if line[15] else 0) )),
						  ])
 

			total_1 += line[11] if line[11] else 0
			total_2 += line[12] if line[12] else 0
			total_3 += line[13] if line[13] else 0
			total_4 += line[14] if line[14] else 0
			total_5 += line[15] if line[15] else 0
			
		if product_index != False:

			ws.append(["","","","","","TOTAL", '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_1) )) , '{:,.2f}'.format(decimal.Decimal ("%0.2f" % (total_3) ))])
			
			total_1 = 0
			total_2 = 0
			total_3 = 0 
			total_4 = 0
			total_5 = 0
			
			ws.append([""])
			ws.append([""])


		workbook.save(output)
		output.seek(0)

		return self.env['popup.it'].get_file('KardexSunat.xlsx',base64.encodestring(output.read()))



	def report_txt(self):
		separador = '|'
		s_prod = [-1,-1,-1]
		s_loca = [-1,-1,-1]
		if self.alllocations == True:
			locat_ids = self.env['stock.location'].search( [('usage','in',('internal','inventory','transit','procurement','production'))] )
			lst_locations = locat_ids.ids
		else:
			lst_locations = self.location_ids.ids
		lst_products  = self.products_ids.ids
		productos='{'
		almacenes='{'
		date_ini=self.fini
		date_fin=self.ffin
		if self.allproducts:
			lst_products = self.env['product.product'].with_context(active_test=False).search([]).ids
		else:
			lst_products = self.products_ids.ids
		if len(lst_products) == 0:
			raise osv.except_osv('Alerta','No existen productos seleccionados')
		for producto in lst_products:
			productos=productos+str(producto)+','
			s_prod.append(producto)
		productos=productos[:-1]+'}'
		for location in lst_locations:
			almacenes=almacenes+str(location)+','
			s_loca.append(location)
		almacenes=almacenes[:-1]+'}'
		t = self.env['make.kardex.valorado'].create({
				'fini':date_ini,
				'ffin':date_fin,
				'fecha_ini_mod':date_ini,
				'fecha_fin_mod':date_fin,
				'allproducts':True,
				'alllocations':True,
				'moneda':'pen',
				})
		t.with_context({'res_model_it':'make.kardex.valorado.formato.sunat','id_it':self.id}).guardado()
		self.env.cr.execute("""
			 select
				fecha_albaran as "Fecha Alb.",	
				fecha as "Fecha",
				type_doc as "T. Doc.",
				serial as "Serie",
				nro as "Nro. Documento",
				numdoc_cuadre as "Nro. Documento",
				doc_partner as "Nro Doc. Partner",
				name as "Proveedor",							
				operation_type as "Tipo de operacion",				 
				name_template as "Producto",
				unidad as "Unidad",			
				ingreso as "Ingreso Fisico",
				round(debit,6) as "Ingreso Valorado.",
				salida as "Salida Fisico",
				round(credit,6) as "Salida Valorada",
				saldof as "Saldo Fisico",
				round(saldov,6) as "Saldo valorado",
				round(cadquiere,6) as "Costo adquisicion",
				round(cprom,6) as "Costo promedio",
					origen as "Origen",
					destino as "Destino",
				almacen AS "Almacen",
				default_code as "Cod Producto",
				null::varchar "guia remision",
				product_id,
				stock_moveid,
				location_id,
				ubicacion_origen
			from get_kardex_v("""+  str(date_ini).replace('-','') +""","""+  str(date_fin).replace('-','') + ",'" + productos + """'::INT[], '""" + almacenes + """'::INT[], """ +str(self.env.company.id)+ """)
		order by almacen, product_id, fecha, esingreso
		""")
		product_index = False
		almacen_index = False
		total_1 = 0
		total_2 = 0
		total_3 = 0 
		total_4 = 0
		txt=""
		contador = {}
		for linea in self.env.cr.fetchall():			
			if str(linea[1])>= str(date_ini):
				move_obj = self.env['stock.move'].browse(linea[25])
				txt+=str(linea[1])[0:4] +	str(linea[1])[5:7] + '00|'
				llavemaster = move_obj.invoice_id.id or linea[25]
				if llavemaster in contador:
					contador[llavemaster] = contador[llavemaster]+1
				else:
					contador[llavemaster] = 1

				almacen_origen = self.env['stock.location'].browse(linea[27])

				txt+=str(llavemaster) + str(contador[llavemaster]) + '|'
				txt+=('A' if almacen_origen.id and almacen_origen.usage == 'inventory' else 'M') + str(move_obj.invoice_id.name or move_obj.id) + '|'

				almacen = self.env['stock.location'].browse(linea[26])
				txt+= str(almacen.l10n_pe_edi_branch_code) + "|" #EL CODIGO DE ESTABLECIMIENTO
				txt+='1|'
				#txt+=str(linea[25]) + str(contador[linea[25]]) + '|'##esta de mas
				productoobj = self.env['product.product'].browse(linea[24])
				txt+= (productoobj.categ_id.existence_type_id.code or '') + '|'
				txt+= (productoobj.default_code or '') + '|'
				txt+= '1|'
				txt+= (productoobj.onu_code.code or '') + '|'
				fechaespecial = str(move_obj.invoice_id.invoice_date if move_obj.invoice_id.id else ((move_obj.kardex_date - timedelta(hours=5) if move_obj.id and move_obj.kardex_date else str(linea[1]))  ) )
				txt+= fechaespecial[8:10] + "/" + fechaespecial[5:7] + "/" + fechaespecial[:4] + '|'
				txt+= (str(move_obj.invoice_id.l10n_latam_document_type_id.code) if move_obj.invoice_id.id else '00' ) + '|'
				txt+= (str(move_obj.invoice_id.ref).split('-')[0] if move_obj.invoice_id.id and move_obj.invoice_id.ref else '0' ) + '|'
				txt+= (str(move_obj.invoice_id.ref).split('-')[1] if move_obj.invoice_id.id and move_obj.invoice_id.ref and len(move_obj.invoice_id.ref.split('-'))>1 else '0' ) + '|'
				txt+= (str(move_obj.picking_id.type_operation_sunat_id.code) if move_obj.picking_id.type_operation_sunat_id.id else '' ) + '|'
				txt+= (productoobj.name_get()[0][1] or '') + '|'
				txt+= (productoobj.uom_id.code_sunat.code or '') + '|'
				txt+= '1|'			
				txt+=  "%.2f"%(linea[11])+ '|' #entrada
				txt+=  "%.2f"%(linea[12]/linea[11] if linea[11]!= 0 else 0)+ '|' #entrada			
				txt+=  "%.2f"%(linea[12])+ '|' #entrada
				txt+=  "%.2f"%(linea[13])+ '|' #salida
				txt+=  "%.2f"%(linea[14]/linea[13] if linea[13]!= 0 else 0)+ '|' #salida
				txt+=  "%.2f"%(linea[14])+ '|' #salida
				txt+=  "%.2f"%(linea[15])+ '|' #saldo
				txt+=  "%.2f"%(linea[16]/linea[15] if linea[15]!= 0 else 0)+ '|' #saldo
				txt+=  "%.2f"%(linea[16])+ '|' #saldo
				txt+=  '1|\n' #saldo


		import importlib
		import sys
		importlib.reload(sys)        
		nombre = "LE" + str(self.env.company.partner_id.vat) + str(self.fini.year) + str(self.fini.month).rjust(2,'0') + '00120100001111.txt'
		return self.env['popup.it'].get_file(nombre,base64.encodestring(b''+txt.encode("utf-8")))





	def report_txt(self):
		separador = '|'
		s_prod = [-1,-1,-1]
		s_loca = [-1,-1,-1]
		if self.alllocations == True:
			locat_ids = self.env['stock.location'].search( [('usage','in',('internal','inventory','transit','procurement','production'))] )
			lst_locations = locat_ids.ids
		else:
			lst_locations = self.location_ids.ids
		lst_products  = self.products_ids_it.ids
		productos='{'
		almacenes='{'
		date_ini=self.fini
		date_fin=self.ffin
		if self.allproducts:
			lst_products = self.env['product.product'].with_context(active_test=False).search([]).ids
		else:
			lst_products = self.products_ids_it.ids
		if len(lst_products) == 0:
			raise osv.except_osv('Alerta','No existen productos seleccionados')
		for producto in lst_products:
			productos=productos+str(producto)+','
			s_prod.append(producto)
		productos=productos[:-1]+'}'
		for location in lst_locations:
			almacenes=almacenes+str(location)+','
			s_loca.append(location)
		almacenes=almacenes[:-1]+'}'
		t = self.env['make.kardex.valorado'].create({
				'fini':date_ini,
				'ffin':date_fin,
				'fecha_ini_mod':date_ini,
				'fecha_fin_mod':date_fin,
				'allproducts':True,
				'alllocations':True,
				'moneda':'pen',
				})
		t.with_context({'res_model_it':'make.kardex.valorado.formato.sunat','id_it':self.id}).guardado()
		self.env.cr.execute("""
			 select
				fecha_albaran as "Fecha Alb.",	
				fecha as "Fecha",
				type_doc as "T. Doc.",
				serial as "Serie",
				nro as "Nro. Documento",
				numdoc_cuadre as "Nro. Documento",
				doc_partner as "Nro Doc. Partner",
				name as "Proveedor",							
				operation_type as "Tipo de operacion",				 
				name_template as "Producto",
				unidad as "Unidad",			
				ingreso as "Ingreso Fisico",
				round(debit,6) as "Ingreso Valorado.",
				salida as "Salida Fisico",
				round(credit,6) as "Salida Valorada",
				saldof as "Saldo Fisico",
				round(saldov,6) as "Saldo valorado",
				round(cadquiere,6) as "Costo adquisicion",
				round(cprom,6) as "Costo promedio",
					origen as "Origen",
					destino as "Destino",
				almacen AS "Almacen",
				default_code as "Cod Producto",
				null::varchar "guia remision",
				product_id,
				stock_moveid,
				location_id,
				ubicacion_origen
			from get_kardex_v("""+  str(date_ini).replace('-','') +""","""+  str(date_fin).replace('-','') + ",'" + productos + """'::INT[], '""" + almacenes + """'::INT[], """ +str(self.env.company.id)+ """)
		order by almacen, product_id, fecha, esingreso
		""")
		product_index = False
		almacen_index = False
		total_1 = 0
		total_2 = 0
		total_3 = 0 
		total_4 = 0
		txt=""
		contador = {}
		for linea in self.env.cr.fetchall():			
			if str(linea[1])>= str(date_ini):
				move_obj = self.env['stock.move'].browse(linea[25])
				txt+=str(linea[1])[0:4] +	str(linea[1])[5:7] + '00|'
				llavemaster = move_obj.invoice_id.id or linea[25]
				if llavemaster in contador:
					contador[llavemaster] = contador[llavemaster]+1
				else:
					contador[llavemaster] = 1

				almacen_origen = self.env['stock.location'].browse(linea[27])

				txt+=str(llavemaster) + str(contador[llavemaster]) + '|'
				txt+=('A' if almacen_origen.id and almacen_origen.usage == 'inventory' else 'M') + str(move_obj.invoice_id.name) + '|'

				almacen = self.env['stock.location'].browse(linea[26])
				txt+= str(almacen.l10n_pe_edi_branch_code) + "|" #EL CODIGO DE ESTABLECIMIENTO
				txt+='1|'
				#txt+=str(linea[25]) + str(contador[linea[25]]) + '|'##esta de mas
				productoobj = self.env['product.product'].browse(linea[24])
				txt+= (productoobj.categ_id.existence_type_id.code or '') + '|'
				txt+= (productoobj.default_code or '') + '|'
				txt+= '1|'
				txt+= (productoobj.onu_code.code or '') + '|'
				fechaespecial = str(move_obj.invoice_id.invoice_date if move_obj.invoice_id.id else ((move_obj.kardex_date - timedelta(hours=5) if move_obj.id and move_obj.kardex_date else str(linea[1]))  ) )
				txt+= fechaespecial[8:10] + "/" + fechaespecial[5:7] + "/" + fechaespecial[:4] + '|'
				txt+= (str(move_obj.invoice_id.l10n_latam_document_type_id.code) if move_obj.invoice_id.id else '00' ) + '|'
				txt+= (str(move_obj.invoice_id.ref).split('-')[0] if move_obj.invoice_id.id and move_obj.invoice_id.ref else '0' ) + '|'
				txt+= (str(move_obj.invoice_id.ref).split('-')[1] if move_obj.invoice_id.id and move_obj.invoice_id.ref and len(move_obj.invoice_id.ref.split('-'))>1 else '0' ) + '|'
				txt+= (str(move_obj.picking_id.type_operation_sunat_id.code) if move_obj.picking_id.type_operation_sunat_id.id else '' ) + '|'
				txt+= (productoobj.name_get()[0][1] or '') + '|'
				txt+= (productoobj.uom_id.code_sunat.code or '') + '|'
				txt+= '1|'			
				txt+=  "%.2f"%(linea[11])+ '|' #entrada
				txt+=  "%.2f"%(linea[12]/linea[11] if linea[11]!= 0 else 0)+ '|' #entrada			
				txt+=  "%.2f"%(linea[12])+ '|' #entrada
				txt+=  "%.2f"%(linea[13])+ '|' #salida
				txt+=  "%.2f"%(linea[14]/linea[13] if linea[13]!= 0 else 0)+ '|' #salida
				txt+=  "%.2f"%(linea[14])+ '|' #salida
				txt+=  "%.2f"%(linea[15])+ '|' #saldo
				txt+=  "%.2f"%(linea[16]/linea[15] if linea[15]!= 0 else 0)+ '|' #saldo
				txt+=  "%.2f"%(linea[16])+ '|' #saldo
				txt+=  '1|\n' #saldo


		import importlib
		import sys
		importlib.reload(sys)        
		nombre = "LE" + str(self.env.company.partner_id.vat) + str(self.fini.year) + str(self.fini.month).rjust(2,'0') + '00120100001111.txt'
		return self.env['popup.it'].get_file(nombre,base64.encodestring(b''+txt.encode("utf-8")))



