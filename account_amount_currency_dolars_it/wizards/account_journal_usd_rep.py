# -*- coding: utf-8 -*-

from odoo import models, fields, api
from datetime import date
from odoo.exceptions import UserError
import base64
from io import BytesIO
import subprocess
import sys

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')

class AccountJournalUsdRep(models.TransientModel):
	_name = 'account.journal.usd.rep'

	name = fields.Char()
	company_id = fields.Many2one('res.company',string=u'Compañia',required=True, default=lambda self: self.env.company,readonly=True)

	def get_fiscal_year(self):
		today = fields.Date.context_today(self)
		fiscal_year = self.env['account.fiscal.year'].search([('name','=',str(today.year))],limit=1)
		return fiscal_year.id if fiscal_year else None

	fiscal_year_id = fields.Many2one('account.fiscal.year',string=u'Año Fiscal',default=lambda self:self.get_fiscal_year())
	period_from_id = fields.Many2one('account.period',string='Periodo Inicial',required=True)
	period_to_id = fields.Many2one('account.period',string='Periodo Final',required=True)
	level = fields.Selection([('balance','Balance'),('register','Registro')],default='register',string='Nivel')
	type_show =  fields.Selection([('pantalla','Pantalla'),('excel','Excel')],string=u'Mostrar en',default='pantalla')
	
	def get_report_diario(self):
		if self.type_show == 'pantalla':
			self._cr.execute("""DROP VIEW IF EXISTS account_journal_usd_book CASCADE;
				CREATE OR REPLACE VIEW account_journal_usd_book AS 
				(SELECT row_number() OVER () AS id, * FROM (%s)T)"""%(self.sql_report_diario()))
			return {
				'name': 'Libro Diario USD',
				'type': 'ir.actions.act_window',
				'res_model': 'account.journal.usd.book',
				'view_mode': 'tree,pivot,graph',
				'view_type': 'form',
			}
		else:
			return self.excel_report_diario()
	
	def sql_report_diario(self):
		sql_not_journal = "'{%s}'"%('0')
		param = self.env['account.main.parameter'].search([('company_id','=',self.env.company.id)],limit=1)
		if param.journal_exchange_exclude:
			sql_not_journal = "'{%s}'" % (','.join(str(i) for i in param.journal_exchange_exclude.ids))
		sql = """SELECT periodo,fecha,libro,voucher,cuenta,debe,haber,balance,debe_me,haber_me,importe_me as balance_me,moneda,tc,cta_analitica,
		glosa,td_partner,doc_partner,partner,td_sunat,nro_comprobante,fecha_doc,fecha_ven,col_reg,monto_reg,medio_pago,ple_diario,ple_compras,ple_ventas	
		FROM get_diariog_usd('%s','%s',%d,%s)"""%(self.period_from_id.code,self.period_to_id.code,self.company_id.id,sql_not_journal)
		return sql
	
	def excel_report_diario(self):
		self.ensure_one()
		self.env.cr.execute(self.sql_report_diario())
		res = self.env.cr.fetchall()
		colnames = [
			desc[0] for desc in self.env.cr.description
		]
		res.insert(0, colnames)

		wb = openpyxl.Workbook()
		ws = wb.active
		row_position = 1
		col_position = 1
		for index, row in enumerate(res, row_position):
			for col, val in enumerate(row, col_position):
				ws.cell(row=index, column=col).value = val
		output = BytesIO()
		wb.save(output)
		output.getvalue()
		output_datas = base64.b64encode(output.getvalue())
		output.close()

		return self.env['popup.it'].get_file('Libro Diario.xlsx',output_datas)
	
	def get_report_mayor(self):
		self.ensure_one()
		sql_not_journal = "'{%s}'"%('0')
		param = self.env['account.main.parameter'].search([('company_id','=',self.env.company.id)],limit=1)
		if param.journal_exchange_exclude:
			sql_not_journal = "'{%s}'" % (','.join(str(i) for i in param.journal_exchange_exclude.ids))
		sql_query = """select 
			vst.periodo,
			vst.fecha,
			vst.libro,
			vst.voucher,
			vst.cuenta,
			vst.debe,
			vst.haber,
			sum(vst.balance) OVER (partition by vst.cuenta order by vst.cuenta,vst.fecha,vst.move_line_id) as saldomn,
			vst.debe_me,
			vst.haber_me,
			sum(vst.importe_me) OVER (partition by vst.cuenta order by vst.cuenta,vst.fecha,vst.move_line_id) as saldome,
			vst.moneda,
			vst.tc,
			vst.cta_analitica,
			replace(regexp_replace(vst.glosa, '[^\w]+',' ','g'),',','') as glosa,
			vst.td_partner,
			vst.doc_partner,
			vst.partner,
			vst.td_sunat,
			vst.nro_comprobante,
			vst.fecha_doc,
			vst.fecha_ven
			from get_diariog_usd('%s','%s',%d,%s) vst 
			order by vst.cuenta,vst.fecha"""%(self.period_from_id.code,self.period_to_id.code,self.company_id.id,sql_not_journal)

		self.env.cr.execute(sql_query)
		sql_query = "COPY (%s) TO STDOUT WITH %s" % (sql_query, "CSV HEADER DELIMITER ';'")

		try:
			output = BytesIO()
			self.env.cr.copy_expert(sql_query, output)
			res = base64.b64encode(output.getvalue())
			output.close()
		finally:
			res = res.decode('utf-8')

		return self.env['popup.it'].get_file('Libro Mayor.csv',res)
	
	def sql_htf1(self):
		sql_not_journal = "'{%s}'"%('0')
		param = self.env['account.main.parameter'].search([('company_id','=',self.env.company.id)],limit=1)
		if param.journal_exchange_exclude:
			sql_not_journal = "'{%s}'" % (','.join(str(i) for i in param.journal_exchange_exclude.ids))
		if self.level == 'register':
			sql = """select 
				mayor,cuenta,nomenclatura,debe,haber,saldo_deudor,saldo_acreedor,activo,pasivo,perdinat,ganannat,perdifun,gananfun,rubro
				from get_f1_register_usd('{period_from}','{period_to}',{company},{sql_not_journal})
				UNION ALL
				SELECT 
				null::text as mayor,
				null::character varying as cuenta,
				'SUMAS'::text as nomenclatura,
				sum(debe) as debe,
				sum(haber) as haber,
				sum(saldo_deudor) as saldo_deudor,
				sum(saldo_acreedor) as saldo_acreedor,
				sum(activo) as activo,
				sum(pasivo) as pasivo,
				sum(perdinat) as perdinat,
				sum(ganannat) as ganannat,
				sum(perdifun) as perdifun,
				sum(gananfun) as gananfun,
				null::text as rubro
				FROM get_f1_register_usd('{period_from}','{period_to}',{company},{sql_not_journal})
				UNION ALL
				SELECT 
				null::text as mayor,
				null::character varying as cuenta,
				'UTILIDAD O PERDIDA'::text as nomenclatura,
				case
					when sum(debe) < sum(haber)
					then sum(haber) - sum(debe)
					else 0
				end as debe,
				case
					when sum(debe) > sum(haber)
					then sum(debe) - sum(haber) 
					else 0
				end as haber,
				case
					when sum(saldo_deudor) < sum(saldo_acreedor)
					then sum(saldo_acreedor) - sum(saldo_deudor)
					else 0
				end as saldo_deudor,
				case
					when sum(saldo_deudor) > sum(saldo_acreedor)
					then sum(saldo_deudor) - sum(saldo_acreedor)
					else 0
				end as saldo_acreedor,
				case
					when sum(activo) < sum(pasivo)
					then sum(pasivo) - sum(activo)
					else 0
				end as activo,
				case
					when sum(activo) > sum(pasivo)
					then sum(activo) - sum(pasivo)
					else 0
				end as pasivo,
				case
					when sum(perdinat) < sum(ganannat)
					then sum(ganannat) - sum(perdinat)
					else 0
				end as perdinat,
				case
					when sum(perdinat) > sum(ganannat)
					then sum(perdinat) - sum(ganannat)
					else 0
				end as ganannat,
				case
					when sum(perdifun) < sum(gananfun)
					then sum(gananfun) - sum(perdifun)
					else 0
				end as perdifun,
				case
					when sum(perdifun) > sum(gananfun)
					then sum(perdifun) - sum(gananfun)
					else 0
				end as gananfun,
				null::text as rubro
				FROM get_f1_register_usd('{period_from}','{period_to}',{company},{sql_not_journal})""".format(
					period_from = self.period_from_id.code,
					period_to = self.period_to_id.code,
					company = self.company_id.id,
					sql_not_journal = sql_not_journal
					)
		else:
			sql = """select 
				mayor,nomenclatura,debe,haber,saldo_deudor,saldo_acreedor,activo,pasivo,perdinat,ganannat,perdifun,gananfun
				from get_f1_balance_usd('{period_from}','{period_to}',{company},{sql_not_journal})
				UNION ALL
				SELECT 
				null::text as mayor,
				'SUMAS'::text as nomenclatura,
				sum(debe) as debe,
				sum(haber) as haber,
				sum(saldo_deudor) as saldo_deudor,
				sum(saldo_acreedor) as saldo_acreedor,
				sum(activo) as activo,
				sum(pasivo) as pasivo,
				sum(perdinat) as perdinat,
				sum(ganannat) as ganannat,
				sum(perdifun) as perdifun,
				sum(gananfun) as gananfun
				FROM get_f1_balance_usd('{period_from}','{period_to}',{company},{sql_not_journal})
				UNION ALL
				SELECT 
				null::text as mayor,
				'UTILIDAD O PERDIDA'::text as nomenclatura,
				case
					when sum(debe) < sum(haber)
					then sum(haber) - sum(debe)
					else 0
				end as debe,
				case
					when sum(debe) > sum(haber)
					then sum(debe) - sum(haber) 
					else 0
				end as haber,
				case
					when sum(saldo_deudor) < sum(saldo_acreedor)
					then sum(saldo_acreedor) - sum(saldo_deudor)
					else 0
				end as saldo_deudor,
				case
					when sum(saldo_deudor) > sum(saldo_acreedor)
					then sum(saldo_deudor) - sum(saldo_acreedor)
					else 0
				end as saldo_acreedor,
				case
					when sum(activo) < sum(pasivo)
					then sum(pasivo) - sum(activo)
					else 0
				end as activo,
				case
					when sum(activo) > sum(pasivo)
					then sum(activo) - sum(pasivo)
					else 0
				end as pasivo,
				case
					when sum(perdinat) < sum(ganannat)
					then sum(ganannat) - sum(perdinat)
					else 0
				end as perdinat,
				case
					when sum(perdinat) > sum(ganannat)
					then sum(perdinat) - sum(ganannat)
					else 0
				end as ganannat,
				case
					when sum(perdifun) < sum(gananfun)
					then sum(gananfun) - sum(perdifun)
					else 0
				end as perdifun,
				case
					when sum(perdifun) > sum(gananfun)
					then sum(perdifun) - sum(gananfun)
					else 0
				end as gananfun
				FROM get_f1_balance_usd('{period_from}','{period_to}',{company},{sql_not_journal})""".format(
					period_from = self.period_from_id.code,
					period_to = self.period_to_id.code,
					company = self.company_id.id,
					sql_not_journal = sql_not_journal
					)
		return sql
	
	def get_report_htf1(self):
		if self.type_show == 'pantalla':
			if self.level == 'register':
				self._cr.execute("""DROP VIEW IF EXISTS f1_register_usd CASCADE;
				CREATE OR REPLACE VIEW f1_register_usd AS 
				(SELECT row_number() OVER () AS id, * FROM (%s)T)"""%(self.sql_htf1()))
				return {
					'name': 'Hoja de Trabajo USD - Registro',
					'type': 'ir.actions.act_window',
					'res_model': 'f1.register.usd',
					'view_mode': 'tree',
					'views': [(self.env.ref('account_amount_currency_dolars_it.view_f1_register_usd_tree').id, 'tree')],
					}
			else:
				self._cr.execute("""DROP VIEW IF EXISTS f1_balance_usd CASCADE;
				CREATE OR REPLACE VIEW f1_balance_usd AS 
				(SELECT row_number() OVER () AS id, * FROM (%s)T)"""%(self.sql_htf1()))
				return {
					'name': 'Hoja de Trabajo USD - Balance',
					'type': 'ir.actions.act_window',
					'res_model': 'f1.balance.usd',
					'view_mode': 'tree',
					'views': [(self.env.ref('account_amount_currency_dolars_it.view_f1_balance_usd_tree').id, 'tree')],
					}
		else:
			return self.get_excel_htf1()
	
	def get_excel_htf1(self):
		self.ensure_one()
		self.env.cr.execute(self.sql_htf1())
		res = self.env.cr.fetchall()
		colnames = [
			desc[0] for desc in self.env.cr.description
		]
		res.insert(0, colnames)

		wb = openpyxl.Workbook()
		ws = wb.active
		row_position = 1
		col_position = 1
		for index, row in enumerate(res, row_position):
			for col, val in enumerate(row, col_position):
				ws.cell(row=index, column=col).value = val
		output = BytesIO()
		wb.save(output)
		output.getvalue()
		output_datas = base64.b64encode(output.getvalue())
		output.close()

		return self.env['popup.it'].get_file('Hoja de Trabajo.xlsx',output_datas)