# -*- coding: utf-8 -*-
from odoo.tools.misc import DEFAULT_SERVER_DATETIME_FORMAT
import time
import odoo.addons.decimal_precision as dp
from openerp.osv import osv
import base64
from odoo import models, fields, api
import codecs
from odoo.exceptions import UserError, ValidationError
from datetime import datetime, timedelta
def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')

from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl import Workbook
values = {}
from openpyxl.utils import get_column_letter
from openpyxl.cell import WriteOnlyCell
values = {}

def border(ws,texto):
	cell = WriteOnlyCell(ws, value=texto)
	cell.font = Font(name='Arial',size=12.5,bold=True)
	cell.border = Border(
	left=Side(border_style=BORDER_THIN, color='00000000'),
	right=Side(border_style=BORDER_THIN, color='00000000'),
	top=Side(border_style=BORDER_THIN, color='00000000'),
	bottom=Side(border_style=BORDER_THIN, color='00000000'))
	return cell
def number_format_saldo(ws,numero):
	cell = WriteOnlyCell(ws, value=numero)
	cell.number_format = "0.000000"
	return cell
def number_format_cantidad(ws,numero):
	cell = WriteOnlyCell(ws, value=numero)
	cell.number_format = "0.00"
	return cell


class kardex_stock_cost_tmp(models.TransientModel):
	_name = 'kardex.stock.cost.tmp'

	almacen = fields.Char('Almacen')
	producto = fields.Char('Producto')
	costo = fields.Float('Costo')

class product_template(models.Model):
	_inherit = 'product.template'

	def get_ultimocosto(self):
		productos = self.env['product.product'].search([('product_tmpl_id','=',self.id)])

		import datetime
		from datetime import timedelta
		
		posible_fecha = self.env["kardex.cerrado.config"].sudo().search([("company_id","=",self.env.company.id)], limit=1, order="fecha_fin desc")
		fecha = False
		if posible_fecha.id:
			fecha = posible_fecha.fecha_fin + timedelta(days=1)
		if not fecha:
			posible_fecha = self.env["stock.move"].sudo().search([("picking_id.company_id","=",self.env.company.id),("state","=",'done')], limit=1, order="kardex_date")
			if posible_fecha.id:
				fecha = (posible_fecha.kardex_date - timedelta(hours=5)).date()
		mes = str(fecha.month if fecha else (datetime.datetime.now() - timedelta(hours=5)).date().month)
		nuevo = self.env['valor.unitario.kardex'].with_context(force_company=self.env.company.id).create({'fecha_inicio': str(fecha) if fecha else str(datetime.datetime.now())[0:4]+ '-'+mes+'-01', 'fecha_final': str(datetime.datetime.now())[0:4]+'-12-31' })
		nuevo.with_context(force_company=self.env.company.id,product_ids =productos.ids).do_valor()

		registro_nuevo = self.env['make.kardex.valorado.stock'].create({
			'allproducts':False,
			'products_ids':productos.ids,
			})
		return registro_nuevo.do_mostrar_pantalla()

class stock_balance_report_lote_costeo(models.Model):
	_name = 'stock.balance.report.lote.costeo'

	producto = fields.Many2one('product.product',string='N.Producto',store=True)
	codigo = fields.Char(related='producto.default_code',string='Cod. Producto',store=True)
	almacen = fields.Many2one('stock.location',string=u'N.Almacén',store=True)
	entrada = fields.Float(string='Stock', digits=(12,2),store=True)
	salida = fields.Float(string='Salida', digits=(12,2),store=True)
	saldo = fields.Float(string='Disponible', digits=(12,2),store=True)
	unidad = fields.Many2one(related='producto.uom_id',string='Unidad',store=True)
	categoria_1 = fields.Char(related='producto.categ_id.name',string='Categoria 1',store=True)
	categoria_2 = fields.Char(related='producto.categ_id.parent_id.name',string='Categoria 2',store=True)
	categoria_3 = fields.Char(related='producto.categ_id.parent_id.parent_id.name',string='Categoria 3',store=True)
	lote = fields.Many2one('stock.production.lot',string='Lote',store=True)
	#fecha_venc = fields.Date(related='lote.vencimiento',string='Fecha Vencimiento',store=True)

	reservado = fields.Float(string='Reservado', digits=(12,2),store=True)
	product_id = fields.Many2one('product.product','Producto',store=True)
	almacen_id = fields.Many2one('stock.location','Almacen',store=True)
	costeo = fields.Float('Costeo', digits=(12,2))

	tiempoutil = fields.Integer(string='Tiempo Vida Util',store=True)
	rango_vencimiento = fields.Char(string='Rango de Vencimiento',store=True)


class kardex_stock_cost_tmp(models.TransientModel):
	_name = 'kardex.stock.cost.tmp'

	almacen = fields.Char('Almacen')
	producto = fields.Char('Producto')
	costo = fields.Float('Costo')

class product_template(models.Model):
	_inherit = 'product.template'

	def get_ultimocosto(self):
		productos = self.env['product.product'].search([('product_tmpl_id','=',self.id)])

		import datetime
		from datetime import timedelta
		
		posible_fecha = self.env["kardex.cerrado.config"].sudo().search([("company_id","=",self.env.company.id)], limit=1, order="fecha_fin desc")
		fecha = False
		if posible_fecha.id:
			fecha = posible_fecha.fecha_fin + timedelta(days=1)
		if not fecha:
			posible_fecha = self.env["stock.move"].sudo().search([("picking_id.company_id","=",self.env.company.id),("state","=",'done')], limit=1, order="kardex_date")
			if posible_fecha.id:
				fecha = (posible_fecha.kardex_date - timedelta(hours=5)).date()
		mes = str(fecha.month if fecha else (datetime.datetime.now() - timedelta(hours=5)).date().month)
		nuevo = self.env['valor.unitario.kardex'].with_context(force_company=self.env.company.id).create({'fecha_inicio': str(fecha) if fecha else str(datetime.datetime.now())[0:4]+ '-'+mes+'-01', 'fecha_final': str(datetime.datetime.now())[0:4]+'-12-31' })
		nuevo.with_context(force_company=self.env.company.id,product_ids =productos.ids).do_valor()

		registro_nuevo = self.env['make.kardex.valorado.stock'].create({
			'allproducts':False,
			'products_ids':productos.ids,
			})
		return registro_nuevo.do_mostrar_pantalla()


class make_kardex_valorado_stock(models.TransientModel):
	_name = "make.kardex.valorado.stock"


	fini= fields.Date('Fecha inicial',required=True)
	ffin= fields.Date('Fecha final',required=True)
	products_ids=fields.Many2many('product.product','rel_wiz_kardex_valorado_stock','product_id','kardex_id','Product')
	location_ids=fields.Many2many('stock.location','rel_kardex_location_valorado_stock','location_id','kardex_id','Ubicacion',required=True)
	allproducts=fields.Boolean('Todos los productos',default=True)
	destino = fields.Selection([('csv','CSV')],'Destino')
	check_fecha = fields.Boolean('Editar Fecha')
	alllocations = fields.Boolean('Todos los almacenes',default=True)

	moneda = fields.Selection([('pen','PEN'),('usd','USD')],'Moneda',default='pen')
	fecha_ini_mod = fields.Date('Fecha Inicial')
	fecha_fin_mod = fields.Date('Fecha Final')
	analizador = fields.Boolean('Analizador')

	@api.onchange('fecha_ini_mod')
	def onchange_fecha_ini_mod(self):
		self.fini = self.fecha_ini_mod


	@api.onchange('fecha_fin_mod')
	def onchange_fecha_fin_mod(self):
		self.ffin = self.fecha_fin_mod


	@api.model
	def default_get(self, fields):
		res = super(make_kardex_valorado_stock, self).default_get(fields)
		import datetime
		fecha_hoy = str(datetime.datetime.now())[:10]
		fecha_inicial = fecha_hoy[:4] + '-01-01'
		res.update({'fecha_ini_mod':fecha_inicial})
		res.update({'fecha_fin_mod':fecha_hoy})
		res.update({'fini':fecha_inicial})
		res.update({'ffin':fecha_hoy})

		#locat_ids = self.pool.get('stock.location').search(cr, uid, [('usage','in',('internal','inventory','transit','procurement','production'))])
		locat_ids = self.env['stock.location'].search([('usage','in',('internal','inventory','transit','procurement','production'))])
		locat_ids = [elemt.id for elemt in locat_ids]
		res.update({'location_ids':[(6,0,locat_ids)]})
		return res

	@api.onchange('alllocations')
	def onchange_alllocations(self):
		if self.alllocations == True:
			locat_ids = self.env['stock.location'].search( [('usage','in',('internal','inventory','transit','procurement','production'))] )
			self.location_ids = [(6,0,locat_ids.ids)]
		else:
			self.location_ids = [(6,0,[])]




	def do_csvtoexcel(self):
		cad = ""
		s_prod = [-1,-1,-1]
		s_loca = [-1,-1,-1]
		if self.alllocations == True:
			locat_ids = self.env['stock.location'].search( [('usage','in',('internal','inventory','transit','procurement','production'))] )
			lst_locations = locat_ids.ids
		else:
			lst_locations = self.location_ids.ids
		lst_products  = self.products_ids.ids
		productos='('
		almacenes='('
		date_ini=self.fini
		date_fin=self.ffin
		if self.allproducts:
			lst_products = self.env['product.product'].with_context(active_test=False).search([]).ids

		else:
			lst_products = self.products_ids.ids

		if len(lst_products) == 0:
			raise osv.except_osv('Alerta','No existen productos seleccionados')

		for producto in lst_products:
			productos=productos+str(producto)+','
			s_prod.append(producto)
		productos=productos[:-1]+')'
		for location in lst_locations:
			almacenes=almacenes+str(location)+','
			s_loca.append(location)
		almacenes=almacenes[:-1]+')'



		import io
		output = io.BytesIO()

		workbook = Workbook(write_only=True)
		ws = workbook.create_sheet("Saldos Valorados")
		x= 9
		tam_col = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
		ws.column_dimensions['A'].width = 10
		ws.column_dimensions['B'].width = 20
		ws.column_dimensions['C'].width = 20
		ws.column_dimensions['D'].width = 20
		ws.column_dimensions['E'].width = 20
		ws.column_dimensions['F'].width = 20
		ws.column_dimensions['G'].width = 20
		ws.column_dimensions['H'].width = 20
		ws.column_dimensions['I'].width = 100

		ws.column_dimensions['J'].width = 20
		ws.column_dimensions['K'].width = 20
		ws.column_dimensions['L'].width = 20
		ws.column_dimensions['M'].width = 20
		ws.column_dimensions['N'].width = 20

		cell = WriteOnlyCell(ws, value="SALDOS VALORADOS")
		cell.font = Font(name='Bahnschrift Light SemiCondensed',size=33,bold=True)
		cell.alignment = Alignment(horizontal='center')


		cell_fini = WriteOnlyCell(ws, value="FECHA INICIO:" + str(self.fini))
		cell_fini.font = Font(name='Arial',size=12.5,bold=True)

		cell_ffin = WriteOnlyCell(ws, value="FECHA FINAL:" + str(self.ffin))
		cell_ffin.font = Font(name='Arial',size=12.5,bold=True)



		ws.merged_cells.ranges.append(get_column_letter(1)+ "1:" + get_column_letter(14) + '1')
		ws.append([cell,"","","","","","","","","","","","","",""])

		ws.append([""])


		ws.append([cell_fini])
		ws.append([cell_ffin])

		ws.append([""])
		ws.append([""])
		ws.append([""])

			
		import datetime		

		linea = [border(ws,u"IDOdoo"),border(ws,u"Categoría de Producto N1"),border(ws,u"Categoría de Producto N2"),border(ws,u"Categoría de Producto N3"),border(ws,u"Categoría de Producto N4"),border(ws,u"Categoría de Producto N5"),border(ws,u"Marca"),border(ws,u"Codigo Producto"),border(ws,u"Producto"),border(ws,u"Unidad"),border(ws,u"Precio De Venta"),border(ws,u"Saldo Cantidad"),border(ws,u"Saldo Soles"),border(ws,u"C/U"),border(ws,u"Almacen")]
		ws.append(linea)


		config = self.env['kardex.parameter'].search([('company_id','=',self.env.company.id)])

		date_ini = '%d-01-01' % ( config._get_anio_start(date_fin.year) )

		kardex_save_obj = False

		if self.moneda == 'pen':
			kardex_save_obj = self.env['kardex.save'].search([('company_id','=',self.env.company.id),('state','=','done'),('name.date_end','<',self.fini)]).sorted(lambda l: l.name.code , reverse=True)
			if len(kardex_save_obj)>0:
				kardex_save_obj = kardex_save_obj[0]
				date_ini = kardex_save_obj.name.date_end + timedelta(days=1)
			else:
				kardex_save_obj = False
		
		text_report_linea = ''
		if kardex_save_obj:
			text_report_linea = "<b>Se usara el saldo guardado: "+str(kardex_save_obj.name.code)+ " </b>"
			

		si_existe = ""
		if kardex_save_obj:
			si_existe = """union all


select 
				ksp.producto as product_id,
				ksp.almacen as location_id,
				'' as origen_usage,
				sl.usage as destino_usage,
				ksp.cprom * ksp.stock as debit,
				0 as credit,
				(fecha || ' 00:00:00')::timestamp as fechax,
				'' as type_doc,
				'' as serial,
				'' as nro,
				'' as numdoc_cuadre,
				'' as nro_documento,
				'Saldo Inicial' as name,
				'' as operation_type,
				pname.new_name,
				coalesce(pp.default_code,pt.default_code) as default_code,
				uu.name as unidad,
				ksp.stock as ingreso,
				0 as salida,
				ksp.cprom as cadquiere,
				'' as origen,
				sl.complete_name as destino,
				sl.complete_name as almacen

			from kardex_save_period ksp 
			inner join stock_location sl on sl.id = ksp.almacen
			inner join product_product pp on pp.id = ksp.producto
			inner join product_template pt on pt.id = pp.product_tmpl_id
			inner join uom_uom uu on uu.id = pt.uom_id
			inner join product_category pc on pc.id = pt.categ_id
			left join stock_production_lot spt on spt.id = ksp.lote
			 LEFT JOIN ( SELECT t_pp.id,
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
					 LEFT JOIN product_variant_combination pvc ON pvc.product_product_id = t_pp.id
					 LEFT JOIN product_template_attribute_value ptav ON ptav.id = pvc.product_template_attribute_value_id
					 LEFT JOIN product_attribute_value pav ON pav.id = ptav.product_attribute_value_id
				  GROUP BY t_pp.id) pname ON pname.id = pp.id
			 where save_id = """+str(kardex_save_obj.id)+""" 
			   and sl.id in """+str(almacenes)+""" and pp.id in """ +str(productos)+ """
				
			 """

		text_report = "<b>Cargando Saldos Valorados</b><br/><center>Ejecutando SQL del kardex ... Espere por favor</center><br/>" + text_report_linea
		self.send_message(text_report)

		total_all = []


			
		if self.moneda == 'pen':

			self.env.cr.execute("""  
					select  
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
			 from (
					select 
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
					from (
		select vst_kardex_sunat.*,sp.name as doc_almac,sm.kardex_date as fecha_albaran,
		vst_kardex_sunat.fecha - interval '5' hour as fechax,sl_o.usage as origen_usage , sl_d.usage as destino_usage, np.new_name
					from vst_kardex_fisico_valorado as vst_kardex_sunat
		left join stock_move sm on sm.id = vst_kardex_sunat.stock_moveid
		left join stock_picking sp on sp.id = sm.picking_id

							inner join stock_location sl_o on sl_o.id = sm.location_id
							inner join stock_location sl_d on sl_d.id = sm.location_dest_id
		left join (
			select t_pp.id, 
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
				left join product_variant_combination pvc on pvc.product_product_id = t_pp.id
				left join product_template_attribute_value ptav on ptav.id = pvc.product_template_attribute_value_id
				left join product_attribute_value pav on pav.id = ptav.product_attribute_value_id
				group by t_pp.id
				) np on np.id = vst_kardex_sunat.product_id
							
			   where (fecha_num((vst_kardex_sunat.fecha - interval '5' hour)::date) between """+str(date_ini).replace('-','')+""" and """+str(date_fin).replace('-','')+""")    
			   and vst_kardex_sunat.location_id in """+str(almacenes)+""" and vst_kardex_sunat.product_id in """ +str(productos)+ """
					 and vst_kardex_sunat.company_id = """ +str(self.env.company.id)+ """
					order by vst_kardex_sunat.location_id,vst_kardex_sunat.product_id,vst_kardex_sunat.fecha,vst_kardex_sunat.esingreso,vst_kardex_sunat.stock_moveid,vst_kardex_sunat.nro
					)Total   """+si_existe+"""			
	) A order by location_id,product_id,fechax
					
				""")
			total_all = self.env.cr.fetchall()
		else:
			self.env.cr.execute("""  
					select  
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
			 from (
					select 
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
					from (
		select vst_kardex_sunat.*,sp.name as doc_almac,sm.kardex_date as fecha_albaran,
		vst_kardex_sunat.fecha - interval '5' hour as fechax,sl_o.usage as origen_usage , sl_d.usage as destino_usage, np.new_name
					from vst_kardex_fisico_valorado_dolar as vst_kardex_sunat
		left join stock_move sm on sm.id = vst_kardex_sunat.stock_moveid
	left join purchase_order_line pol on pol.id = sm.purchase_line_id
	left join purchase_order po on po.id = pol.order_id
		left join stock_picking sp on sp.id = sm.picking_id

							inner join stock_location sl_o on sl_o.id = sm.location_id
							inner join stock_location sl_d on sl_d.id = sm.location_dest_id
		left join (
			select t_pp.id, 
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
				left join product_variant_combination pvc on pvc.product_product_id = t_pp.id
				left join product_template_attribute_value ptav on ptav.id = pvc.product_template_attribute_value_id
				left join product_attribute_value pav on pav.id = ptav.product_attribute_value_id
				group by t_pp.id
				) np on np.id = vst_kardex_sunat.product_id
							
			   where (fecha_num((vst_kardex_sunat.fecha - interval '5' hour)::date) between """+str(date_ini).replace('-','')+""" and """+str(date_fin).replace('-','')+""")    
			   and vst_kardex_sunat.location_id in """+str(almacenes)+""" and vst_kardex_sunat.product_id in """ +str(productos)+ """
					 and vst_kardex_sunat.company_id = """ +str(self.env.company.id)+ """
					order by vst_kardex_sunat.location_id,vst_kardex_sunat.product_id,vst_kardex_sunat.fecha,vst_kardex_sunat.esingreso,vst_kardex_sunat.stock_moveid,vst_kardex_sunat.nro
					)Total   """+si_existe+"""			
	) A order by location_id,product_id,fechax
					
				""")
			total_all = self.env.cr.fetchall()


		self.send_message(text_report_linea+"Saldos Valorados<br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center>")
		cont_report = 0
		import datetime
		tiempo_inicial = datetime.datetime.now()
		tiempo_pasado = [0,0]
		cprom_data = {}

		ingreso1 =0
		ingreso2 =0
		salida1 =0
		salida2 =0
		producto = 0
		almacen = 0
		array_grabar = None
		flag = False
		for xl in total_all:
			l = {
				'product_id':xl[0],
				'location_id':xl[1],
				'origen_usage':xl[2],
				'destino_usage':xl[3],
				'debit':xl[4],
				'credit':xl[5],
				'fechax':xl[6],
				'type_doc':xl[7],
				'serial':xl[8],
				'nro':xl[9],
				'numdoc_cuadre':xl[10],
				'nro_documento':xl[11],
				'name':xl[12],
				'operation_type':xl[13],
				'new_name':xl[14],
				'default_code':xl[15],
				'unidad':xl[16],
				'ingreso':xl[17],
				'salida':xl[18],
				'cadquiere':xl[19],
				'origen':xl[20],
				'destino':xl[21],
				'almacen':xl[22],
			}

			if producto == None:
				producto = l['product_id']
				almacen = l['almacen']
			if producto != 	l['product_id'] or almacen != l['almacen']:
				producto = l['product_id']
				almacen = l['almacen']	
				if array_grabar and array_grabar[11].value != 0:			
					ws.append(array_grabar)


			cont_report += 1
			if cont_report%300 == 0:
				tiempo_pasado = divmod((datetime.datetime.now()-tiempo_inicial).seconds,60)
				text_report = ""
				text_report = text_report_linea+ "<b>Saldos Valorados</b><br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center><br/><center>Procesado:"+str(cont_report)+"/"+str(len(total_all))+"</center><br/> Tiempo Procesado: "+ str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos" 
				text_report += """<div id="myProgress" style="position: relative; width: 100%;  height: 30px;   background-color: white;">
				  <div id="myBar" style="position: absolute;  width: """+"%.2f"%(cont_report*100/len(total_all))+"""%;  height: 100%; background-color: #875A7B;">
					<div id="label" style="text-align: center; line-height: 30px; color: white;">""" +"%.2f"%(cont_report*100/len(total_all))+ """%</div>
				  </div>
				</div>"""
				self.send_message(text_report)
				
			llave = (l['product_id'],l['location_id'])
			cprom_acum = [0,0]
			if llave in cprom_data:
				cprom_acum = cprom_data[llave]
			else:
				cprom_data[llave] = cprom_acum

			cprom_act_antes = cprom_data[llave][1] / cprom_data[llave][0] if cprom_data[llave][0] != 0 else 0

			data_temp = {}
			
			data_temp = {'origen':l['origen_usage'] or '','destino':l['destino_usage'] or ''}
			ingreso_v = 0
			egreso_v = 0
			if l['ingreso'] or l['debit']:
				if (data_temp['origen'] == 'internal' and data_temp['destino'] == 'internal') or (data_temp['origen'] == 'transit' and data_temp['destino'] == 'internal'):
					cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] + (l['debit'] if l['debit'] else 0) -  (l['credit'] if l['credit'] else 0)

					ingreso_v = (l['debit'] if l['debit'] else 0) 
				else:	
					cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] + (l['debit'] if l['debit'] else 0) -  (l['credit'] if l['credit'] else 0)

					ingreso_v = (l['debit'] if l['debit'] else 0) 
			else:
				if (data_temp['origen'] == 'internal' and data_temp['destino'] == 'supplier'):
					cprom_acum[0] = cprom_acum[0] -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] - (l['debit'] if l['debit'] else 0) - ( (l['credit'] if l['credit'] else 0) * (l['salida'] if l['salida'] else 0) )
					egreso_v = (l['credit'] if l['credit'] else 0) * (l['salida'] if l['salida'] else 0)
				else:
					if l['salida']:
						cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
						cprom_acum[1] = cprom_acum[1] - (l['salida'] if l['salida'] else 0)*(cprom_act_antes if cprom_act_antes else 0)

						egreso_v = (l['salida'] if l['salida'] else 0)*(cprom_act_antes if cprom_act_antes else 0)
					else:
						cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
						cprom_acum[1] = cprom_acum[1] - (l['credit'] if l['credit'] else 0)

						egreso_v = (l['credit'] if l['credit'] else 0)

			cprom_acum[1] = cprom_acum[1] if cprom_acum[0] > 0 else 0
			cprom_act = cprom_acum[1] / cprom_acum[0] if cprom_acum[0] != 0 else 0

			cprom_data[llave] = cprom_acum

			linea = []
			producto_obj = self.env['product.product'].browse(l['product_id'])
			cat_name = []
			cat_obj = producto_obj.categ_id
			while (cat_obj.id):
				cat_name.append(cat_obj.name)
				cat_obj = cat_obj.parent_id
			cat_name.reverse()
			cat_recortado = cat_name[2:]
			n1= ""
			n2= ""
			n3= ""
			n4= ""
			n5= ""
			n1= cat_recortado[0] if len(cat_recortado)> 0 else ''
			n2= cat_recortado[1] if len(cat_recortado)> 1 else ''
			n3= cat_recortado[2] if len(cat_recortado)> 2 else ''
			n4= cat_recortado[3] if len(cat_recortado)> 3 else ''
			n5= cat_recortado[4] if len(cat_recortado)> 4 else ''
	
			linea.append( producto_obj.id or '')
			linea.append( n1)
			linea.append( n2)
			linea.append( n3)
			linea.append( n4)
			linea.append( n5)
			linea.append( producto_obj.product_brand_id.name_get()[0][1] if producto_obj.product_brand_id.id else '')
			linea.append( l['default_code'] if l['default_code'] else '')
			linea.append( l['new_name'] if l['new_name'] else '' )
			linea.append( l['unidad'] if l['unidad'] else '' )
			linea.append (number_format_cantidad(ws, producto_obj.lst_price if producto_obj.lst_price else 0))

			linea.append( number_format_cantidad(ws, cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0 ))
			linea.append( number_format_saldo(ws, cprom_acum[1] if len(cprom_acum)>1 and cprom_acum[1] else 0 ))
			linea.append( number_format_saldo(ws, (cprom_acum[1] if len(cprom_acum)>1 and cprom_acum[1] else 0) / (cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0) if (cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0) != 0 else 0 ))

			linea.append( l['almacen'] if l['almacen'] else '' )
			array_grabar = linea.copy()
			
			ingreso1 += l['ingreso'] or 0
			ingreso2 += ingreso_v or 0
			salida1 += l['salida'] or 0
			salida2 += egreso_v or 0


		tam_col = [11,11,5,5,7,5,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11]
		if array_grabar and array_grabar[11].value != 0:
			ws.append(array_grabar)

		workbook.save(output)
		output.seek(0)

		attach_id = self.env['ir.attachment'].create({
					'name': "Kardex Valorado.xlsx",
					'type': 'binary',
					'datas': base64.encodebytes(output.getvalue()),
					'eliminar_automatico': True
				})
		output.close()



		return {
			'type': 'ir.actions.client',
			'tag': 'notification_llikha',
			'params': {
				'title':'Kardex Valorado Excel' if self.moneda == 'pen' else 'Kardex Valorado Excel (USD)',
				'type': 'success',
				'sticky': True,
				'message': 'Se proceso de '+str(self.fini)+' al ' + str(self.ffin)+ '.<br/>Lineas procesadas: '+ str(len(total_all)) +'<br/>Tiempo: ' + str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos",
				'next': {'type': 'ir.actions.act_window_close'},
				'buttons':[{
					'label':'Descargar Kardex Valorado',
					'model':'ir.attachment',
					'method':'get_download_ls',
					'id':attach_id.id,
					}
				],
			}
		}



	def do_csvtoexcellotecosteado(self):
		cad = ""
		s_prod = [-1,-1,-1]
		s_loca = [-1,-1,-1]
		if self.alllocations == True:
			locat_ids = self.env['stock.location'].search( [('usage','in',('internal','inventory','transit','procurement','production'))] )
			lst_locations = locat_ids.ids
		else:
			lst_locations = self.location_ids.ids
		lst_products  = self.products_ids.ids
		productos='('
		almacenes='('
		date_ini=self.fini
		date_fin=self.ffin
		if self.allproducts:
			lst_products = self.env['product.product'].with_context(active_test=False).search([]).ids

		else:
			lst_products = self.products_ids.ids

		if len(lst_products) == 0:
			raise osv.except_osv('Alerta','No existen productos seleccionados')

		for producto in lst_products:
			productos=productos+str(producto)+','
			s_prod.append(producto)
		productos=productos[:-1]+')'
		for location in lst_locations:
			almacenes=almacenes+str(location)+','
			s_loca.append(location)
		almacenes=almacenes[:-1]+')'
		x= 9
			
		import datetime		

		config = self.env['kardex.parameter'].search([('company_id','=',self.env.company.id)])

		date_ini = '%d-01-01' % ( config._get_anio_start(date_fin.year) )

		kardex_save_obj = False

		if self.moneda == 'pen':
			kardex_save_obj = self.env['kardex.save'].search([('company_id','=',self.env.company.id),('state','=','done'),('name.date_end','<',self.fini)]).sorted(lambda l: l.name.code , reverse=True)
			if len(kardex_save_obj)>0:
				kardex_save_obj = kardex_save_obj[0]
				date_ini = kardex_save_obj.name.date_end + timedelta(days=1)
			else:
				kardex_save_obj = False
		
		text_report_linea = ''
		if kardex_save_obj:
			text_report_linea = "<b>Se usara el saldo guardado: "+str(kardex_save_obj.name.code)+ " </b>"
			

		si_existe = ""
		if kardex_save_obj:
			si_existe = """union all


select 
				ksp.producto as product_id,
				ksp.almacen as location_id,
				'' as origen_usage,
				sl.usage as destino_usage,
				ksp.cprom * ksp.stock as debit,
				0 as credit,
				(fecha || ' 00:00:00')::timestamp as fechax,
				'' as type_doc,
				'' as serial,
				'' as nro,
				'' as numdoc_cuadre,
				'' as nro_documento,
				'Saldo Inicial' as name,
				'' as operation_type,
				pname.new_name,
				coalesce(pp.default_code,pt.default_code) as default_code,
				uu.name as unidad,
				ksp.stock as ingreso,
				0 as salida,
				ksp.cprom as cadquiere,
				'' as origen,
				sl.complete_name as destino,
				sl.complete_name as almacen

			from kardex_save_period ksp 
			inner join stock_location sl on sl.id = ksp.almacen
			inner join product_product pp on pp.id = ksp.producto
			inner join product_template pt on pt.id = pp.product_tmpl_id
			inner join uom_uom uu on uu.id = pt.uom_id
			inner join product_category pc on pc.id = pt.categ_id
			left join stock_production_lot spt on spt.id = ksp.lote
			 LEFT JOIN ( SELECT t_pp.id,
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
					 LEFT JOIN product_variant_combination pvc ON pvc.product_product_id = t_pp.id
					 LEFT JOIN product_template_attribute_value ptav ON ptav.id = pvc.product_template_attribute_value_id
					 LEFT JOIN product_attribute_value pav ON pav.id = ptav.product_attribute_value_id
				  GROUP BY t_pp.id) pname ON pname.id = pp.id
			 where save_id = """+str(kardex_save_obj.id)+""" 
			   and sl.id in """+str(almacenes)+""" and pp.id in """ +str(productos)+ """
				
			 """

		text_report = "<b>Cargando Saldos Valorados</b><br/><center>Ejecutando SQL del kardex ... Espere por favor</center><br/>" + text_report_linea
		self.send_message(text_report)

		total_all = []


			
		if self.moneda == 'pen':

			self.env.cr.execute("""  
					select  
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
			 from (
					select 
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
					from (
		select vst_kardex_sunat.*,sp.name as doc_almac,sm.kardex_date as fecha_albaran,
		vst_kardex_sunat.fecha - interval '5' hour as fechax,sl_o.usage as origen_usage , sl_d.usage as destino_usage, np.new_name
					from vst_kardex_fisico_valorado as vst_kardex_sunat
		left join stock_move sm on sm.id = vst_kardex_sunat.stock_moveid
		left join stock_picking sp on sp.id = sm.picking_id

							inner join stock_location sl_o on sl_o.id = sm.location_id
							inner join stock_location sl_d on sl_d.id = sm.location_dest_id
		left join (
			select t_pp.id, 
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
				left join product_variant_combination pvc on pvc.product_product_id = t_pp.id
				left join product_template_attribute_value ptav on ptav.id = pvc.product_template_attribute_value_id
				left join product_attribute_value pav on pav.id = ptav.product_attribute_value_id
				group by t_pp.id
				) np on np.id = vst_kardex_sunat.product_id
							
			   where (fecha_num((vst_kardex_sunat.fecha - interval '5' hour)::date) between """+str(date_ini).replace('-','')+""" and """+str(date_fin).replace('-','')+""")    
			   and vst_kardex_sunat.location_id in """+str(almacenes)+""" and vst_kardex_sunat.product_id in """ +str(productos)+ """
					 and vst_kardex_sunat.company_id = """ +str(self.env.company.id)+ """
					order by vst_kardex_sunat.location_id,vst_kardex_sunat.product_id,vst_kardex_sunat.fecha,vst_kardex_sunat.esingreso,vst_kardex_sunat.stock_moveid,vst_kardex_sunat.nro
					)Total   """+si_existe+"""			
	) A order by location_id,product_id,fechax
					
				""")
			total_all = self.env.cr.fetchall()
		else:
			self.env.cr.execute("""  
					select  
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
			 from (
					select 
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
					from (
		select vst_kardex_sunat.*,sp.name as doc_almac,sm.kardex_date as fecha_albaran,
		vst_kardex_sunat.fecha - interval '5' hour as fechax,sl_o.usage as origen_usage , sl_d.usage as destino_usage, np.new_name
					from vst_kardex_fisico_valorado_dolar as vst_kardex_sunat
		left join stock_move sm on sm.id = vst_kardex_sunat.stock_moveid
	left join purchase_order_line pol on pol.id = sm.purchase_line_id
	left join purchase_order po on po.id = pol.order_id
		left join stock_picking sp on sp.id = sm.picking_id

							inner join stock_location sl_o on sl_o.id = sm.location_id
							inner join stock_location sl_d on sl_d.id = sm.location_dest_id
		left join (
			select t_pp.id, 
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
				left join product_variant_combination pvc on pvc.product_product_id = t_pp.id
				left join product_template_attribute_value ptav on ptav.id = pvc.product_template_attribute_value_id
				left join product_attribute_value pav on pav.id = ptav.product_attribute_value_id
				group by t_pp.id
				) np on np.id = vst_kardex_sunat.product_id
							
			   where (fecha_num((vst_kardex_sunat.fecha - interval '5' hour)::date) between """+str(date_ini).replace('-','')+""" and """+str(date_fin).replace('-','')+""")    
			   and vst_kardex_sunat.location_id in """+str(almacenes)+""" and vst_kardex_sunat.product_id in """ +str(productos)+ """
					 and vst_kardex_sunat.company_id = """ +str(self.env.company.id)+ """
					order by vst_kardex_sunat.location_id,vst_kardex_sunat.product_id,vst_kardex_sunat.fecha,vst_kardex_sunat.esingreso,vst_kardex_sunat.stock_moveid,vst_kardex_sunat.nro
					)Total   """+si_existe+"""			
	) A order by location_id,product_id,fechax
					
				""")
			total_all = self.env.cr.fetchall()


		self.send_message(text_report_linea+"Saldos Valorados<br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center>")
		cont_report = 0
		import datetime
		tiempo_inicial = datetime.datetime.now()
		tiempo_pasado = [0,0]
		cprom_data = {}

		ingreso1 =0
		ingreso2 =0
		salida1 =0
		salida2 =0
		producto = 0
		almacen = 0
		array_grabar = None
		flag = False
		costeosops = {}
		for xl in total_all:
			l = {
				'product_id':xl[0],
				'location_id':xl[1],
				'origen_usage':xl[2],
				'destino_usage':xl[3],
				'debit':xl[4],
				'credit':xl[5],
				'fechax':xl[6],
				'type_doc':xl[7],
				'serial':xl[8],
				'nro':xl[9],
				'numdoc_cuadre':xl[10],
				'nro_documento':xl[11],
				'name':xl[12],
				'operation_type':xl[13],
				'new_name':xl[14],
				'default_code':xl[15],
				'unidad':xl[16],
				'ingreso':xl[17],
				'salida':xl[18],
				'cadquiere':xl[19],
				'origen':xl[20],
				'destino':xl[21],
				'almacen':xl[22],
			}

			if producto == None:
				producto = l['product_id']
				almacen = l['almacen']
			if producto != 	l['product_id'] or almacen != l['almacen']:
				producto = l['product_id']
				almacen = l['almacen']	
				if array_grabar and array_grabar[11] != 0:			
					llave = (array_grabar[0],array_grabar[14])
					if llave in costeosops:
						pass
					else:
						costeosops[llave] = array_grabar[13]


			cont_report += 1
			if cont_report%300 == 0:
				tiempo_pasado = divmod((datetime.datetime.now()-tiempo_inicial).seconds,60)
				text_report = ""
				text_report = text_report_linea+ "<b>Saldos Valorados</b><br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center><br/><center>Procesado:"+str(cont_report)+"/"+str(len(total_all))+"</center><br/> Tiempo Procesado: "+ str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos" 
				text_report += """<div id="myProgress" style="position: relative; width: 100%;  height: 30px;   background-color: white;">
				  <div id="myBar" style="position: absolute;  width: """+"%.2f"%(cont_report*100/len(total_all))+"""%;  height: 100%; background-color: #875A7B;">
					<div id="label" style="text-align: center; line-height: 30px; color: white;">""" +"%.2f"%(cont_report*100/len(total_all))+ """%</div>
				  </div>
				</div>"""
				self.send_message(text_report)
				
			llave = (l['product_id'],l['location_id'])
			cprom_acum = [0,0]
			if llave in cprom_data:
				cprom_acum = cprom_data[llave]
			else:
				cprom_data[llave] = cprom_acum

			cprom_act_antes = cprom_data[llave][1] / cprom_data[llave][0] if cprom_data[llave][0] != 0 else 0

			data_temp = {}
			
			data_temp = {'origen':l['origen_usage'] or '','destino':l['destino_usage'] or ''}
			ingreso_v = 0
			egreso_v = 0
			if l['ingreso'] or l['debit']:
				if (data_temp['origen'] == 'internal' and data_temp['destino'] == 'internal') or (data_temp['origen'] == 'transit' and data_temp['destino'] == 'internal'):
					cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] + (l['debit'] if l['debit'] else 0) -  (l['credit'] if l['credit'] else 0)

					ingreso_v = (l['debit'] if l['debit'] else 0) 
				else:	
					cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] + (l['debit'] if l['debit'] else 0) -  (l['credit'] if l['credit'] else 0)

					ingreso_v = (l['debit'] if l['debit'] else 0) 
			else:
				if (data_temp['origen'] == 'internal' and data_temp['destino'] == 'supplier'):
					cprom_acum[0] = cprom_acum[0] -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] - (l['debit'] if l['debit'] else 0) - ( (l['credit'] if l['credit'] else 0) * (l['salida'] if l['salida'] else 0) )
					egreso_v = (l['credit'] if l['credit'] else 0) * (l['salida'] if l['salida'] else 0)
				else:
					if l['salida']:
						cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
						cprom_acum[1] = cprom_acum[1] - (l['salida'] if l['salida'] else 0)*(cprom_act_antes if cprom_act_antes else 0)

						egreso_v = (l['salida'] if l['salida'] else 0)*(cprom_act_antes if cprom_act_antes else 0)
					else:
						cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
						cprom_acum[1] = cprom_acum[1] - (l['credit'] if l['credit'] else 0)

						egreso_v = (l['credit'] if l['credit'] else 0)

			cprom_acum[1] = cprom_acum[1] if cprom_acum[0] > 0 else 0
			cprom_act = cprom_acum[1] / cprom_acum[0] if cprom_acum[0] != 0 else 0

			cprom_data[llave] = cprom_acum

			linea = []
			producto_obj = self.env['product.product'].browse(l['product_id'])
			cat_name = []
			cat_obj = producto_obj.categ_id
			while (cat_obj.id):
				cat_name.append(cat_obj.name)
				cat_obj = cat_obj.parent_id
			cat_name.reverse()
			cat_recortado = cat_name[2:]
			n1= ""
			n2= ""
			n3= ""
			n4= ""
			n5= ""
			n1= cat_recortado[0] if len(cat_recortado)> 0 else ''
			n2= cat_recortado[1] if len(cat_recortado)> 1 else ''
			n3= cat_recortado[2] if len(cat_recortado)> 2 else ''
			n4= cat_recortado[3] if len(cat_recortado)> 3 else ''
			n5= cat_recortado[4] if len(cat_recortado)> 4 else ''
	
			linea.append( producto_obj.id or '')
			linea.append( n1)
			linea.append( n2)
			linea.append( n3)
			linea.append( n4)
			linea.append( n5)
			linea.append( producto_obj.product_brand_id.name_get()[0][1] if producto_obj.product_brand_id.id else '')
			linea.append( l['default_code'] if l['default_code'] else '')
			linea.append( l['new_name'] if l['new_name'] else '' )
			linea.append( l['unidad'] if l['unidad'] else '' )
			linea.append ( producto_obj.lst_price if producto_obj.lst_price else 0)

			linea.append(  cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0 )
			linea.append(  cprom_acum[1] if len(cprom_acum)>1 and cprom_acum[1] else 0 )
			linea.append(  (cprom_acum[1] if len(cprom_acum)>1 and cprom_acum[1] else 0) / (cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0) if (cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0) != 0 else 0 )

			linea.append( l['almacen'] if l['almacen'] else '' )
			array_grabar = linea.copy()
			
			ingreso1 += l['ingreso'] or 0
			ingreso2 += ingreso_v or 0
			salida1 += l['salida'] or 0
			salida2 += egreso_v or 0

		if array_grabar and array_grabar[11] != 0:			
			llave = (array_grabar[0],array_grabar[14])
			if llave in costeosops:
				pass
			else:
				costeosops[llave] = array_grabar[13]
			



		self.env['stock.balance.report.lote.costeo'].search([]).unlink()

		si_existe = ""
		if kardex_save_obj:
			si_existe = """ select ksp.almacen as alm_id, ksp.producto as p_id,pt.categ_id as categoria_id, '"""+str(kardex_save_obj.name.date_end)+"""'::date as fecha,
			'SALDO INICIAL' as origen, sl.complete_name as destino, sl.complete_name as almacen, ksp.stock as entrada, 0 as salida, null as stock_move,
			'' as motivo_guia, pname.new_name as producto, 'done' as estado, 'Saldo Inicial' as name, coalesce(pp.default_code,pt.default_code) as cod_pro,
			pc.name as categoria, uu.name as unidad, ksp.producto as product_id, ksp.almacen as almacen_id, spt.name as lote, spt.id as lote_id
			from kardex_save_period ksp 
			inner join stock_location sl on sl.id = ksp.almacen
			inner join product_product pp on pp.id = ksp.producto
			inner join product_template pt on pt.id = pp.product_tmpl_id
			inner join uom_uom uu on uu.id = pt.uom_id
			inner join product_category pc on pc.id = pt.categ_id
			left join stock_production_lot spt on spt.id = ksp.lote
			 LEFT JOIN ( SELECT t_pp.id,
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
					 LEFT JOIN product_variant_combination pvc ON pvc.product_product_id = t_pp.id
					 LEFT JOIN product_template_attribute_value ptav ON ptav.id = pvc.product_template_attribute_value_id
					 LEFT JOIN product_attribute_value pav ON pav.id = ptav.product_attribute_value_id
				  GROUP BY t_pp.id) pname ON pname.id = pp.id
			 where save_id = """+str(kardex_save_obj.id)+"""
			union all """

			#for alm in kardex_save_obj.lineas:
			#	self.env.cr.execute(""" select sum(product_uom_qty) 
			#		from stock_move_line 
			#		inner join stock_production_lot  on stock_production_lot.id = stock_move_line.lot_id 
			#		where stock_production_lot.id = """ +str(alm.lote.id)+ """ and 
			#		stock_move_line.location_id = """ +str(alm.almacen.id)+ """
			#		and stock_move_line.product_id = """ +str(alm.producto.id)+ """ and 
			#		stock_move_line.state in ('partially_available','assigned') """)
			#	cont1 = 0
			#	for ex in self.env.cr.fetchall():
			#		cont1 = ex[0]
			#	data[(alm.almacen.id,alm.producto.id,alm.lote.id)] = [alm.almacen.id,alm.producto.id,alm.stock,alm.lote.id,cont1]


		self.env.cr.execute("""
			select 
			vstf.p_id,
			vstf.alm_id,
			coalesce(sum(coalesce(vstf.entrada,0)),0) -  coalesce(sum(coalesce(vstf.salida,0)),0),
			0 as salida,
			coalesce(sum(coalesce(vstf.entrada,0)),0) -  coalesce(sum(coalesce(vstf.salida,0)),0) - coalesce(max(reservado.reservado),0),
			vstf.lote_id,
			coalesce(max(reservado.reservado),0),
			vstf.p_id,
			vstf.alm_id
			from
			( """ +si_existe+ """			
			select location_dest_id as alm_id, product_id as p_id, categoria_id, vst_kardex_fisico.date as fecha,u_origen as origen, u_destino as destino, u_destino as almacen, vst_kardex_fisico.product_qty as entrada, 0 as salida,vst_kardex_fisico.id  as stock_move,vst_kardex_fisico.guia as motivo_guia, producto,vst_kardex_fisico.estado,vst_kardex_fisico.name, cod_pro, categoria, unidad,product_id,location_dest_id as almacen_id, lote,lote_id from vst_kardex_fisico_lote() as vst_kardex_fisico where company_id = """+str(self.env.company.id)+"""
			and (date - interval '5' hour)::date >='""" +str(date_ini)+ """' and (date - interval '5' hour)::date <='""" +str(date_fin)+ """'
			union all
			select location_id as alm_id, product_id as p_id, categoria_id, vst_kardex_fisico.date as fecha,u_origen as origen, u_destino as destino, u_origen as almacen, 0 as entrada, vst_kardex_fisico.product_qty as salida,vst_kardex_fisico.id  as stock_move ,vst_kardex_fisico.guia as motivo_guia ,producto ,vst_kardex_fisico.estado,vst_kardex_fisico.name, cod_pro, categoria, unidad,product_id, location_id as almacen_id, lote, lote_id from vst_kardex_fisico_lote() as vst_kardex_fisico where company_id = """+str(self.env.company.id)+"""
			and (date - interval '5' hour)::date >='""" +str(date_ini)+ """' and (date - interval '5' hour)::date <='""" +str(date_fin)+ """'
			) as vstf
			left join stock_production_lot spl on spl.id = vstf.lote_id
			left join (
				select stock_move_line.location_id,sum(stock_move_line.product_uom_qty) as reservado, stock_move_line.product_id,stock_move_line.lot_id  from stock_move_line left join stock_production_lot  on stock_production_lot.id = stock_move_line.lot_id where stock_move_line.state in ('partially_available','assigned') 
				group by stock_move_line.location_id, stock_move_line.product_id,stock_move_line.lot_id
			) reservado on coalesce(reservado.lot_id,-1) = coalesce(vstf.lote_id,-1) and reservado.product_id = vstf.p_id and reservado.location_id = vstf.alm_id
			where 
			vstf.product_id in """ +str(tuple(s_prod))+ """
			and vstf.almacen_id in """ +str(tuple(s_loca))+ """
			and vstf.estado = 'done'
			group by
			producto,cod_pro,categoria_id, p_id, alm_id,lote,lote_id
			having coalesce(sum(coalesce(vstf.entrada,0)),0) -  coalesce(sum(coalesce(vstf.salida,0)),0) != 0
			;
		""")

		todos = self.env.cr.fetchall()

		text_report_line = text_report + u"---Se procesaran: "+ str(len(todos)) + " lineas---"
		self.send_message(text_report_line)
		cont_report = 0

		for item in todos:
			costoguardado = 0
			llave = (item[7],item[1])
			if llave in costeosops:
				costoguardado = costeosops[llave]

			tiempoutil = 0
			rango_vencimiento = ""
			if item[5]:
				loteobj = self.env['stock.production.lot'].browse(item[5])
				if loteobj.id and loteobj.expiration_date:
					tiempoutil = (datetime.datetime.now()- loteobj.expiration_date).days
					if tiempoutil <= 90:
						rango_vencimiento = '0 a 3 meses'
					elif tiempoutil >90 and tiempoutil <= 180:
						rango_vencimiento = '3 a 6 meses'
					elif tiempoutil > 180 and tiempoutil <= 360:
						rango_vencimiento = '6 a 12 meses'
					else:
						rango_vencimiento = '1 año a más'

			data = {
				'producto':item[0],
				'almacen':item[1],
				'entrada':item[2],
				'salida':item[3],
				'saldo':item[4],
				'lote':item[5],
				'reservado':item[6],
				'product_id':item[7],
				'almacen_id':item[8],
				'costeo':costoguardado,
				'tiempoutil': tiempoutil,
				'rango_vencimiento': rango_vencimiento,
			}
			self.env['stock.balance.report.lote.costeo'].create(data)

			if cont_report%300 == 0:
				tiempo_pasado = divmod((datetime.datetime.now()-tiempo_inicial).seconds,60)
				
				text_report_line = "<b>Generando Saldos por Lotes.</b><br/><center>Total lineas a procesar: "+str(len(todos)) + "</center><br/><center>Procesado:"+str(cont_report)+"/"+str(len(todos))+"</center><br/> Tiempo Procesado: "+ str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos" 
				text_report_line += """<div id="myProgress" style="position: relative; width: 100%;  height: 30px;   background-color: white;">
				  <div id="myBar" style="position: absolute;  width: """+"%.2f"%(cont_report*100/len(todos))+"""%;  height: 100%; background-color: #875A7B;">
				    <div id="label" style="text-align: center; line-height: 30px; color: white;">""" +"%.2f"%(cont_report*100/len(todos))+ """%</div>
				  </div>
				</div>"""
				self.send_message(text_report)

			cont_report += 1
		
		return {
			'name': 'Reporte de Saldos x Lote',
			'type': 'ir.actions.act_window',
			'res_model': 'stock.balance.report.lote.costeo',
			'view_mode': 'tree,pivot,graph',
			'views': [(False, 'tree'), (False, 'pivot'), (False, 'graph')]
		}



	def do_mostrar_pantalla(self):
		cad = ""
		s_prod = [-1,-1,-1]
		s_loca = [-1,-1,-1]
		if self.alllocations == True:
			locat_ids = self.env['stock.location'].search( [('usage','in',('internal','inventory','transit','procurement','production'))] )
			lst_locations = locat_ids.ids
		else:
			lst_locations = self.location_ids.ids
		lst_products  = self.products_ids.ids
		productos='('
		almacenes='('
		date_ini=self.fini
		date_fin=self.ffin
		if self.allproducts:
			lst_products = self.env['product.product'].with_context(active_test=False).search([]).ids

		else:
			lst_products = self.products_ids.ids

		if len(lst_products) == 0:
			raise osv.except_osv('Alerta','No existen productos seleccionados')

		for producto in lst_products:
			productos=productos+str(producto)+','
			s_prod.append(producto)
		productos=productos[:-1]+')'
		for location in lst_locations:
			almacenes=almacenes+str(location)+','
			s_loca.append(location)
		almacenes=almacenes[:-1]+')'

			
		import datetime		


		config = self.env['kardex.parameter'].search([('company_id','=',self.env.company.id)])

		date_ini = '%d-01-01' % ( config._get_anio_start(date_fin.year) )

		kardex_save_obj = False

		if self.moneda == 'pen':
			kardex_save_obj = self.env['kardex.save'].search([('company_id','=',self.env.company.id),('state','=','done'),('name.date_end','<',self.fini)]).sorted(lambda l: l.name.code , reverse=True)
			if len(kardex_save_obj)>0:
				kardex_save_obj = kardex_save_obj[0]
				date_ini = kardex_save_obj.name.date_end + timedelta(days=1)
			else:
				kardex_save_obj = False
		
		text_report_linea = ''
		if kardex_save_obj:
			text_report_linea = "<b>Se usara el saldo guardado: "+str(kardex_save_obj.name.code)+ " </b>"
			

		si_existe = ""
		if kardex_save_obj:
			si_existe = """union all


select 
				ksp.producto as product_id,
				ksp.almacen as location_id,
				'' as origen_usage,
				sl.usage as destino_usage,
				ksp.cprom * ksp.stock as debit,
				0 as credit,
				(fecha || ' 00:00:00')::timestamp as fechax,
				'' as type_doc,
				'' as serial,
				'' as nro,
				'' as numdoc_cuadre,
				'' as nro_documento,
				'Saldo Inicial' as name,
				'' as operation_type,
				pname.new_name,
				coalesce(pp.default_code,pt.default_code) as default_code,
				uu.name as unidad,
				ksp.stock as ingreso,
				0 as salida,
				ksp.cprom as cadquiere,
				'' as origen,
				sl.complete_name as destino,
				sl.complete_name as almacen

			from kardex_save_period ksp 
			inner join stock_location sl on sl.id = ksp.almacen
			inner join product_product pp on pp.id = ksp.producto
			inner join product_template pt on pt.id = pp.product_tmpl_id
			inner join uom_uom uu on uu.id = pt.uom_id
			inner join product_category pc on pc.id = pt.categ_id
			left join stock_production_lot spt on spt.id = ksp.lote
			 LEFT JOIN ( SELECT t_pp.id,
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
					 LEFT JOIN product_variant_combination pvc ON pvc.product_product_id = t_pp.id
					 LEFT JOIN product_template_attribute_value ptav ON ptav.id = pvc.product_template_attribute_value_id
					 LEFT JOIN product_attribute_value pav ON pav.id = ptav.product_attribute_value_id
				  GROUP BY t_pp.id) pname ON pname.id = pp.id
			 where save_id = """+str(kardex_save_obj.id)+""" 
			   and sl.id in """+str(almacenes)+""" and pp.id in """ +str(productos)+ """
				
			 """

		text_report = "<b>Cargando Saldos Valorados</b><br/><center>Ejecutando SQL del kardex ... Espere por favor</center><br/>" + text_report_linea
		self.send_message(text_report)

		total_all = []


			
		if self.moneda == 'pen':

			self.env.cr.execute("""  
					select  
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
			 from (
					select 
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
					from (
		select vst_kardex_sunat.*,sp.name as doc_almac,sm.kardex_date as fecha_albaran,
		vst_kardex_sunat.fecha - interval '5' hour as fechax,sl_o.usage as origen_usage , sl_d.usage as destino_usage, np.new_name
					from vst_kardex_fisico_valorado as vst_kardex_sunat
		left join stock_move sm on sm.id = vst_kardex_sunat.stock_moveid
		left join stock_picking sp on sp.id = sm.picking_id

							inner join stock_location sl_o on sl_o.id = sm.location_id
							inner join stock_location sl_d on sl_d.id = sm.location_dest_id
		left join (
			select t_pp.id, 
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
				left join product_variant_combination pvc on pvc.product_product_id = t_pp.id
				left join product_template_attribute_value ptav on ptav.id = pvc.product_template_attribute_value_id
				left join product_attribute_value pav on pav.id = ptav.product_attribute_value_id
				group by t_pp.id
				) np on np.id = vst_kardex_sunat.product_id
							
			   where (fecha_num((vst_kardex_sunat.fecha - interval '5' hour)::date) between """+str(date_ini).replace('-','')+""" and """+str(date_fin).replace('-','')+""")    
			   and vst_kardex_sunat.location_id in """+str(almacenes)+""" and vst_kardex_sunat.product_id in """ +str(productos)+ """
					 and vst_kardex_sunat.company_id = """ +str(self.env.company.id)+ """
					order by vst_kardex_sunat.location_id,vst_kardex_sunat.product_id,vst_kardex_sunat.fecha,vst_kardex_sunat.esingreso,vst_kardex_sunat.stock_moveid,vst_kardex_sunat.nro
					)Total   """+si_existe+"""			
	) A order by location_id,product_id,fechax
					
				""")
			total_all = self.env.cr.fetchall()
		else:
			self.env.cr.execute("""  
					select  
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
			 from (
					select 
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
					from (
		select vst_kardex_sunat.*,sp.name as doc_almac,sm.kardex_date as fecha_albaran,
		vst_kardex_sunat.fecha - interval '5' hour as fechax,sl_o.usage as origen_usage , sl_d.usage as destino_usage, np.new_name
					from vst_kardex_fisico_valorado_dolar as vst_kardex_sunat
		left join stock_move sm on sm.id = vst_kardex_sunat.stock_moveid
	left join purchase_order_line pol on pol.id = sm.purchase_line_id
	left join purchase_order po on po.id = pol.order_id
		left join stock_picking sp on sp.id = sm.picking_id

							inner join stock_location sl_o on sl_o.id = sm.location_id
							inner join stock_location sl_d on sl_d.id = sm.location_dest_id
		left join (
			select t_pp.id, 
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
				left join product_variant_combination pvc on pvc.product_product_id = t_pp.id
				left join product_template_attribute_value ptav on ptav.id = pvc.product_template_attribute_value_id
				left join product_attribute_value pav on pav.id = ptav.product_attribute_value_id
				group by t_pp.id
				) np on np.id = vst_kardex_sunat.product_id
							
			   where (fecha_num((vst_kardex_sunat.fecha - interval '5' hour)::date) between """+str(date_ini).replace('-','')+""" and """+str(date_fin).replace('-','')+""")    
			   and vst_kardex_sunat.location_id in """+str(almacenes)+""" and vst_kardex_sunat.product_id in """ +str(productos)+ """
					 and vst_kardex_sunat.company_id = """ +str(self.env.company.id)+ """
					order by vst_kardex_sunat.location_id,vst_kardex_sunat.product_id,vst_kardex_sunat.fecha,vst_kardex_sunat.esingreso,vst_kardex_sunat.stock_moveid,vst_kardex_sunat.nro
					)Total   """+si_existe+"""			
	) A order by location_id,product_id,fechax
					
				""")
			total_all = self.env.cr.fetchall()


		self.send_message(text_report_linea+"Saldos Valorados<br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center>")
		cont_report = 0
		import datetime
		tiempo_inicial = datetime.datetime.now()
		tiempo_pasado = [0,0]
		cprom_data = {}

		ingreso1 =0
		ingreso2 =0
		salida1 =0
		salida2 =0
		producto = 0
		almacen = 0
		array_grabar = None
		flag = False
		datafinal = []
		for xl in total_all:
			l = {
				'product_id':xl[0],
				'location_id':xl[1],
				'origen_usage':xl[2],
				'destino_usage':xl[3],
				'debit':xl[4],
				'credit':xl[5],
				'fechax':xl[6],
				'type_doc':xl[7],
				'serial':xl[8],
				'nro':xl[9],
				'numdoc_cuadre':xl[10],
				'nro_documento':xl[11],
				'name':xl[12],
				'operation_type':xl[13],
				'new_name':xl[14],
				'default_code':xl[15],
				'unidad':xl[16],
				'ingreso':xl[17],
				'salida':xl[18],
				'cadquiere':xl[19],
				'origen':xl[20],
				'destino':xl[21],
				'almacen':xl[22],
			}

			if producto == None:
				producto = l['product_id']
				almacen = l['almacen']
			if producto != 	l['product_id'] or almacen != l['almacen']:
				producto = l['product_id']
				almacen = l['almacen']	
				if array_grabar and array_grabar[3] != 0:			
					datafinal.append(array_grabar)


			cont_report += 1
			if cont_report%300 == 0:
				tiempo_pasado = divmod((datetime.datetime.now()-tiempo_inicial).seconds,60)
				text_report = ""
				text_report = text_report_linea+ "<b>Saldos Valorados</b><br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center><br/><center>Procesado:"+str(cont_report)+"/"+str(len(total_all))+"</center><br/> Tiempo Procesado: "+ str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos" 
				text_report += """<div id="myProgress" style="position: relative; width: 100%;  height: 30px;   background-color: white;">
				  <div id="myBar" style="position: absolute;  width: """+"%.2f"%(cont_report*100/len(total_all))+"""%;  height: 100%; background-color: #875A7B;">
					<div id="label" style="text-align: center; line-height: 30px; color: white;">""" +"%.2f"%(cont_report*100/len(total_all))+ """%</div>
				  </div>
				</div>"""
				self.send_message(text_report)
				
			llave = (l['product_id'],l['location_id'])
			cprom_acum = [0,0]
			if llave in cprom_data:
				cprom_acum = cprom_data[llave]
			else:
				cprom_data[llave] = cprom_acum

			cprom_act_antes = cprom_data[llave][1] / cprom_data[llave][0] if cprom_data[llave][0] != 0 else 0

			data_temp = {}
			
			data_temp = {'origen':l['origen_usage'] or '','destino':l['destino_usage'] or ''}
			ingreso_v = 0
			egreso_v = 0
			if l['ingreso'] or l['debit']:
				if (data_temp['origen'] == 'internal' and data_temp['destino'] == 'internal') or (data_temp['origen'] == 'transit' and data_temp['destino'] == 'internal'):
					cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] + (l['debit'] if l['debit'] else 0) -  (l['credit'] if l['credit'] else 0)

					ingreso_v = (l['debit'] if l['debit'] else 0) 
				else:	
					cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] + (l['debit'] if l['debit'] else 0) -  (l['credit'] if l['credit'] else 0)

					ingreso_v = (l['debit'] if l['debit'] else 0) 
			else:
				if (data_temp['origen'] == 'internal' and data_temp['destino'] == 'supplier'):
					cprom_acum[0] = cprom_acum[0] -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] - (l['debit'] if l['debit'] else 0) - ( (l['credit'] if l['credit'] else 0) * (l['salida'] if l['salida'] else 0) )
					egreso_v = (l['credit'] if l['credit'] else 0) * (l['salida'] if l['salida'] else 0)
				else:
					if l['salida']:
						cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
						cprom_acum[1] = cprom_acum[1] - (l['salida'] if l['salida'] else 0)*(cprom_act_antes if cprom_act_antes else 0)

						egreso_v = (l['salida'] if l['salida'] else 0)*(cprom_act_antes if cprom_act_antes else 0)
					else:
						cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
						cprom_acum[1] = cprom_acum[1] - (l['credit'] if l['credit'] else 0)

						egreso_v = (l['credit'] if l['credit'] else 0)

			cprom_acum[1] = cprom_acum[1] if cprom_acum[0] > 0 else 0
			cprom_act = cprom_acum[1] / cprom_acum[0] if cprom_acum[0] != 0 else 0

			cprom_data[llave] = cprom_acum

			linea = []
			producto_obj = self.env['product.product'].browse(l['product_id'])

			linea.append( producto_obj.id or '')
			linea.append( producto_obj.name_get()[0][1] if producto_obj.id else '')
			linea.append( l['almacen'] if l['almacen'] else '' )
			
			linea.append( (cprom_acum[1] if len(cprom_acum)>1 and cprom_acum[1] else 0) / (cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0) if (cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0) != 0 else 0 )
			array_grabar = linea.copy()
			
			ingreso1 += l['ingreso'] or 0
			ingreso2 += ingreso_v or 0
			salida1 += l['salida'] or 0
			salida2 += egreso_v or 0


		tam_col = [11,11,5,5,7,5,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11]
		elemtos_mostrar = []
		if array_grabar and array_grabar[3] != 0:
			datafinal.append(array_grabar)

		texto = ""
		product_txt = ""
		for ix in datafinal:
			texto += ix[2] + " (" + "%.2f"%(ix[3]) + " S/.). \n\r"
			product_txt = ix[1]
			tmp_mostra = self.env['kardex.stock.cost.tmp'].create({
				'almacen':ix[2],
				'producto':ix[1],
				'costo':ix[3],
				})
			elemtos_mostrar.append(tmp_mostra.id)

		return {
			'view_mode': 'tree',
			'view_id': False,
			'res_model': 'kardex.stock.cost.tmp',
			'type': 'ir.actions.act_window',
			'domain': [('id','in',elemtos_mostrar)],
			'target':'new',
		}

		notification = {
		   'type': 'ir.actions.client',
		   'tag': 'display_notification',
		   'params': {
		       'title': "Costos " + product_txt  ,
		       'type': 'success',
		       'message': 'Los costos son: \n\r' +  texto ,
		       'sticky': True,
		   }
		}
		return notification



	def update_costo(self):
		cad = ""
		s_prod = [-1,-1,-1]
		s_loca = [-1,-1,-1]
		if self.alllocations == True:
			locat_ids = self.env['stock.location'].search( [('usage','in',('internal','inventory','transit','procurement','production'))] )
			lst_locations = locat_ids.ids
		else:
			lst_locations = self.location_ids.ids
		lst_products  = self.products_ids.ids
		productos='('
		almacenes='('
		date_ini=self.fini
		date_fin=self.ffin
		if self.allproducts:
			lst_products = self.env['product.product'].with_context(active_test=False).search([]).ids

		else:
			lst_products = self.products_ids.ids

		if len(lst_products) == 0:
			raise osv.except_osv('Alerta','No existen productos seleccionados')

		for producto in lst_products:
			productos=productos+str(producto)+','
			s_prod.append(producto)
		productos=productos[:-1]+')'
		for location in lst_locations:
			almacenes=almacenes+str(location)+','
			s_loca.append(location)
		almacenes=almacenes[:-1]+')'



		import io
		output = io.BytesIO()

		workbook = Workbook(write_only=True)
		ws = workbook.create_sheet("Saldos Valorados")
		x= 9
		tam_col = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
		ws.column_dimensions['A'].width = 10
		ws.column_dimensions['B'].width = 20
		ws.column_dimensions['C'].width = 20
		ws.column_dimensions['D'].width = 20
		ws.column_dimensions['E'].width = 20
		ws.column_dimensions['F'].width = 20
		ws.column_dimensions['G'].width = 20
		ws.column_dimensions['H'].width = 20
		ws.column_dimensions['I'].width = 100

		ws.column_dimensions['J'].width = 20
		ws.column_dimensions['K'].width = 20
		ws.column_dimensions['L'].width = 20
		ws.column_dimensions['M'].width = 20
		ws.column_dimensions['N'].width = 20

		cell = WriteOnlyCell(ws, value="SALDOS VALORADOS")
		cell.font = Font(name='Bahnschrift Light SemiCondensed',size=33,bold=True)
		cell.alignment = Alignment(horizontal='center')


		cell_fini = WriteOnlyCell(ws, value="FECHA INICIO:" + str(self.fini))
		cell_fini.font = Font(name='Arial',size=12.5,bold=True)

		cell_ffin = WriteOnlyCell(ws, value="FECHA FINAL:" + str(self.ffin))
		cell_ffin.font = Font(name='Arial',size=12.5,bold=True)



		ws.merged_cells.ranges.append(get_column_letter(1)+ "1:" + get_column_letter(14) + '1')
		ws.append([cell,"","","","","","","","","","","","","",""])

		ws.append([""])


		ws.append([cell_fini])
		ws.append([cell_ffin])

		ws.append([""])
		ws.append([""])
		ws.append([""])

			
		import datetime		

		linea = [border(ws,u"IDOdoo"),border(ws,u"Categoría de Producto N1"),border(ws,u"Categoría de Producto N2"),border(ws,u"Categoría de Producto N3"),border(ws,u"Categoría de Producto N4"),border(ws,u"Categoría de Producto N5"),border(ws,u"Marca"),border(ws,u"Codigo Producto"),border(ws,u"Producto"),border(ws,u"Unidad"),border(ws,u"Precio De Venta"),border(ws,u"Saldo Cantidad"),border(ws,u"Saldo Soles"),border(ws,u"C/U"),border(ws,u"Almacen")]
		ws.append(linea)


		config = self.env['kardex.parameter'].search([('company_id','=',self.env.company.id)])

		date_ini = '%d-01-01' % ( config._get_anio_start(date_fin.year) )

		kardex_save_obj = False

		if self.moneda == 'pen':
			kardex_save_obj = self.env['kardex.save'].search([('company_id','=',self.env.company.id),('state','=','done'),('name.date_end','<',self.fini)]).sorted(lambda l: l.name.code , reverse=True)
			if len(kardex_save_obj)>0:
				kardex_save_obj = kardex_save_obj[0]
				date_ini = kardex_save_obj.name.date_end + timedelta(days=1)
			else:
				kardex_save_obj = False
		
		text_report_linea = ''
		if kardex_save_obj:
			text_report_linea = "<b>Se usara el saldo guardado: "+str(kardex_save_obj.name.code)+ " </b>"
			

		si_existe = ""
		if kardex_save_obj:
			si_existe = """union all


select 
				ksp.producto as product_id,
				ksp.almacen as location_id,
				'' as origen_usage,
				sl.usage as destino_usage,
				ksp.cprom * ksp.stock as debit,
				0 as credit,
				(fecha || ' 00:00:00')::timestamp as fechax,
				'' as type_doc,
				'' as serial,
				'' as nro,
				'' as numdoc_cuadre,
				'' as nro_documento,
				'Saldo Inicial' as name,
				'' as operation_type,
				pname.new_name,
				coalesce(pp.default_code,pt.default_code) as default_code,
				uu.name as unidad,
				ksp.stock as ingreso,
				0 as salida,
				ksp.cprom as cadquiere,
				'' as origen,
				sl.complete_name as destino,
				sl.complete_name as almacen

			from kardex_save_period ksp 
			inner join stock_location sl on sl.id = ksp.almacen
			inner join product_product pp on pp.id = ksp.producto
			inner join product_template pt on pt.id = pp.product_tmpl_id
			inner join uom_uom uu on uu.id = pt.uom_id
			inner join product_category pc on pc.id = pt.categ_id
			left join stock_production_lot spt on spt.id = ksp.lote
			 LEFT JOIN ( SELECT t_pp.id,
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
					 LEFT JOIN product_variant_combination pvc ON pvc.product_product_id = t_pp.id
					 LEFT JOIN product_template_attribute_value ptav ON ptav.id = pvc.product_template_attribute_value_id
					 LEFT JOIN product_attribute_value pav ON pav.id = ptav.product_attribute_value_id
				  GROUP BY t_pp.id) pname ON pname.id = pp.id
			 where save_id = """+str(kardex_save_obj.id)+"""
			 """

		text_report = "<b>Cargando Saldos Valorados</b><br/><center>Ejecutando SQL del kardex ... Espere por favor</center><br/>" + text_report_linea
		self.send_message(text_report)

		total_all = []


			
		if self.moneda == 'pen':

			self.env.cr.execute("""  
					select  
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
			 from (
					select 
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
					from (
		select vst_kardex_sunat.*,sp.name as doc_almac,sm.kardex_date as fecha_albaran,
		vst_kardex_sunat.fecha - interval '5' hour as fechax,sl_o.usage as origen_usage , sl_d.usage as destino_usage, np.new_name
					from vst_kardex_fisico_valorado as vst_kardex_sunat
		left join stock_move sm on sm.id = vst_kardex_sunat.stock_moveid
		left join stock_picking sp on sp.id = sm.picking_id

							inner join stock_location sl_o on sl_o.id = sm.location_id
							inner join stock_location sl_d on sl_d.id = sm.location_dest_id
		left join (
			select t_pp.id, 
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
				left join product_variant_combination pvc on pvc.product_product_id = t_pp.id
				left join product_template_attribute_value ptav on ptav.id = pvc.product_template_attribute_value_id
				left join product_attribute_value pav on pav.id = ptav.product_attribute_value_id
				group by t_pp.id
				) np on np.id = vst_kardex_sunat.product_id
							
			   where (fecha_num((vst_kardex_sunat.fecha - interval '5' hour)::date) between """+str(date_ini).replace('-','')+""" and """+str(date_fin).replace('-','')+""")    
			   and vst_kardex_sunat.location_id in """+str(almacenes)+""" and vst_kardex_sunat.product_id in """ +str(productos)+ """
					 and vst_kardex_sunat.company_id = """ +str(self.env.company.id)+ """
					order by vst_kardex_sunat.location_id,vst_kardex_sunat.product_id,vst_kardex_sunat.fecha,vst_kardex_sunat.esingreso,vst_kardex_sunat.stock_moveid,vst_kardex_sunat.nro
					)Total   """+si_existe+"""			
	) A order by location_id,product_id,fechax
					
				""")
			total_all = self.env.cr.fetchall()
		else:
			self.env.cr.execute("""  
					select  
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
			 from (
					select 
					product_id,
					location_id,
					origen_usage,
					destino_usage,
					debit,
					credit,
					fechax,
					type_doc,
					serial,
					nro,
					numdoc_cuadre,
					nro_documento,
					name,
					operation_type,
					new_name,
					default_code,
					unidad,
					ingreso,
					salida,
					cadquiere,
					origen,
					destino,
					almacen
					from (
		select vst_kardex_sunat.*,sp.name as doc_almac,sm.kardex_date as fecha_albaran,
		vst_kardex_sunat.fecha - interval '5' hour as fechax,sl_o.usage as origen_usage , sl_d.usage as destino_usage, np.new_name
					from vst_kardex_fisico_valorado_dolar as vst_kardex_sunat
		left join stock_move sm on sm.id = vst_kardex_sunat.stock_moveid
	left join purchase_order_line pol on pol.id = sm.purchase_line_id
	left join purchase_order po on po.id = pol.order_id
		left join stock_picking sp on sp.id = sm.picking_id

							inner join stock_location sl_o on sl_o.id = sm.location_id
							inner join stock_location sl_d on sl_d.id = sm.location_dest_id
		left join (
			select t_pp.id, 
					((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
				   FROM product_product t_pp
					 JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
					 left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
				left join product_variant_combination pvc on pvc.product_product_id = t_pp.id
				left join product_template_attribute_value ptav on ptav.id = pvc.product_template_attribute_value_id
				left join product_attribute_value pav on pav.id = ptav.product_attribute_value_id
				group by t_pp.id
				) np on np.id = vst_kardex_sunat.product_id
							
			   where (fecha_num((vst_kardex_sunat.fecha - interval '5' hour)::date) between """+str(date_ini).replace('-','')+""" and """+str(date_fin).replace('-','')+""")    
			   and vst_kardex_sunat.location_id in """+str(almacenes)+""" and vst_kardex_sunat.product_id in """ +str(productos)+ """
					 and vst_kardex_sunat.company_id = """ +str(self.env.company.id)+ """
					order by vst_kardex_sunat.location_id,vst_kardex_sunat.product_id,vst_kardex_sunat.fecha,vst_kardex_sunat.esingreso,vst_kardex_sunat.stock_moveid,vst_kardex_sunat.nro
					)Total   """+si_existe+"""			
	) A order by location_id,product_id,fechax
					
				""")
			total_all = self.env.cr.fetchall()


		self.send_message(text_report_linea+"Actualizando Costos Valorados<br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center>")
		cont_report = 0
		import datetime
		tiempo_inicial = datetime.datetime.now()
		tiempo_pasado = [0,0]
		cprom_data = {}

		ingreso1 =0
		ingreso2 =0
		salida1 =0
		salida2 =0
		producto = 0
		almacen = 0
		array_grabar = None
		flag = False
		actualizar_montos = []
		for xl in total_all:
			l = {
				'product_id':xl[0],
				'location_id':xl[1],
				'origen_usage':xl[2],
				'destino_usage':xl[3],
				'debit':xl[4],
				'credit':xl[5],
				'fechax':xl[6],
				'type_doc':xl[7],
				'serial':xl[8],
				'nro':xl[9],
				'numdoc_cuadre':xl[10],
				'nro_documento':xl[11],
				'name':xl[12],
				'operation_type':xl[13],
				'new_name':xl[14],
				'default_code':xl[15],
				'unidad':xl[16],
				'ingreso':xl[17],
				'salida':xl[18],
				'cadquiere':xl[19],
				'origen':xl[20],
				'destino':xl[21],
				'almacen':xl[22],
			}

			if producto == None:
				producto = l['product_id']
				almacen = l['almacen']
			if producto != 	l['product_id'] or almacen != l['almacen']:
				producto = l['product_id']
				almacen = l['almacen']	
				if array_grabar and array_grabar[11].value != 0:			
					producto_obj = self.env['product.product'].browse(array_grabar[16])
					costopromedio = array_grabar[15]
					#producto_obj.standard_price =  costopromedio if costopromedio >producto_obj.standard_price else producto_obj.standard_price
					#producto_obj.standard_price =  costopromedio
					costo_actual = 0
					cantidad_actual = 0
					for i in self.env['stock.valuation.layer'].search([('product_id','=',producto_obj.id)]):
						costo_actual += i.value
						cantidad_actual += i.quantity	
 					
					#producto_obj.standard_price = costo_actual/cantidad_actual if cantidad_actual != 0 else 0

					if producto_obj.id in actualizar_montos:
						if costopromedio > producto_obj.standard_price:
							
							self.env.cr.execute("""
							update ir_property set value_float = """ + str(costo_actual/cantidad_actual if cantidad_actual != 0 else 0) + """
							where name = 'standard_price' and company_id = """ + str(self.env.company.id)+ """ and res_id = 'product.product,"""+str(producto_obj.id)+"""' 
							""")
							producto_obj.standard_price = costopromedio
							
					else:
						self.env.cr.execute("""
							update ir_property set value_float = """ + str(costo_actual/cantidad_actual if cantidad_actual != 0 else 0) + """
							where name = 'standard_price' and company_id = """ + str(self.env.company.id)+ """ and res_id = 'product.product,"""+str(producto_obj.id)+"""' 
							""")
						actualizar_montos.append(producto_obj.id)						
						producto_obj.standard_price = costopromedio

			
					ws.append(array_grabar)


			cont_report += 1
			if cont_report%300 == 0:
				tiempo_pasado = divmod((datetime.datetime.now()-tiempo_inicial).seconds,60)
				text_report = ""
				text_report = text_report_linea+ "<b>Saldos Valorados</b><br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center><br/><center>Procesado:"+str(cont_report)+"/"+str(len(total_all))+"</center><br/> Tiempo Procesado: "+ str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos" 
				text_report += """<div id="myProgress" style="position: relative; width: 100%;  height: 30px;   background-color: white;">
				  <div id="myBar" style="position: absolute;  width: """+"%.2f"%(cont_report*100/len(total_all))+"""%;  height: 100%; background-color: #875A7B;">
					<div id="label" style="text-align: center; line-height: 30px; color: white;">""" +"%.2f"%(cont_report*100/len(total_all))+ """%</div>
				  </div>
				</div>"""
				self.send_message(text_report)
				
			llave = (l['product_id'],l['location_id'])
			cprom_acum = [0,0]
			if llave in cprom_data:
				cprom_acum = cprom_data[llave]
			else:
				cprom_data[llave] = cprom_acum

			cprom_act_antes = cprom_data[llave][1] / cprom_data[llave][0] if cprom_data[llave][0] != 0 else 0

			data_temp = {}
			
			data_temp = {'origen':l['origen_usage'] or '','destino':l['destino_usage'] or ''}
			ingreso_v = 0
			egreso_v = 0
			if l['ingreso'] or l['debit']:
				if (data_temp['origen'] == 'internal' and data_temp['destino'] == 'internal') or (data_temp['origen'] == 'transit' and data_temp['destino'] == 'internal'):
					cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] + (l['debit'] if l['debit'] else 0) -  (l['credit'] if l['credit'] else 0)

					ingreso_v = (l['debit'] if l['debit'] else 0) 
				else:	
					cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] + (l['debit'] if l['debit'] else 0) -  (l['credit'] if l['credit'] else 0)

					ingreso_v = (l['debit'] if l['debit'] else 0) 
			else:
				if (data_temp['origen'] == 'internal' and data_temp['destino'] == 'supplier'):
					cprom_acum[0] = cprom_acum[0] -  (l['salida'] if l['salida'] else 0)
					cprom_acum[1] = cprom_acum[1] - (l['debit'] if l['debit'] else 0) - ( (l['credit'] if l['credit'] else 0) * (l['salida'] if l['salida'] else 0) )
					egreso_v = (l['credit'] if l['credit'] else 0) * (l['salida'] if l['salida'] else 0)
				else:
					if l['salida']:
						cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
						cprom_acum[1] = cprom_acum[1] - (l['salida'] if l['salida'] else 0)*(cprom_act_antes if cprom_act_antes else 0)

						egreso_v = (l['salida'] if l['salida'] else 0)*(cprom_act_antes if cprom_act_antes else 0)
					else:
						cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
						cprom_acum[1] = cprom_acum[1] - (l['credit'] if l['credit'] else 0)

						egreso_v = (l['credit'] if l['credit'] else 0)

			cprom_acum[1] = cprom_acum[1] if cprom_acum[0] > 0 else 0
			cprom_act = cprom_acum[1] / cprom_acum[0] if cprom_acum[0] != 0 else 0

			cprom_data[llave] = cprom_acum

			linea = []
			producto_obj = self.env['product.product'].browse(l['product_id'])
			cat_name = []
			cat_obj = producto_obj.categ_id
			while (cat_obj.id):
				cat_name.append(cat_obj.name)
				cat_obj = cat_obj.parent_id
			cat_name.reverse()
			cat_recortado = cat_name[2:]
			n1= ""
			n2= ""
			n3= ""
			n4= ""
			n5= ""
			n1= cat_recortado[0] if len(cat_recortado)> 0 else ''
			n2= cat_recortado[1] if len(cat_recortado)> 1 else ''
			n3= cat_recortado[2] if len(cat_recortado)> 2 else ''
			n4= cat_recortado[3] if len(cat_recortado)> 3 else ''
			n5= cat_recortado[4] if len(cat_recortado)> 4 else ''
	
			linea.append( producto_obj.id or '')
			linea.append( n1)
			linea.append( n2)
			linea.append( n3)
			linea.append( n4)
			linea.append( n5)
			linea.append( producto_obj.product_brand_id.name_get()[0][1] if producto_obj.product_brand_id.id else '')
			linea.append( l['default_code'] if l['default_code'] else '')
			linea.append( l['new_name'] if l['new_name'] else '' )
			linea.append( l['unidad'] if l['unidad'] else '' )
			linea.append (number_format_cantidad(ws, producto_obj.lst_price if producto_obj.lst_price else 0))

			linea.append( number_format_cantidad(ws, cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0 ))
			linea.append( number_format_saldo(ws, cprom_acum[1] if len(cprom_acum)>1 and cprom_acum[1] else 0 ))
			linea.append( number_format_saldo(ws, (cprom_acum[1] if len(cprom_acum)>1 and cprom_acum[1] else 0) / (cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0) if (cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0) != 0 else 0 ))

			linea.append( l['almacen'] if l['almacen'] else '' )
			linea.append( (cprom_acum[1] if len(cprom_acum)>1 and cprom_acum[1] else 0) / (cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0) if (cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0) != 0 else 0 )
			linea.append( producto_obj.id )
			
			array_grabar = linea.copy()

			#producto_obj.standard_price =  costopromedio if costopromedio >producto_obj.standard_price else producto_obj.standard_price
			

			ingreso1 += l['ingreso'] or 0
			ingreso2 += ingreso_v or 0
			salida1 += l['salida'] or 0
			salida2 += egreso_v or 0


		tam_col = [11,11,5,5,7,5,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11]
		if array_grabar and array_grabar[11].value != 0:			
			producto_obj = self.env['product.product'].browse(array_grabar[16])
			costopromedio = array_grabar[15]
			costo_actual = 0
			cantidad_actual = 0
			for i in self.env['stock.valuation.layer'].search([('product_id','=',producto_obj.id)]):
				costo_actual += i.value
				cantidad_actual += i.quantity		
	
			#producto_obj.standard_price = costo_actual/cantidad_actual if cantidad_actual != 0 else 0
			#producto_obj.standard_price =  costopromedio if costopromedio >producto_obj.standard_price else producto_obj.standard_price		
			#std_price_wiz = self.env['stock.change.standard.price'].with_context(active_id=producto_obj.id, active_model='product.product').create({'new_price' : costopromedio, 'counterpart_account_id_required':False})
			#std_price_wiz.with_context(active_id=producto_obj.id, active_model='product.product').change_price()
			#producto_obj.standard_price =  costopromedio
			
			if producto_obj.id in actualizar_montos:
				if costopromedio > producto_obj.standard_price:
					
					self.env.cr.execute("""
							update ir_property set value_float = """ + str(costo_actual/cantidad_actual if cantidad_actual != 0 else 0) + """
							where name = 'standard_price' and company_id = """ + str(self.env.company.id)+ """ and res_id = 'product.product,"""+str(producto_obj.id)+"""' 
							""") 	
					producto_obj.standard_price = costopromedio
			else:
				self.env.cr.execute("""
							update ir_property set value_float = """ + str(costo_actual/cantidad_actual if cantidad_actual != 0 else 0) + """
							where name = 'standard_price' and company_id = """ + str(self.env.company.id)+ """ and res_id = 'product.product,"""+str(producto_obj.id)+"""' 
							""") 	
				actualizar_montos.append(producto_obj.id)				
				producto_obj.standard_price = costopromedio

			ws.append(array_grabar)

		workbook.save(output)
		output.seek(0)

		attach_id = self.env['ir.attachment'].create({
					'name': "Kardex Valorado.xlsx",
					'type': 'binary',
					'datas': base64.encodebytes(output.getvalue()),
					'eliminar_automatico': True
				})
		output.close()


		return {
		'notify':{'title':'Update Costos','message':"Se actualizo los costos de los productos ",'sticky':True,'type':'success'},
		}
			
