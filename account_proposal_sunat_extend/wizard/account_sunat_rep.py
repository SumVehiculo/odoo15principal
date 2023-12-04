from odoo import _, api, fields, models, tools
from odoo.exceptions import UserError
import base64

from io import BytesIO
import re
import uuid
import zipfile
from datetime import datetime, timedelta
import io

import urllib.parse
import requests
import json
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

import subprocess
import sys

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')

BOOK_SIRE = [
	{'name': 'ACEPTA LA PROPUESTA (VENTAS)' ,'nro': 1,'type':'v'},
	{'name': 'REEMPLAZA LA PROPUESTA (VENTAS)' ,'nro': 2,'type':'v'},
	{'name': 'AJUSTES POSTERIORES (VENTAS)' ,'nro': 3,'type':'v'},
	{'name': 'AJUSTES POSTERIORES PLE (VENTAS)' ,'nro': 4,'type':'v'},
	{'name': 'COMPLEMENTAR PROPUESTA (VENTAS)' ,'nro': 5,'type':'v'},
	{'name': 'ACEPTA LA PROPUESTA (COMPRAS)' ,'nro': 1,'type':'c'},
	{'name': u'REEMPLAZA/COMPARA PROPUESTA (COMPRAS)' ,'nro': 2,'type':'c'},
	{'name': 'COMPLEMENTA PROPUESTA (COMPRAS)' ,'nro': 3,'type':'c'},
	{'name': 'AJUSTES POSTERIORES (COMPRAS)' ,'nro': 4,'type':'c'},
	{'name': 'AJUSTES POSTERIORES PLE (COMPRAS)' ,'nro': 5,'type':'c'},
	{'name': 'OPERACIONES NO DOMICILIADOS (COMPRAS)' ,'nro': 6,'type':'c'},
	{'name': 'AJUSTE OPERACIONES NO DOMICILIADOS (COMPRAS)' ,'nro': 7,'type':'c'},
	{'name': 'AJUSTE OPERACIONES NO DOMICILIADOS PLE (COMPRAS)' ,'nro': 8,'type':'c'},
]

class AccountSunatRep(models.TransientModel):
	_inherit = 'account.sunat.wizard'



	

	def comparison_odoo_sunat_sale(self):
		if self.sire_proposal == 'api':
			self.get_data_from_api_sale()
			try:
				sql = self.get_sql_sire(2,"v")			
				#sql = self.sql_sire_1_2_sale(self.period_id.date_start,self.period_id.date_end,self.company_id,2)
				#sql = self.sql_sire_1_2_sale(self.period_id.date_start,self.period_id.date_end,self.company_id, 2)
				#raise UserError (sql)
				self.env.cr.execute(sql)
				res = self.env.cr.fetchall()
				
			except Exception as e:
				raise UserError(f"Error en la consulta SQL: {e}")
			
			value_2_map = {
				str(line.codCar): [				
					line.perPeriodoTributario,
					line.codCar,
					line.codTipoCDP,
					line.numSerieCDP,
					line.numCDP,
					line.codTipoCarga,
					line.codSituacion,
					line.fecEmision,
					line.fecVencPag,
					line.codTipoDocIdentidad,
					line.numDocIdentidad,
					line.nomRazonSocialCliente,
					line.mtoValFactExpo,
					line.mtoBIGravada,
					line.mtoDsctoBI,
					line.mtoIGV,
					line.mtoDsctoIGV,
					line.mtoExonerado,
					line.mtoInafecto,
					line.mtoISC,
					line.mtoBIIvap,
					line.mtoIvap,
					line.mtoIcbp,
					line.mtoOtrosTrib,
					line.mtoTotalCP,
					line.codMoneda,
					line.mtoTipoCambio,
					line.codEstadoComprobante,
					line.desEstadoComprobante,
					line.indOperGratuita,
					line.mtoValorOpGratuitas,
					line.mtoValorFob,
					line.indTipoOperacion,
					line.mtoPorcParticipacion,
					line.mtoValorFobDolar,
					line.fecEmisionMod,
					line.codTipoCDPMod,
					line.numSerieCDPMod,
					line.numCDPMod
				] for line in self.env['account.sunat.sire.sale.data'].search([])
			}

			value_1_map = {(str(row_1[11]) if row_1[11] else '') + (str(row_1[6]) if row_1[6] else '') + (str(row_1[7]) if row_1[7] else '') + (str(row_1[8]).zfill(10) if row_1[8] else ''): row_1 for row_1 in res}
			value_2 = list(value_2_map.keys())
			value_1 = list(value_1_map.keys())
			
			odoo_list = [value_1_map[item] for item in set(value_1) - set(value_2)]
			sunat_list = [value_2_map[item] for item in set(value_2) - set(value_1)]

			workbook = self.comparison_export_to_excel(sunat_list,odoo_list)
			file_name='Diferencias_ODOO_SUNAT_ventas'
			return self.env['popup.it'].get_file('%s.xlsx'%file_name,workbook)	
		else:
			self.get_data_from_txt_sale()
			try:
				txt_data = base64.b64decode(self.document_file)
				lines = txt_data.decode('utf-8').split('\n')
			except Exception as e:
				raise UserError(f"Error al decodificar el archivo: {e}")
			data = [line.split('|') for line in lines if line]
			if data:
				data.pop(0) 
				#value_2 = [str(line[3]) for line in data]
				try:
					sql = self.get_sql_sire(2,"v")							
					self.env.cr.execute(sql)
					res = self.env.cr.fetchall()
				except Exception as e:
					raise UserError(f"Error en la consulta SQL: {e}")
				#raise UserError(str(res))
				value_2_map = {str(line[3]): line for line in data}
			
				value_1_map = {(str(row_1[11]) if row_1[11] else '') + (str(row_1[6]) if row_1[6] else '') + (str(row_1[7]) if row_1[7] else '') + (str(row_1[8]).zfill(10) if row_1[8] else ''): row_1 for row_1 in res}
				#value_1 = [str(row[12]) + str(row[6]) + str(row[7]) + str(row[9]).zfill(10) for row in res if len(row) > 12]
				
				value_2 = list(value_2_map.keys())
				value_1 = list(value_1_map.keys())

				odoo_list = [value_1_map[item] for item in set(value_1) - set(value_2)]
				sunat_list = [value_2_map[item] for item in set(value_2) - set(value_1)]
				#odoo_list = list(set(value_1) - set(value_2))
				#sunat_list = list(set(value_2) - set(value_1))
				workbook = self.comparison_export_to_excel(sunat_list,odoo_list)
				file_name='Diferencias_ODOO_SUNAT_ventas'
				return self.env['popup.it'].get_file('%s.xlsx'%file_name,workbook)		
			
			
		
	def comparison_odoo_sunat_purchase(self):
			try:
				txt_data = base64.b64decode(self.document_file)
				lines = txt_data.decode('utf-8').split('\n')
			except Exception as e:
				raise UserError(f"Error al decodificar el archivo: {e}")
			data = [line.split('|') for line in lines if line]
			if data:
				data.pop(0) 
				#value_2 = [str(line[3]) for line in data]
				try:
					sql = self.sql_sire_2_purchase(self.period_id.date_start, self.period_id.date_end, self.company_id)
					self.env.cr.execute(sql)
					res = self.env.cr.fetchall()
				except Exception as e:
					raise UserError(f"Error en la consulta SQL: {e}")
				
				value_2_map = {str(line[3]): line for line in data}
				value_1_map = {str(row[12]) + str(row[6]) + str(row[7]) + str(row[9]).zfill(10): row for row in res}
				#value_1 = [str(row[12]) + str(row[6]) + str(row[7]) + str(row[9]).zfill(10) for row in res if len(row) > 12]
				
				value_2 = list(value_2_map.keys())
				value_1 = list(value_1_map.keys())

				odoo_list = [value_1_map[item] for item in set(value_1) - set(value_2)]
				sunat_list = [value_2_map[item] for item in set(value_2) - set(value_1)]
				#odoo_list = list(set(value_1) - set(value_2))
				#sunat_list = list(set(value_2) - set(value_1))
				workbook = self.comparison_export_to_excel(sunat_list,odoo_list)
				file_name='Diferencias_ODOO_SUNAT_compras'
				return self.env['popup.it'].get_file('%s.xlsx'%file_name,workbook)			

	def comparison_export_to_excel(self,sunat_list, odoo_list):
		

		header_font = Font(bold=True, size=12)
		header_alignment = Alignment(horizontal='center', vertical='center')
		header_fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type = "solid")
		border_side = Side(border_style="thin", color="000000")
		border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

		wb = openpyxl.Workbook()

		def apply_cell_style(ws, cell_range):
			for row in ws[cell_range]:
				for cell in row:
					cell.border = border

		ws1 = wb.active
		ws1.title = "SUNAT"
		if sunat_list:
			num_columns = len(sunat_list[0])
			for col in range(1, num_columns + 1):
				cell = ws1.cell(row=1, column=col)
				cell.value = f"Campo {col}"
				cell.font = header_font
				cell.alignment = header_alignment
				cell.fill = header_fill
				ws1.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

			apply_cell_style(ws1, f'A1:{openpyxl.utils.get_column_letter(num_columns)}{len(sunat_list)+1}')

			for i, row in enumerate(sunat_list, start=2):
				for j, value in enumerate(row, start=1):
					ws1.cell(row=i, column=j).value = value

		ws2 = wb.create_sheet(title="ODOO")
		if odoo_list:
			num_columns = len(odoo_list[0])
			for col in range(1, num_columns + 1):
				cell = ws2.cell(row=1, column=col)
				cell.value = f"Campo {col}"
				cell.font = header_font
				cell.alignment = header_alignment
				cell.fill = header_fill
				ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

			apply_cell_style(ws2, f'A1:{openpyxl.utils.get_column_letter(num_columns)}{len(odoo_list)+1}')

			for i, row in enumerate(odoo_list, start=2):
				for j, value in enumerate(row, start=1):
					ws2.cell(row=i, column=j).value = value
		output = BytesIO()
		wb.save(output)
		output.getvalue()
		output_datas = base64.b64encode(output.getvalue())
		output.close()
		return output_datas