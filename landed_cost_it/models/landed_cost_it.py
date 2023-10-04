# -*- coding: utf-8 -*-

from mimetypes import init
from odoo import models, fields, api
from odoo.exceptions import UserError, ValidationError
import base64
from io import BytesIO
import subprocess
import sys

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')


##SE VUELVE A PONER EL CAMPO AQUI YA QUE SE NECESITA PARA CALCULO DE GV Y NO ES POSIBLE DEPENDER DEL CAMPO KARDEX_VALORADO_IT YA QUE ESTE DEPENDE DE GV
class stock_move(models.Model):
	_inherit = 'stock.move'

	price_unit_it = fields.Float('Precio Unitario',digits=(12,8))

class LandedCostIt(models.Model):
	_name = 'landed.cost.it'
	_inherit = ['mail.thread']

	name = fields.Char(string='Nombre')

	prorratear_en = fields.Selection([('cantidad', 'Por Cantidad'), ('valor', 'Por Valor')],string='Prorratear en funcion', required=True, default='cantidad')

	picking_ids = fields.Many2many('stock.picking', 'gastos_vinculado_picking_rel', 'gastos_id', 'picking_id', string='Albaranes')
	detalle_ids = fields.One2many('landed.cost.it.line', 'gastos_id', 'Detalle')
	company_id = fields.Many2one('res.company',string=u'Compañia',required=True, default=lambda self: self.env.company,readonly=True)
	invoice_ids = fields.One2many('landed.cost.invoice.line', 'landed_id',string='Facturas')
	purchase_ids = fields.One2many('landed.cost.purchase.line', 'landed_id',string='Ordenes de Compra')
	advalorem_ids = fields.One2many('landed.cost.advalorem.line', 'landed_id',string='Advalorem')

	state = fields.Selection([('draft', 'Borrador'), ('done', 'Finalizado')],string='Estado', default='draft')
	total_flete = fields.Float(string='Total GV', digits=(12, 2), store=True)
	total_flete_usd = fields.Float(string='Total GV USD', digits=(12, 2))
	total_factor = fields.Float(string='Total Factor', digits=(12, 2), store=True)
	date_kardex = fields.Datetime(string='Fecha Kardex')
	purchase_origin_id = fields.Many2one('purchase.order',string='Pedido de Compra')

	def get_info(self):
		
		picks = self.env['stock.picking'].search([('landed_cost_id','=',self.id)])
		
		if picks:
			for p in picks:
				self.picking_ids = [(6, 0, [p.id])]
		
		self.agregar_lineas()
		invoices = self.env['account.move'].search([('landed_cost_id','=',self.id),('state','=','posted')])
		for move in invoices:
			for line in move.line_ids:
				if line.product_id.is_landed_cost:
					vals = {
						'invoice_id': line.id,
						'invoice_date': line.move_id.invoice_date,
						'type_document_id': line.type_document_id.id,
						'nro_comp': line.nro_comp,
						'date': line.move_id.date,
						'partner_id': line.partner_id.id,
						'product_id': line.product_id.id,
						'debit': (line.debit - line.credit),
						'amount_currency': (line.debit - line.credit)/line.tc if line.move_id.currency_id.id == line.company_id.currency_id.id else line.amount_currency,
						'tc': line.tc,
						'type_landed_cost_id': line.product_id.type_landed_cost_id.id,
						'company_id': line.company_id.id,
					}
					self.write({'invoice_ids' :([(0,0,vals)]) })
					self._change_flete()
		
		
	def get_invoices(self):
		wizard = self.env['get.landed.invoices.wizard'].create({
			'landed_id': self.id,
			'company_id':self.company_id.id
		})
		module = __name__.split('addons.')[1].split('.')[0]
		view = self.env.ref('%s.view_get_landed_invoices_wizard' % module)
		return {
			'name':u'Seleccionar Facturas',
			'res_id':wizard.id,
			'view_mode': 'form',
			'res_model': 'get.landed.invoices.wizard',
			'view_id': view.id,
			'context': self.env.context,
			'target': 'new',
			'type': 'ir.actions.act_window',
		}

	def get_purchases(self):
		wizard = self.env['get.landed.purchases.wizard'].create({
			'landed_id': self.id,
			'company_id':self.company_id.id
		})
		module = __name__.split('addons.')[1].split('.')[0]
		view = self.env.ref('%s.view_get_landed_purchases_wizard' % module)
		return {
			'name':u'Seleccionar Compras',
			'res_id':wizard.id,
			'view_mode': 'form',
			'res_model': 'get.landed.purchases.wizard',
			'view_id': view.id,
			'context': self.env.context,
			'target': 'new',
			'type': 'ir.actions.act_window',
		}

	@api.onchange('invoice_ids','purchase_ids')
	def _change_flete(self):
		flete = 0
		flete_usd = 0
		for elem in self.purchase_ids:
			flete += elem.price_total_signed
		for elem in self.invoice_ids:
			flete += elem.debit
			flete_usd += elem.amount_currency
		self.total_flete = flete
		self.total_flete_usd = flete_usd

	@api.model
	def create(self, vals):
		id_seq = self.env['ir.sequence'].search([('name', '=', 'Gastos Vinculados IT'),('company_id','=',self.env.company.id)],limit=1)

		if not id_seq:
			id_seq = self.env['ir.sequence'].create({'name': 'Gastos Vinculados IT', 'company_id': self.env.company.id, 'implementation': 'no_gap','active': True, 'prefix': 'GV-', 'padding': 4, 'number_increment': 1, 'number_next_actual': 1})

		vals['name'] = id_seq._next()
		t = super(LandedCostIt, self).create(vals)
		return t

	def unlink(self):
		if self.state == 'done':
			raise UserError('No se puede eliminar un Gasto Vinculado Terminado')

		for i in self.picking_ids:
			i.unlink()

		for i in self.detalle_ids:
			i.unlink()

		t = super(LandedCostIt, self).unlink()
		return t

	def borrador(self):
		self.state = 'draft'
		#for i in self.detalle_ids:
		#	costo_actual = 0
		#	cantidad_actual = 0
		#	for ij in self.env['stock.valuation.layer'].search([('product_id','=',i.stock_move_id.product_id.id)]):
		#		costo_actual += ij.value
		#		cantidad_actual += ij.quantity					
		#	i.stock_move_id.product_id.standard_price = costo_actual/cantidad_actual if cantidad_actual != 0 else 0

		#	costopromedio = (costo_actual - i.flete) / cantidad_actual if cantidad_actual else 0
						
		#	std_price_wiz = self.env['stock.change.standard.price'].with_context(active_id=i.stock_move_id.product_id.id, active_model='product.product').create({'new_price' : costopromedio, 'counterpart_account_id_required':False})
		#	std_price_wiz.with_context(active_id=i.stock_move_id.product_id.id, active_model='product.product').change_price()

	def procesar(self):
		self.state = 'done'
		#for i in self.detalle_ids:
		#	costo_actual = 0
		#	cantidad_actual = 0
		#	for ij in self.env['stock.valuation.layer'].search([('product_id','=',i.stock_move_id.product_id.id)]):
		#		costo_actual += ij.value
		#		cantidad_actual += ij.quantity					
		#	i.stock_move_id.product_id.standard_price = costo_actual/cantidad_actual if cantidad_actual != 0 else 0
		
		#	costopromedio = (costo_actual + i.flete) / cantidad_actual  if cantidad_actual else 0
				
		#	std_price_wiz = self.env['stock.change.standard.price'].with_context(active_id=i.stock_move_id.product_id.id, active_model='product.product').create({'new_price' : costopromedio, 'counterpart_account_id_required':False})
		#	std_price_wiz.with_context(active_id=i.stock_move_id.product_id.id, active_model='product.product').change_price()
			

	def calcular(self):
		self.refresh()
		self._change_flete()
		total_fle_lines = 0
		for i in self.detalle_ids:
			i.refresh()
			total_prorrateo = 0
			for m in self.detalle_ids:
				total_prorrateo += m.cantidad_rel if self.prorratear_en == 'cantidad' else m.valor_rel_signed

			i.factor = ((i.cantidad_rel if self.prorratear_en == 'cantidad' else i.valor_rel_signed) /
						total_prorrateo) if total_prorrateo != 0 else 0
			i.refresh()
			valor_mn = sum(line['valormn'] for line in self.advalorem_ids.filtered(lambda line: line.product_id.id == i.stock_move_id.product_id.id and line.picking_id.id == i.stock_move_id.picking_id.id))
			valor_me = sum(line['valorme'] for line in self.advalorem_ids.filtered(lambda line: line.product_id.id == i.stock_move_id.product_id.id and line.picking_id.id == i.stock_move_id.picking_id.id))
			i.flete = (i.factor * self.total_flete)
			i.flete_usd = (i.factor * self.total_flete_usd)
			i.advalorem = valor_mn or 0
			i.advalorem_usd = valor_me or 0
			i.total = i.flete + i.advalorem + i.valor_rel_signed
			i.total_usd = i.flete_usd + i.advalorem_usd + i.valor_rel
			total_fle_lines +=  i.flete
			
		#REDONDEO
		if len(self.detalle_ids)>0:
			diferencia_flete = 0
			if total_fle_lines < self.total_flete:
				diferencia_flete = self.total_flete - total_fle_lines
				self.detalle_ids[0].flete = self.detalle_ids[0].flete + diferencia_flete

			if total_fle_lines > self.total_flete:
				diferencia_flete = total_fle_lines - self.total_flete
				self.detalle_ids[0].flete = self.detalle_ids[0].flete - diferencia_flete

		return True

	def agregar_lineas(self):
		self.ensure_one()
		for i in self.detalle_ids:
			i.unlink()

		for i in self.picking_ids:
			for j in i.move_lines: 
				data = {
					'stock_move_id': j.id,
					'gastos_id': self.id,
				}
				self.env['landed.cost.it.line'].create(data)
	
	def generate_excel(self):
		import io
		from xlsxwriter.workbook import Workbook
		from xlsxwriter.utility import xl_rowcol_to_cell

		ReportBase = self.env['report.base']
		direccion = self.env['account.main.parameter'].search([('company_id','=',self.company_id.id)],limit=1).dir_create_file

		if not direccion:
			raise UserError(u'No existe un Directorio Exportadores configurado en Parametros Principales de Contabilidad para su Compañía')

		workbook = Workbook(direccion +'Reporte_GV.xlsx')
		workbook, formats = ReportBase.get_formats(workbook)

		numbertotalocho = workbook.add_format({'num_format':'0.00000000','bold': True})
		numbertotalocho.set_align('right')
		numbertotalocho.set_align('vcenter')
		numbertotalocho.set_border(style=1)
		numbertotalocho.set_font_size(10.5)
		numbertotalocho.set_font_name('Times New Roman')
		numbertotalocho.set_underline()

		import importlib
		import sys
		importlib.reload(sys)

		worksheet = workbook.add_worksheet("GV")
		worksheet.set_tab_color('blue')

		HEADERS = ['REFERENCIA','DE','PARA','PRODUCTO','UNIDAD','CANTIDAD','P.UNIT','VALOR','FACTOR','GASTO V','ADVALOREM','VALOR TOTAL','C.UNIT']
		worksheet = ReportBase.get_headers(worksheet,HEADERS,0,0,formats['boldbord'])

		x=1
		init = 1

		for line in self.detalle_ids:
			worksheet.write(x,0,line.picking_rel.display_name if line.picking_rel else '',formats['especial1'])
			worksheet.write(x,1,line.origen_rel.display_name if line.origen_rel else '',formats['especial1'])
			worksheet.write(x,2,line.destino_rel.display_name if line.destino_rel else '',formats['especial1'])
			worksheet.write(x,3,line.producto_rel.display_name if line.producto_rel else '',formats['especial1'])
			worksheet.write(x,4,line.unidad_rel.display_name if line.unidad_rel else '',formats['especial1'])
			worksheet.write(x,5,line.cantidad_rel if line.cantidad_rel else 0,formats['numberdos'])
			worksheet.write(x,6,line.precio_unit_signed if line.precio_unit_signed else 0,formats['numberocho'])
			worksheet.write(x,7,line.valor_rel_signed if line.valor_rel_signed else 0,formats['numberdos'])
			worksheet.write(x,8,line.factor if line.factor else '',formats['numberocho'])
			worksheet.write(x,9,line.flete if line.flete else '',formats['numberocho'])
			worksheet.write(x,10,line.advalorem if line.advalorem else '',formats['numberdos'])
			worksheet.write(x,11,line.total if line.total else '',formats['numberdos'])
			worksheet.write(x,12,line.total/line.cantidad_rel if line.cantidad_rel and line.cantidad_rel != 0 else 0,formats['numberocho'])
			x += 1

		worksheet.write_formula(x,5, '=SUM(' + xl_rowcol_to_cell(init,5) + ':' + xl_rowcol_to_cell(x-1,5) + ')', formats['numbertotal'])
		worksheet.write_formula(x,7, '=SUM(' + xl_rowcol_to_cell(init,7) + ':' + xl_rowcol_to_cell(x-1,7) + ')', formats['numbertotal'])
		worksheet.write_formula(x,9, '=SUM(' + xl_rowcol_to_cell(init,9) + ':' + xl_rowcol_to_cell(x-1,9) + ')', numbertotalocho)
		worksheet.write_formula(x,10, '=SUM(' + xl_rowcol_to_cell(init,10) + ':' + xl_rowcol_to_cell(x-1,10) + ')', formats['numbertotal'])
		worksheet.write_formula(x,11, '=SUM(' + xl_rowcol_to_cell(init,11) + ':' + xl_rowcol_to_cell(x-1,11) + ')', formats['numbertotal'])

		widths = [15,25,20,30,15,15,20,20,20,20,20,20,17]
		worksheet = ReportBase.resize_cells(worksheet,widths)
		workbook.close()

		f = open(direccion +'Reporte_GV.xlsx', 'rb')

		return self.env['popup.it'].get_file('Reporte GV.xlsx',base64.encodestring(b''.join(f.readlines())))

	def get_excel_saldos(self):
		self.ensure_one()
		sql = sql1 = sqlsum = ""
		sql2 = """valor_p + valormn """
		sql3 = """ valormn """
		for elem in self.env['landed.cost.it.type'].search([]):
			sql2 += "+"
			sql3 += "+"
			sql2 += """ coalesce("%s",0) """%(elem.code)
			sql3 += """ coalesce("%s",0) """%(elem.code)
			sql += ", \n"
			sql1 += """ coalesce("%s",0) as "%s", """%(elem.code, elem.name)
			sqlsum += """ sum(coalesce("%s",0)) as "%s", """%(elem.code, elem.name)
			sql += """ a.factor*(select sum(debit) from landed_cost_invoice_line where landed_id=%d and type_landed_cost_id = %d) as "%s" """%(self.id,elem.id,elem.code)

		self.env.cr.execute("""(select almacen,codigo,producto,cantidad,factor,valor_p, valormn as "Advalorem ", 
		%s
								%s as total_gv,
								%s as costo_total,
								(%s)/valor_p as factor_d,
								(%s)/cantidad as costo_unitario

								from 
								(
								select   
								b.almacen, 
								b.default_code as codigo, 
								b.name_template as producto, 
								b.ingreso as cantidad,
								a.factor,
								b.debit as valor_p,
								coalesce(c.valormn,0) as valormn %s
								from landed_cost_it_line a
								LEFT JOIN 
								(
								select sl.complete_name as almacen,
								line.stock_move_id as stock_moveid,
								sm.product_id,
								pp.default_code,
								pt.name as name_template,
								sum(sm.product_qty) as ingreso,
								sum(line.valor_rel_signed) as debit
								from landed_cost_it_line line
								left join stock_move sm on sm.id = line.stock_move_id
								left join stock_location sl on sl.id = sm.location_dest_id
								left join product_product pp on pp.id = sm.product_id
								left join product_template pt on pt.id = pp.product_tmpl_id
								where line.gastos_id = %d
								group by sl.complete_name, line.stock_move_id, sm.product_id, pp.default_code, pt.name

								) b ON b.stock_moveid = a.stock_move_id
								LEFT JOIN stock_move SM on SM.id = b.stock_moveid
								LEFT JOIN (select landed_id, picking_id , product_id, sum(coalesce(valormn,0)) as valormn from landed_cost_advalorem_line GROUP BY landed_id, picking_id, product_id) c ON c.product_id = b.product_id AND a.gastos_id = c.landed_id and SM.picking_id = c.picking_id
								where a.gastos_id = %d
								)tt)
								UNION ALL
								(select '' as almacen,'' as codigo,'' as producto,null as cantidad, null as factor,sum(tt2.valor_p) as valor_p, sum(tt2.valormn) as "Advalorem ", 
		%s
								sum(%s) as total_gv,
								sum(%s) as costo_total,
								null as factor_d,
								null as costo_unitario

								from 
								(
								select   
								b.almacen, 
								b.default_code as codigo, 
								b.name_template as producto, 
								b.ingreso as cantidad,
								a.factor,
								b.debit as valor_p,
								coalesce(c.valormn,0) as valormn %s
								from landed_cost_it_line a
								LEFT JOIN 
								(
									
								select sl.complete_name as almacen,
								line.stock_move_id as stock_moveid,
								sm.product_id,
								pp.default_code,
								pt.name as name_template,
								sum(sm.product_qty) as ingreso,
								sum(line.valor_rel_signed) as debit
								from landed_cost_it_line line
								left join stock_move sm on sm.id = line.stock_move_id
								left join stock_location sl on sl.id = sm.location_dest_id
								left join product_product pp on pp.id = sm.product_id
								left join product_template pt on pt.id = pp.product_tmpl_id
								where line.gastos_id = %d
								group by sl.complete_name, line.stock_move_id, sm.product_id, pp.default_code, pt.name

								) b ON b.stock_moveid = a.stock_move_id
								LEFT JOIN stock_move SM on SM.id = b.stock_moveid
								LEFT JOIN (select landed_id , picking_id, product_id, sum(coalesce(valormn,0)) as valormn from landed_cost_advalorem_line GROUP BY landed_id, picking_id, product_id) c ON c.product_id = b.product_id AND a.gastos_id = c.landed_id and SM.picking_id = c.picking_id
								where a.gastos_id = %d
								)tt2)"""%(sql1,
									sql3,
									sql2,
									sql2,
									sql2,
									sql,
									self.id,
									self.id,
									sqlsum,
									sql3,
									sql2,
									sql,
									self.id,
									self.id
								))
		res = self.env.cr.fetchall()
		colnames = [
			desc[0] for desc in self.env.cr.description
		]
		res.insert(0, colnames)

		wb = openpyxl.Workbook()
		ws = wb.active
		row_position = 1
		col_position = 1
		for index, row in enumerate(res, row_position):
			for col, val in enumerate(row, col_position):
				ws.cell(row=index, column=col).value = val
		output = BytesIO()
		wb.save(output)
		output.getvalue()
		output_datas = base64.b64encode(output.getvalue())
		output.close()

		return self.env['popup.it'].get_file('%s.xlsx'%(self.name),output_datas)

class LandedCostItLine(models.Model):
	_name = 'landed.cost.it.line'
	stock_move_id = fields.Many2one('stock.move', 'Stock Move')
	gastos_id = fields.Many2one('landed.cost.it', 'Gastos Vinculado',ondelete="cascade")

	picking_rel = fields.Many2one('stock.picking',string='Referencia', related='stock_move_id.picking_id')
	origen_rel = fields.Many2one('stock.location', string='De',related='stock_move_id.location_id')
	destino_rel = fields.Many2one('stock.location',string='Para', related='stock_move_id.location_dest_id')
	producto_rel = fields.Many2one('product.product',string='Producto', related='stock_move_id.product_id')
	unidad_rel = fields.Many2one('uom.uom',string='Unidad de Medida', related='stock_move_id.product_uom')
	cantidad_rel = fields.Float(string='Cantidad', related='stock_move_id.product_qty')
	precio_unitario_rel = fields.Float(string='Precio Unitario', related='stock_move_id.price_unit_it')
	precio_unit_signed = fields.Float(string='Precio Unitario Soles', compute="get_price_unit_signed",store=True, digits=(64,8))
	valor_rel = fields.Float(string='Valor', compute="get_valor_rel",store=True)
	valor_rel_signed = fields.Float(string='Valor Soles', compute="get_valor_rel",store=True)
	
	valuation_id = fields.Many2one('stock.valuation.layer','Valoracion')

	factor = fields.Float(string='Factor', digits=(12, 10))
	flete = fields.Float(string='Total GV PEN', digits=(12, 6))
	flete_usd = fields.Float(string='Total GV USD', digits=(12, 6))
	advalorem = fields.Float(string='Advalorem PEN', digits=(12, 6))
	advalorem_usd = fields.Float(string='Advalorem USD', digits=(12, 6))
	total = fields.Float(string='Valor Total PEN', digits=(12, 6))
	total_usd = fields.Float(string='Valor Total USD', digits=(12, 6))

	@api.depends('stock_move_id.price_unit_it','stock_move_id.picking_id.tc')
	def get_price_unit_signed(self):
		for record in self:
			record.precio_unit_signed = (record.stock_move_id.price_unit_it or 0) * (record.stock_move_id.picking_id.tc or 1)

	@api.depends('stock_move_id.product_qty','precio_unit_signed')
	def get_valor_rel(self):
		for record in self:
			record.valor_rel = record.stock_move_id.product_qty * record.stock_move_id.price_unit_it
			record.valor_rel_signed = record.stock_move_id.product_qty * record.precio_unit_signed

class LandedCostInvoiceLine(models.Model):
	_name = 'landed.cost.invoice.line'
	
	landed_id = fields.Many2one('landed.cost.it', 'Gastos Vinculado',ondelete="cascade")
	invoice_id = fields.Many2one('account.move.line',string='Factura')
	invoice_date = fields.Date(string='Fecha Factura')
	type_document_id = fields.Many2one('l10n_latam.document.type',string='Tipo de Documento')
	nro_comp = fields.Char(string='Nro Comprobante')
	date = fields.Date(string='Fecha Contable')
	partner_id = fields.Many2one('res.partner',string='Socio')
	product_id = fields.Many2one('product.product',string='Producto')
	debit = fields.Float(string='Debe',digits=(64,2))
	amount_currency = fields.Float(string='Monto Me',digits=(64,2))
	tc = fields.Float(string='TC',digits=(12,4))
	type_landed_cost_id = fields.Many2one('landed.cost.it.type',string='Tipo G.V.')
	company_id = fields.Many2one('res.company',string=u'Compañía')

class LandedCostPurchaseLine(models.Model):
	_name = 'landed.cost.purchase.line'
	
	landed_id = fields.Many2one('landed.cost.it', 'Gastos Vinculado',ondelete="cascade")
	purchase_id = fields.Many2one('purchase.order.line',string='Compra')
	purchase_date = fields.Date(string='Fecha Pedido')
	name = fields.Char(string='Pedido')
	partner_id = fields.Many2one('res.partner',string='Socio')
	product_id = fields.Many2one('product.product',string='Producto')
	price_total_signed = fields.Float(string='Total Soles',digits=(64,2))
	tc = fields.Float(string='TC',digits=(12,4))
	currency_id = fields.Many2one('res.currency',string='Moneda')
	price_total = fields.Float(string='Total',digits=(64,2))
	company_id = fields.Many2one('res.company',string=u'Compañía')

class LandedCostAdvaloremLine(models.Model):
	_name = 'landed.cost.advalorem.line'
	_description = 'Landed Cost Advalorem Line'

	@api.depends('product_id','landed_id','landed_id.detalle_ids')
	def _check_products(self):
		for object in self:
			object.correct_product = False
			product_list = []
			if object.landed_id:
				for x in object.landed_id.detalle_ids:
					if x.stock_move_id.product_id:
						product_list.append(x.stock_move_id.product_id.id)
			if object.product_id.id in product_list:
				object.correct_product = True
			  
	landed_id = fields.Many2one('landed.cost.it', 'Gastos Vinculado')
	invoice_id = fields.Many2one('account.move.line',string='Factura')
	picking_id = fields.Many2one('stock.picking',string='Referencia')
	product_id = fields.Many2one('product.product',string='Producto')
	valormn = fields.Float(string='Valor MN',digits=(12,2))
	valorme = fields.Float(string='Valor ME',digits=(12,2))
	correct_product = fields.Boolean(string='Producto Pertenece a GV',store=True,compute='_check_products')