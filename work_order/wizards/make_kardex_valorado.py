from odoo  import models, fields, api
from odoo.exceptions import UserError


from datetime import timedelta
from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font,  Alignment
from openpyxl.styles.borders import BORDER_THIN
import base64

def border(ws,texto):
	cell = WriteOnlyCell(ws, value=texto)
	cell.font = Font(name='Arial',size=13,bold=True)
	cell.border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000'))
	return cell


def number_format_costo(ws,numero):
    cell = WriteOnlyCell(ws, value=numero)
    cell.number_format = "0.000000"
    return cell

def number_format_cantidad(ws,numero):
    cell = WriteOnlyCell(ws, value=numero)
    cell.number_format = "0.00"
    return cell

def border_sub(ws,texto):
    cell = WriteOnlyCell(ws, value=texto)
    cell.font = Font(name='Arial',size=12,bold=True)
    cell.border = Border(
    left=Side(border_style=BORDER_THIN, color='00000000'),
    right=Side(border_style=BORDER_THIN, color='00000000'),
    top=Side(border_style=BORDER_THIN, color='00000000'),
    bottom=Side(border_style=BORDER_THIN, color='00000000'))
    return cell


class MakeKardexValorado(models.TransientModel):
    _inherit="make.kardex.valorado"

    def do_csvtoexcel(self):
        if self.tipo_kardex == 'promedio':

            import openpyxl
            from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
            from openpyxl.styles.borders import Border, Side, BORDER_THIN
            from openpyxl import Workbook
            values = {}
            from openpyxl.utils import get_column_letter
            from openpyxl.cell import WriteOnlyCell
            cad = ""

            s_prod = [-1,-1,-1]
            s_loca = [-1,-1,-1]
            if self.alllocations == True:
                locat_ids = self.env['stock.location'].search( [('usage','in',('internal','inventory','transit','procurement','production'))] )
                lst_locations = locat_ids.ids
            else:
                lst_locations = self.location_ids.ids
            lst_products  = self.products_ids.ids
            productos='('
            almacenes='('
            date_ini=self.fini
            date_fin=self.ffin
            if self.allproducts:
                lst_products = self.env['product.product'].with_context(active_test=False).search([('check_unidad_especifica','=',False)]).ids

            else:
                lst_products = self.products_ids.filtered(lambda r: r.check_unidad_especifica == False).ids

            if len(lst_products) == 0:
                raise UserError('No existen productos seleccionados')

            for producto in lst_products:
                productos=productos+str(producto)+','
                s_prod.append(producto)
            productos=productos[:-1]+')'
            for location in lst_locations:
                almacenes=almacenes+str(location)+','
                s_loca.append(location)
            almacenes=almacenes[:-1]+')'



            import io
            output = io.BytesIO()

            workbook = Workbook(write_only=True)
            ws = workbook.create_sheet("Kardex Valorado")
            x= 10
            tam_col = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
        
        
            ws.column_dimensions['A'].width = 13.30
            ws.column_dimensions['B'].width = 9.71
            ws.column_dimensions['E'].width = 10
            ws.column_dimensions['F'].width = 24.43
            ws.column_dimensions['G'].width = 13.71
            ws.column_dimensions['H'].width = 86.29
            ws.column_dimensions['J'].width = 42.14
            ws.column_dimensions['K'].width = 20.86
            ws.column_dimensions['L'].width = 8.86
            ws.column_dimensions['M'].width = 11
            ws.column_dimensions['N'].width = 18
            ws.column_dimensions['O'].width = 11.86
            ws.column_dimensions['P'].width = 18.30
            ws.column_dimensions['Q'].width = 11.14
            ws.column_dimensions['R'].width = 18.30
            ws.column_dimensions['S'].width = 20
            ws.column_dimensions['T'].width = 17.43
            ws.column_dimensions['U'].width = 47.43
            ws.column_dimensions['V'].width = 47
            ws.column_dimensions['W'].width = 12.14

        
            cell = WriteOnlyCell(ws, value="KARDEX VALORADO")
            cell.font = Font(name='Bahnschrift Light SemiCondensed',size=33,bold=True)
            cell.alignment = Alignment(horizontal='center')


            cell_fini = WriteOnlyCell(ws, value="FECHA INICIO:")
            cell_fini.font = Font(name='Arial',size=10,bold=True)

            cell_ffin = WriteOnlyCell(ws, value="FECHA FINAL:")
            cell_ffin.font = Font(name='Arial',size=10,bold=True)



            ws.merged_cells.ranges.append(get_column_letter(5)+ "1:" + get_column_letter(10) + '1')
            ws.append(["","","","",cell])

            ws.append([""])


            ws.append([cell_fini,str(self.fini)])
            ws.append([cell_ffin,str(self.ffin)])

            ws.append([""])
            ws.append([""])
            ws.append([""])

                
            import datetime		

            ws.merged_cells.ranges.append(get_column_letter(13)+ "8:" + get_column_letter(14) + '8')

            ws.merged_cells.ranges.append(get_column_letter(15)+ "8:" + get_column_letter(16) + '8')

            ws.merged_cells.ranges.append(get_column_letter(17)+ "8:" + get_column_letter(18) + '8')
            ws.append([""]*12 + [border(ws,"Ingreso"),border(ws,""),border(ws,"Salida"),border(ws,""),border(ws,"Saldo"),border(ws,"")] )


            linea = [border(ws,u"Fecha"),border(ws,u"Hora"),border(ws,u"Tipo"),border(ws,u"Serie"),border(ws,u"Número"),border(ws,u"Doc. Almacen"),border(ws,u"RUC"),border(ws,u"Empresa"),border(ws,u"T. OP."),border(ws,u"Producto"),border(ws,u"Codigo Producto"),border(ws,u"Unidad"),border(ws,u"Cantidad"),border(ws,u"Costo"),border(ws,u"Cantidad"),border(ws,u"Costo"),border(ws,u"Cantidad"),border(ws,u"Costo"),border_sub(ws,u"Costo Adquisición"),border_sub(ws,u"Costo Promedio"),border(ws,u"Ubicacion Origen"),border(ws,u"Ubicacion Destino"),border(ws,u"Almacen"),border(ws,u"IDProducto"),border(ws,u"Pedido")]




            if self.check_account:	

                linea += [border(ws,u"Cuenta Valuación"),border(ws,u"Cuenta Salida"),border(ws,u"Cuenta Analítica")]
            ws.append(linea)


                # worksheet.merge_range(8,26,9,26, u"Etiqueta Analítica",boldbord)

            text_report = "<b>Cargando Kardex Valorado</b><br/><center>Ejecutando SQL del kardex ... Espere por favor</center>"

            if self.moneda == 'usd':
                text_report = "<b>Cargando Kardex Valorado</b><br/><center>Ejecutando SQL del kardex (USD)... Espere por favor</center>"			
            self.send_message(text_report)
            
            total_all = []


            config = self.env['kardex.parameter'].search([('company_id','=',self.env.company.id)])

            date_ini = '%d-01-01' % ( config._get_anio_start(date_fin.year) )

            kardex_save_obj = False

            kardex_save_obj = self.env['kardex.save'].search([('company_id','=',self.env.company.id),('state','=','done'),('name.date_end','<',self.fini)]).sorted(lambda l: l.name.code , reverse=True)
            if len(kardex_save_obj)>0:
                kardex_save_obj = kardex_save_obj[0]
                date_ini = kardex_save_obj.name.date_end + timedelta(days=1)
            else:
                kardex_save_obj = False

                    
            
            text_report_linea = ''
            if kardex_save_obj:
                text_report_linea = "<b>Se usara el saldo guardado: "+str(kardex_save_obj.name.code)+ " </b>"
                
            si_existe = ""
            if kardex_save_obj:
                si_existe = """union all


    select 
                    ksp.producto as product_id,
                    ksp.almacen as location_id,
                    '' as origen_usage,
                    sl.usage as destino_usage,
                    ksp.cprom"""+ ('_dolar' if self.moneda == 'usd' else '')+""" * ksp.stock as debit,
                    0 as credit,
                    (fecha || ' 00:00:00')::timestamp as fechax,

                    '' as type_doc,
                    '' as serial,
                    '' as nro,
                    '' as numdoc_cuadre,
                    '' as nro_documento,

                    'Saldo Inicial' as name,
                    '' as operation_type,
                    pname.new_name,
                    coalesce(pp.default_code,pt.default_code) as default_code,
                    uu.name as unidad,
                    ksp.stock as ingreso,
                    0 as salida,
                    ksp.cprom"""+ ('_dolar' if self.moneda == 'usd' else '')+""" as cadquiere,
                    '' as origen,
                    sl.complete_name as destino,
                    sl.complete_name as almacen,
                    '' as ctanalitica,
                    0 as stock_moveid,
                    '' as pedido

                from kardex_save_period ksp 
                inner join stock_location sl on sl.id = ksp.almacen
                inner join product_product pp on pp.id = ksp.producto
                inner join product_template pt on pt.id = pp.product_tmpl_id
                inner join uom_uom uu on uu.id = pt.uom_id
                inner join product_category pc on pc.id = pt.categ_id
                left join stock_production_lot spt on spt.id = ksp.lote
                    LEFT JOIN ( SELECT t_pp.id,
                        ((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
                        FROM product_product t_pp
                            JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
                            left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
                            LEFT JOIN product_variant_combination pvc ON pvc.product_product_id = t_pp.id
                            LEFT JOIN product_template_attribute_value ptav ON ptav.id = pvc.product_template_attribute_value_id
                            LEFT JOIN product_attribute_value pav ON pav.id = ptav.product_attribute_value_id
                        GROUP BY t_pp.id) pname ON pname.id = pp.id
                    where save_id = """+str(kardex_save_obj.id)+"""
                    and ksp.almacen in  """+str(almacenes)+"""
                    and ksp.producto in  """+str(productos)+"""
                    """

            if self.moneda == 'pen':
                self.env.cr.execute("""  
                    select  
                    product_id,
                    location_id,
                    origen_usage,
                    destino_usage,
                    debit,
                    credit,
                    fechax,
                    type_doc,
                    serial,
                    nro,
                    numdoc_cuadre,
                    nro_documento,
                    name,
                    operation_type,
                    new_name,
                    default_code,
                    unidad,
                    ingreso,
                    salida,
                    cadquiere,
                    origen,
                    destino,
                    almacen,
                    ctanalitica,
                    stock_moveid,
                    pedido
                from (
                    select 
                    product_id,
                    location_id,
                    origen_usage,
                    destino_usage,
                    debit,
                    credit,
                    fechax,
                    type_doc,
                    serial,
                    nro,
                    numdoc_cuadre,
                    nro_documento,
                    name,
                    operation_type,
                    new_name,
                    default_code,
                    unidad,
                    ingreso,
                    salida,
                    cadquiere,
                    origen,
                    destino,
                    almacen,
                    ctanalitica,
                    stock_moveid,
                    pedido
                    from (
        select vst_kardex_sunat.*,sp.name as doc_almac,sm.kardex_date as fecha_albaran,
        vst_kardex_sunat.fecha - interval '5' hour as fechax,sl_o.usage as origen_usage , sl_d.usage as destino_usage, np.new_name,
        coalesce(po.name,so.name) as pedido
                    from vst_kardex_fisico_valorado as vst_kardex_sunat
        left join stock_move sm on sm.id = vst_kardex_sunat.stock_moveid
        left join purchase_order_line pol on pol.id = sm.purchase_line_id
        left join purchase_order po on po.id = pol.order_id

        left join sale_order_line sol on sol.id = sm.sale_line_id
        left join sale_order so on so.id = sol.order_id

        left join stock_picking sp on sp.id = sm.picking_id

                            left join stock_location sl_o on sl_o.id = sm.location_id
                            left join stock_location sl_d on sl_d.id = sm.location_dest_id
        left join (
            select t_pp.id, 
                    ((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
                    FROM product_product t_pp
                        JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
                        left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
                left join product_variant_combination pvc on pvc.product_product_id = t_pp.id
                left join product_template_attribute_value ptav on ptav.id = pvc.product_template_attribute_value_id
                left join product_attribute_value pav on pav.id = ptav.product_attribute_value_id
                group by t_pp.id
                ) np on np.id = vst_kardex_sunat.product_id
                            
                where (fecha_num((vst_kardex_sunat.fecha - interval '5' hour)::date) between """+str(date_ini).replace('-','')+""" and """+str(date_fin).replace('-','')+""")    
                and vst_kardex_sunat.location_id in """+str(almacenes)+""" and vst_kardex_sunat.product_id in """ +str(productos)+ """
                        and vst_kardex_sunat.company_id = """ +str(self.env.company.id)+ """
                    order by vst_kardex_sunat.location_id,vst_kardex_sunat.product_id,vst_kardex_sunat.fecha,vst_kardex_sunat.esingreso,vst_kardex_sunat.stock_moveid,vst_kardex_sunat.nro
                    )Total	
                    """+si_existe+"""			
    ) A order by location_id,product_id,fechax
                    
                """)
                total_all = self.env.cr.fetchall()
            else:

                self.env.cr.execute("""  
                    select  
                    product_id,
                    location_id,
                    origen_usage,
                    destino_usage,
                    debit,
                    credit,
                    fechax,
                    type_doc,
                    serial,
                    nro,
                    numdoc_cuadre,
                    nro_documento,
                    name,
                    operation_type,
                    new_name,
                    default_code,
                    unidad,
                    ingreso,
                    salida,
                    cadquiere,
                    origen,
                    destino,
                    almacen,
                    ctanalitica,
                    stock_moveid,
                    pedido
                from (
                    select 
                    product_id,
                    location_id,
                    origen_usage,
                    destino_usage,
                    debit,
                    credit,
                    fechax,
                    type_doc,
                    serial,
                    nro,
                    numdoc_cuadre,
                    nro_documento,
                    name,
                    operation_type,
                    new_name,
                    default_code,
                    unidad,
                    ingreso,
                    salida,
                    cadquiere,
                    origen,
                    destino,
                    almacen,
                    ctanalitica,
                    stock_moveid,
                    pedido
                    from (
        select vst_kardex_sunat.*,sp.name as doc_almac,sm.kardex_date as fecha_albaran,
        vst_kardex_sunat.fecha - interval '5' hour as fechax,sl_o.usage as origen_usage , sl_d.usage as destino_usage, np.new_name,
        po.name as pedido
                    from vst_kardex_fisico_valorado_dolar as vst_kardex_sunat
        left join stock_move sm on sm.id = vst_kardex_sunat.stock_moveid
        left join purchase_order_line pol on pol.id = sm.purchase_line_id
        left join purchase_order po on po.id = pol.order_id
        left join stock_picking sp on sp.id = sm.picking_id

                            left join stock_location sl_o on sl_o.id = sm.location_id
                            left join stock_location sl_d on sl_d.id = sm.location_dest_id
        left join (
            select t_pp.id, 
                    ((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
                    FROM product_product t_pp
                        JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
                        left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
                left join product_variant_combination pvc on pvc.product_product_id = t_pp.id
                left join product_template_attribute_value ptav on ptav.id = pvc.product_template_attribute_value_id
                left join product_attribute_value pav on pav.id = ptav.product_attribute_value_id
                group by t_pp.id
                ) np on np.id = vst_kardex_sunat.product_id
                            
                where (fecha_num((vst_kardex_sunat.fecha - interval '5' hour)::date) between """+str(date_ini).replace('-','')+""" and """+str(date_fin).replace('-','')+""")    
                and vst_kardex_sunat.location_id in """+str(almacenes)+""" and vst_kardex_sunat.product_id in """ +str(productos)+ """
                        and vst_kardex_sunat.company_id = """ +str(self.env.company.id)+ """
                    order by vst_kardex_sunat.location_id,vst_kardex_sunat.product_id,vst_kardex_sunat.fecha,vst_kardex_sunat.esingreso,vst_kardex_sunat.stock_moveid,vst_kardex_sunat.nro
                    )Total	
                    """+si_existe+"""			
    ) A order by location_id,product_id,fechax

                """)
                total_all = self.env.cr.fetchall()

            self.send_message(text_report_linea + "Generando Kardex<br/><center>Total lineas a generar: "+str(len(total_all)) + "</center>")
            cont_report = 0
            import datetime
            tiempo_inicial = datetime.datetime.now()
            tiempo_pasado = [0,0]
            cprom_data = {}

            ingreso1 =0
            ingreso2 =0
            salida1 =0
            salida2 =0

            for xl in total_all:
                l = {
                    'product_id':xl[0],
                    'location_id':xl[1],
                    'origen_usage':xl[2],
                    'destino_usage':xl[3],
                    'debit':xl[4],
                    'credit':xl[5],
                    'fechax':xl[6],
                    'type_doc':xl[7],
                    'serial':xl[8],
                    'nro':xl[9],
                    'numdoc_cuadre':xl[10],
                    'nro_documento':xl[11],
                    'name':xl[12],
                    'operation_type':xl[13],
                    'new_name':xl[14],
                    'default_code':xl[15],
                    'unidad':xl[16],
                    'ingreso':xl[17],
                    'salida':xl[18],
                    'cadquiere':xl[19],
                    'origen':xl[20],
                    'destino':xl[21],
                    'almacen':xl[22],
                    'ctanalitica':xl[23],	
                    'stock_moveid':xl[24],	
                    'pedido':xl[25],	
                }
                cont_report += 1
                if cont_report%300 == 0:
                    tiempo_pasado = divmod((datetime.datetime.now()-tiempo_inicial).seconds,60)
                    text_report = ""
                    if self.moneda == 'pen':
                        text_report = text_report_linea+ "<b>Generando Kardex.</b><br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center><br/><center>Procesado:"+str(cont_report)+"/"+str(len(total_all))+"</center><br/> Tiempo Procesado: "+ str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos" 
                    else:
                        text_report = text_report_linea+ "<b>Generando Kardex (USD).</b><br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center><br/><center>Procesado:"+str(cont_report)+"/"+str(len(total_all))+"</center><br/> Tiempo Procesado: "+ str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos" 
                    text_report += """<div id="myProgress" style="position: relative; width: 100%;  height: 30px;   background-color: white;">
                        <div id="myBar" style="position: absolute;  width: """+"%.2f"%(cont_report*100/len(total_all))+"""%;  height: 100%; background-color: #875A7B;">
                        <div id="label" style="text-align: center; line-height: 30px; color: white;">""" +"%.2f"%(cont_report*100/len(total_all))+ """%</div>
                        </div>
                    </div>"""
                    self.send_message(text_report)
                    
                llave = (l['product_id'],l['location_id'])
                cprom_acum = [0,0]
                if llave in cprom_data:
                    cprom_acum = cprom_data[llave]
                else:
                    cprom_data[llave] = cprom_acum

                cprom_act_antes = cprom_data[llave][1] / cprom_data[llave][0] if cprom_data[llave][0] != 0 else 0

                data_temp = {}
                
                data_temp = {'origen':l['origen_usage'] or '','destino':l['destino_usage'] or ''}
                ingreso_v = 0
                egreso_v = 0
                if l['ingreso'] or l['debit']:
                    if (data_temp['origen'] == 'internal' and data_temp['destino'] == 'internal') or (data_temp['origen'] == 'transit' and data_temp['destino'] == 'internal'):
                        cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
                        cprom_acum[1] = cprom_acum[1] + (l['debit'] if l['debit'] else 0) -  (l['credit'] if l['credit'] else 0)

                        ingreso_v = (l['debit'] if l['debit'] else 0) 
                    else:	
                        cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
                        cprom_acum[1] = cprom_acum[1] + (l['debit'] if l['debit'] else 0) -  (l['credit'] if l['credit'] else 0)

                        ingreso_v = (l['debit'] if l['debit'] else 0) 
                else:
                    if (data_temp['origen'] == 'internal' and data_temp['destino'] == 'supplier'):
                        cprom_acum[0] = cprom_acum[0] -  (l['salida'] if l['salida'] else 0)
                        cprom_acum[1] = cprom_acum[1] - (l['debit'] if l['debit'] else 0) - ( (l['credit'] if l['credit'] else 0) * (l['salida'] if l['salida'] else 0) )
                        egreso_v = (l['credit'] if l['credit'] else 0) * (l['salida'] if l['salida'] else 0)
                    else:
                        if l['salida']:
                            cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
                            cprom_acum[1] = cprom_acum[1] - (l['salida'] if l['salida'] else 0)*(cprom_act_antes if cprom_act_antes else 0)

                            egreso_v = (l['salida'] if l['salida'] else 0)*(cprom_act_antes if cprom_act_antes else 0)
                        else:
                            cprom_acum[0] = cprom_acum[0] + (l['ingreso'] if l['ingreso'] else 0) -  (l['salida'] if l['salida'] else 0)
                            cprom_acum[1] = cprom_acum[1] - (l['credit'] if l['credit'] else 0)

                            egreso_v = (l['credit'] if l['credit'] else 0)

                cprom_acum[1] = cprom_acum[1] if cprom_acum[0] > 0 else 0
                cprom_act = cprom_acum[1] / cprom_acum[0] if cprom_acum[0] != 0 else 0

                cprom_data[llave] = cprom_acum

                linea = []
                linea.append( str(l['fechax'])[:10] if l['fechax'] else '')
                linea.append( (str(l['fechax'])[11:])[:8] if l['fechax'] else  '')
                linea.append( l['type_doc'] if l['type_doc'] else '' )
                linea.append( l['serial'] if l['serial'] else '' )
                linea.append( l['nro'] if l['nro'] else '' )
                linea.append( l['numdoc_cuadre'] if l['numdoc_cuadre'] else '')
                linea.append( l['nro_documento'] if l['nro_documento'] else '' )
                linea.append( l['name'] if l['name'] else '' )
                linea.append( l['operation_type'] if l['operation_type'] else '' )
                linea.append( l['new_name'] if l['new_name'] else '' )
                linea.append( l['default_code'] if l['default_code'] else '')
                linea.append( l['unidad'] if l['unidad'] else '' )
                linea.append( number_format_cantidad(ws, l['ingreso'] if l['ingreso'] else 0 ))
                linea.append( number_format_costo(ws, ingreso_v ))
                linea.append( number_format_cantidad(ws, l['salida'] if l['salida'] else 0 ))
                linea.append( number_format_costo(ws, egreso_v ))
                linea.append( number_format_cantidad(ws, cprom_acum[0] if len(cprom_acum)>1 and cprom_acum[0] else 0 ))
                linea.append( number_format_costo(ws, cprom_acum[1] if len(cprom_acum)>1 and cprom_acum[1] else 0 ))
                linea.append( number_format_costo(ws, l['cadquiere'] if l['cadquiere'] else 0))
                linea.append( number_format_costo(ws, cprom_act ))
                linea.append( l['origen'] if l['origen'] else '' )
                linea.append( l['destino'] if l['destino'] else '' )
                linea.append( l['almacen'] if l['almacen'] else '' )

                linea.append( l['product_id'] if l['product_id'] else '' )
                linea.append( l['pedido'] if l['pedido'] else '' )

                
                if self.check_account:
                    move = self.env['product.product'].browse(l['product_id'])
                    if move.id:
                        linea.append(  move.categ_id.property_stock_valuation_account_id.code)
                        linea.append(  move.categ_id.property_stock_account_output_categ_id.code )
                        linea.append(  l['ctanalitica'] or ''  )

                ws.append(linea)
                
                ingreso1 += l['ingreso'] or 0
                ingreso2 += ingreso_v or 0
                salida1 += l['salida'] or 0
                salida2 += egreso_v or 0

                x = x +1

            tam_col = [11,11,5,5,7,5,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11]

            ws.append([""]*11+["TOTALES:",ingreso1 ,ingreso2 , salida1, salida2])


            workbook.save(output)
            output.seek(0)

            attach_id = self.env['ir.attachment'].create({
                        'name': "Kardex Valorado.xlsx",
                        'type': 'binary',
                        'datas': base64.encodebytes(output.getvalue()),
                        'eliminar_automatico': True
                    })
            output.close()


            return {
                'type': 'ir.actions.client',
                'tag': 'notification_llikha',
                'params': {
                    'title':'Kardex Valorado Excel' if self.moneda == 'pen' else 'Kardex Valorado Excel (USD)',
                    'type': 'success',
                    'sticky': True,
                    'message': 'Se proceso de '+str(self.fini)+' al ' + str(self.ffin)+ '.<br/>Lineas procesadas: '+ str(len(total_all)) +'<br/>Tiempo: ' + str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos",
                    'next': {'type': 'ir.actions.act_window_close'},
                    'buttons':[{
                        'label':'Descargar Kardex Valorado',
                        'model':'ir.attachment',
                        'method':'get_download_ls',
                        'id':attach_id.id,
                        }
                    ],
                }
            }


        else:

            cad = ""

            s_prod = [-1,-1,-1]
            s_loca = [-1,-1,-1]
            if self.alllocations == True:
                locat_ids = self.env['stock.location'].search( [('usage','in',('internal','internal'))] )
                lst_locations = locat_ids.ids
            else:
                lst_locations = self.location_ids.ids
            lst_products  = self.products_ids.ids
            productos='{'
            almacenes='{'
            date_ini=self.fini
            date_fin=self.ffin
            if self.allproducts:
                lst_products = self.env['product.product'].with_context(active_test=False).search([('product_tmpl_id.check_unidad_especifica','=',True)]).ids

            else:
                lst_products = self.products_ids.filtered(lambda r: r.product_tmpl_id.check_unidad_especifica == True).ids

            if len(lst_products) == 0:
                raise UserError('No existen productos seleccionados')

            for producto in lst_products:
                productos=productos+str(producto)+','
                s_prod.append(producto)
            productos=productos[:-1]+'}'
            for location in lst_locations:
                almacenes=almacenes+str(location)+','
                s_loca.append(location)
            almacenes=almacenes[:-1]+'}'


            config = self.env['kardex.parameter'].search([('company_id','=',self.env.company.id)])

            date_ini = '%d-01-01' % ( config._get_anio_start(date_ini.year) )

            kardex_save_obj = self.env['kardex.save'].search([('company_id','=',self.env.company.id),('state','=','done'),('name.date_end','<',self.fini),('name.fiscal_year_id.name','=',str(date_fin.year) )]).sorted(lambda l: l.name.code , reverse=True)
            if len(kardex_save_obj)>0:
                kardex_save_obj = kardex_save_obj[0]
                date_ini = kardex_save_obj.name.date_end + timedelta(days=1)
                
            
            si_existe = ""
            if kardex_save_obj:
                si_existe = """ select '"""+str(kardex_save_obj.name.date_end)+"""'::date as fecha,
                'SALDO INICIAL' as origen, sl.complete_name as destino, sl.complete_name as almacen, ksp.stock as entrada, 0 as salida, null as stock_move,
                '' as motivo_guia, pname.new_name as producto, 'done' as estado, 'Saldo Inicial' as name, coalesce(pp.default_code,pt.default_code) as cod_pro,
                pc.name as categoria, uu.name as unidad, ksp.producto as product_id, ksp.almacen as almacen_id, spt.name as lote, spt.id as lote_id
                from kardex_save_period ksp 
                inner join stock_location sl on sl.id = ksp.almacen
                inner join product_product pp on pp.id = ksp.producto
                inner join product_template pt on pt.id = pp.product_tmpl_id
                inner join uom_uom uu on uu.id = pt.uom_id
                inner join product_category pc on pc.id = pt.categ_id
                left join stock_production_lot spt on spt.id = ksp.lote
                
                    LEFT JOIN ( SELECT t_pp.id,
                        ((     coalesce(max(it.value),max(t_pt.name::text))::character varying::text || ' '::text) || replace(array_agg(pav.name)::character varying::text, '{NULL}'::text, ''::text))::character varying AS new_name
                        FROM product_product t_pp
                            JOIN product_template t_pt ON t_pp.product_tmpl_id = t_pt.id
                            left join ir_translation it ON t_pt.id = it.res_id and it.name = 'product.template,name' and it.lang = 'es_PE' and it.state = 'translated'
                            LEFT JOIN product_variant_combination pvc ON pvc.product_product_id = t_pp.id
                            LEFT JOIN product_template_attribute_value ptav ON ptav.id = pvc.product_template_attribute_value_id
                            LEFT JOIN product_attribute_value pav ON pav.id = ptav.product_attribute_value_id
                        GROUP BY t_pp.id) pname ON pname.id = pp.id
                    where save_id = """+str(kardex_save_obj.id)+"""
                    
                    and ksp.almacen in  """+str(tuple(s_loca))+"""
                    and ksp.producto in  """+str(tuple(s_prod))+"""
                union all """


            import io
            from xlsxwriter.workbook import Workbook
            output = io.BytesIO()

            #direccion = self.env['main.parameter'].search([])[0].dir_create_file
            workbook = Workbook(output, {'constant_memory': True})
            worksheet = workbook.add_worksheet("Kardex")
            bold = workbook.add_format({'bold': True})
            bold.set_font_size(8)
            normal = workbook.add_format()
            boldbord = workbook.add_format({'bold': True})
            boldbord.set_border(style=2)
            boldbord.set_align('center')
            boldbord.set_align('vcenter')
            boldbord.set_text_wrap()
            boldbord.set_font_size(8)
            boldbord.set_bg_color('#DCE6F1')

            especial1 = workbook.add_format({'bold': True})
            especial1.set_align('center')
            especial1.set_align('vcenter')
            especial1.set_text_wrap()
            especial1.set_font_size(15)

            numbertres = workbook.add_format({'num_format':'0.000'})
            numberdos = workbook.add_format({'num_format':'0.00'})
            numberseis = workbook.add_format({'num_format':'0.000000'})
            numberseis.set_font_size(8)
            numberocho = workbook.add_format({'num_format':'0.00000000'})
            numberocho.set_font_size(8)
            bord = workbook.add_format()
            bord.set_border(style=1)
            bord.set_font_size(8)
            numberdos.set_border(style=1)
            numberdos.set_font_size(8)
            numbertres.set_border(style=1)
            numberseis.set_border(style=1)
            numberocho.set_border(style=1)
            numberdosbold = workbook.add_format({'num_format':'0.00','bold':True})
            numberdosbold.set_font_size(8)
            
            formatdate = workbook.add_format({'num_format': 'dd-mm-yyyy'})
            formattime = workbook.add_format({'num_format': 'HH:mm'})
            x= 10
            tam_col = [0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0,0]
            tam_letra = 1.2

            worksheet.merge_range(1,5,1,10, "KARDEX COSTEO UNIDADES ESPECIFICAS", especial1)
            worksheet.write(2,0,'FECHA INICIO:',bold)
            worksheet.write(3,0,'FECHA FIN:',bold)

            worksheet.write(2,1,str(self.fini))
            worksheet.write(3,1,str(self.ffin))
            import datetime

            worksheet.write(8,12, "Ingreso",boldbord)
            worksheet.write(8,13, "Salida",boldbord)		
            worksheet.write(8,14, "Saldo",boldbord)

            worksheet.write(9,0, u"Fecha",boldbord)
            worksheet.write(9,1, u"Hora",boldbord)
            worksheet.write(9,2, u"Ubicacion Origen",boldbord)
            worksheet.write(9,3, u"Ubicacion Destino",boldbord)
            worksheet.write(9,4, u"Almacen",boldbord)
            worksheet.write(9,5, u"Tipo de Operación",boldbord)
            worksheet.write(9,6, u"Nombre del Tipo de Operación",boldbord)
            worksheet.write(9,7, u"Categoria",boldbord)
            worksheet.write(9,8, u"Producto",boldbord)
            worksheet.write(9,9, u"Codigo P.",boldbord)
            worksheet.write(9,10, u"Lote",boldbord)
            worksheet.write(9,11, u"Unidad",boldbord)
            worksheet.write(9,12, u"Doc. Almacen",boldbord)
            worksheet.write(9,13, "Cantidad",boldbord)
            worksheet.write(9,14, "Cantidad",boldbord)
            worksheet.write(9,15, "Cantidad",boldbord)
            worksheet.write(9,16, "Precio Unitario",boldbord)			
            worksheet.write(9,17, "Cliente",boldbord)
            worksheet.write(9,18, "Nro Factura",boldbord)




            self.env.cr.execute("""
                select 
                    (vstf.fecha - interval '5' hour)::date as "Fecha",
                    (vstf.fecha - interval '5' hour) ::time as "Hora",
                    origen AS "Ubicación Origen",
                    destino AS "Ubicación Destino",
                    almacen AS "Almacén",
                    vstf.motivo_guia::varchar AS "Tipo de operación",
                    categoria as "Categoria",
                    producto as "Producto",
                    cod_pro as "Codigo P.",
                    vstf.lote as "Lote",
                    unidad as "unidad",
                    vstf.name as "Doc. Almacén",
                    vstf.entrada as "Entrada",
                    vstf.salida as "Salida",
                    rp.name as partner,
                    am.ref as comprobante,
                    vstf.lote_id as loteid,
                    --sp.numberg as guia
                    tok.name as operation_type_label
                from
                    (""" +si_existe+ """	
                        select 
                            vst_kardex_fisico.date as fecha,
                            u_origen as origen, 
                            u_destino as destino, 
                            u_destino as almacen, 
                            vst_kardex_fisico.product_qty as entrada, 
                            0 as salida,
                            vst_kardex_fisico.id  as stock_move,
                            vst_kardex_fisico.guia as motivo_guia, 
                            producto,
                            vst_kardex_fisico.estado,
                            vst_kardex_fisico.name, 
                            cod_pro, 
                            categoria, 
                            unidad,
                            product_id,
                            location_dest_id as almacen_id,
                            lote, 
                            lote_id 
                        from 
                            vst_kardex_fisico_lote() as vst_kardex_fisico
                        where
                            company_id = """+str(self.env.company.id)+""" and 
                            (date - interval '5' hour)::date >='""" +str(date_ini)+ """' and 
                            (date - interval '5' hour)::date <='""" +str(date_fin)+ """'			
                        union all
                        select 
                            vst_kardex_fisico.date as fecha,
                            u_origen as origen, 
                            u_destino as destino, 
                            u_origen as almacen, 
                            0 as entrada, 
                            vst_kardex_fisico.product_qty as salida,
                            vst_kardex_fisico.id  as stock_move ,
                            vst_kardex_fisico.guia as motivo_guia ,
                            producto ,
                            vst_kardex_fisico.estado,
                            vst_kardex_fisico.name, 
                            cod_pro, 
                            categoria, 
                            unidad,
                            product_id, 
                            location_id as almacen_id , 
                            lote, 
                            lote_id from vst_kardex_fisico_lote() as vst_kardex_fisico 
                        where 
                            company_id = """+str(self.env.company.id)+""" and 
                            (date - interval '5' hour)::date >='""" +str(date_ini)+ """' and 
                            (date - interval '5' hour)::date <='""" +str(date_fin)+ """'			
                    ) as vstf
                    left join stock_move sm on sm.id = vstf.stock_move
                    left join stock_picking sp on sp.id = sm.picking_id
                    left join res_partner rp on rp.id =  sp.partner_id
                    left join account_move am on am.id = sm.invoice_id
                    
                    left join type_operation_kardex tok on tok.id = sp.type_operation_sunat_id
                where 
                    vstf.product_id in """ +str(tuple(s_prod))+ """ and 
                    vstf.almacen_id in """ +str(tuple(s_loca))+ """ and 
                    vstf.estado = 'done'
                order by
                    almacen,producto,lote,vstf.fecha
            ;
            """)

            ingreso1= 0
            ingreso2= 0
            salida1= 0
            salida2= 0

            saldo = 0
            almacen = None
            producto = None
            lote = None

            tiempo_inicial = datetime.datetime.now()
            tiempo_pasado = [0,0]
            cont_report = 0

            total_all = self.env.cr.fetchall()
            for line in total_all:
                cont_report += 1
                if cont_report%300 == 0:
                    tiempo_pasado = divmod((datetime.datetime.now()-tiempo_inicial).seconds,60)
                    text_report = ""
                    text_report = "<b>Generando Kardex Fisico x Lote.</b><br/><center>Total lineas a procesar: "+str(len(total_all)) + "</center><br/><center>Procesado:"+str(cont_report)+"/"+str(len(total_all))+"</center><br/> Tiempo Procesado: "+ str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos" 
                    text_report += """<div id="myProgress" style="position: relative; width: 100%;  height: 30px;   background-color: white;">
                        <div id="myBar" style="position: absolute;  width: """+"%.2f"%(cont_report*100/len(total_all))+"""%;  height: 100%; background-color: #875A7B;">
                        <div id="label" style="text-align: center; line-height: 30px; color: white;">""" +"%.2f"%(cont_report*100/len(total_all))+ """%</div>
                        </div>
                    </div>"""
                    self.send_message(text_report)

                if almacen == None:
                    almacen = (line[4] if line[4] else '')
                    producto = (line[7] if line[7] else '')
                    lote = (line[9] if line[9] else '')
                    saldo = line[12] - line[13]
                elif almacen != (line[4] if line[4] else '') or producto != (line[7] if line[7] else '') or lote != (line[9] if line[9] else ''):
                    almacen = (line[4] if line[4] else '')
                    producto = (line[7] if line[7] else '')
                    lote = (line[9] if line[9] else '')
                    saldo = line[12] - line[13]
                else:
                    saldo = saldo + line[12] - line[13]

                worksheet.write(x,0,line[0] if line[0] else '' ,formatdate )
                worksheet.write(x,1,line[1] if line[1] else '' ,formattime )
                worksheet.write(x,2,line[2] if line[2] else '' ,bord )
                worksheet.write(x,3,line[3] if line[3] else '' ,bord )
                worksheet.write(x,4,line[4] if line[4] else '' ,bord )
                worksheet.write(x,5,line[5] if line[5] else '' ,bord )
                worksheet.write(x,6,line[17] if line[17] else '' ,bord )
                worksheet.write(x,7,line[6] if line[6] else '' ,bord )
                worksheet.write(x,8,line[7] if line[7] else '' ,bord )
                worksheet.write(x,9,line[8] if line[8] else '' ,bord )
                worksheet.write(x,10,line[9] if line[9] else '' ,bord)
                worksheet.write(x,11,line[10] if line[10] else '' ,bord )
                worksheet.write(x,12,line[11] if line[11] else 0 ,numberdos )
                worksheet.write(x,13,line[12] if line[12] else 0 ,numberdos )
                worksheet.write(x,14,line[13] if line[13] else 0 ,numberdos )
                worksheet.write(x,15,saldo ,numberdos )

                idlote = line[16] if line[16] else 0
                lotegg = self.env['stock.production.lot'].browse(idlote)
                pue = 0
                if lotegg:
                    pue = lotegg[0].precio_final

                worksheet.write(x,16,pue ,numberdos )
                worksheet.write(x,17,line[14] if line[14] else '' ,bord )
                #worksheet.write(x,16,line[16] if line[16] else '' ,bord )
                worksheet.write(x,18,line[15] if line[15] else '' ,bord )


                x = x +1

            tam_col = [11,11,5,5,7,5,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11,11]


            worksheet.set_column('A:A', tam_col[0])
            worksheet.set_column('B:B', tam_col[1])
            worksheet.set_column('C:C', tam_col[2])
            worksheet.set_column('D:D', tam_col[3])
            worksheet.set_column('E:E', tam_col[4])
            worksheet.set_column('F:F', tam_col[5])
            worksheet.set_column('G:G', tam_col[6])
            worksheet.set_column('H:H', tam_col[7])
            worksheet.set_column('I:I', tam_col[8])
            worksheet.set_column('J:J', tam_col[9])
            worksheet.set_column('K:K', tam_col[10])
            worksheet.set_column('L:L', tam_col[11])
            worksheet.set_column('M:M', tam_col[12])
            worksheet.set_column('N:N', tam_col[13])
            worksheet.set_column('O:O', tam_col[14])
            worksheet.set_column('P:P', tam_col[15])
            worksheet.set_column('Q:Q', tam_col[16])
            worksheet.set_column('R:R', tam_col[17])
            worksheet.set_column('S:S', tam_col[18])
            worksheet.set_column('T:Z', tam_col[19])


            workbook.close()
            output.seek(0)

            attach_id = self.env['ir.attachment'].create({
                        'name': "Kardex Valorado Unidades Especificas Excel.xlsx",
                        'type': 'binary',
                        'datas': base64.encodebytes(output.getvalue()),
                        'eliminar_automatico': True
                    })
            output.close()
            
            return {
                'type': 'ir.actions.client',
                'tag': 'notification_llikha',
                'params': {
                    'title':'Kardex Valorado Unidades Especificas Excel',
                    'type': 'success',
                    'sticky': True,
                    'message': 'Se proceso de '+str(self.fini)+' al ' + str(self.ffin)+ '.<br/>Lineas procesadas: '+ str(len(total_all)) +'<br/>Tiempo: ' + str(tiempo_pasado[0])+" minutos " +str(tiempo_pasado[1]) + " segundos",
                    'next': {'type': 'ir.actions.act_window_close'},
                    'buttons':[{
                        'label':'Descargar Kardex Valorado Unidades Especificas Excel',
                        'model':'ir.attachment',
                        'method':'get_download_ls',
                        'id':attach_id.id,
                        }
                    ],
                }
            }

