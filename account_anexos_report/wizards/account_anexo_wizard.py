# -*- coding: utf-8 -*-

from odoo import models, fields, api
from odoo.exceptions import UserError
import base64
from io import BytesIO
from datetime import *
import base64
import subprocess
import sys

def install(package):
	subprocess.check_call([sys.executable, "-m", "pip", "install", package])

try:
	import openpyxl
except:
	install('openpyxl==3.0.5')

class AccountAnexoWizard(models.TransientModel):
	_name = 'account.anexo.wizard'
	_description = 'Account Anexo Wizard'

	name = fields.Char()
	company_id = fields.Many2one('res.company',string=u'Compañia',required=True, default=lambda self: self.env.company,readonly=True)
	fiscal_year_id = fields.Many2one('account.fiscal.year',string='Ejercicio',required=True)
	period_from = fields.Many2one('account.period',string='Periodo Inicial',required=True)
	period_to = fields.Many2one('account.period',string='Periodo Final',required=True)
	
	@api.onchange('company_id')
	def get_fiscal_year(self):
		if self.company_id:
			today = fields.Date.context_today(self)
			fiscal_year = self.env['account.fiscal.year'].search([('name','=',str(today.year))],limit=1)
			if fiscal_year:
				self.fiscal_year_id = fiscal_year.id
	
	def get_report(self):
		self.ensure_one()
		self.env.cr.execute(self._get_f1_balance_sql())
		res = self.env.cr.fetchall()
		colnames = [
			desc[0] for desc in self.env.cr.description
		]
		res.insert(0, colnames)

		wb = openpyxl.Workbook()
		ws = wb.active
		ws.title = 'HT BALAN'
		row_position = 1
		col_position = 1
		for index, row in enumerate(res, row_position):
			for col, val in enumerate(row, col_position):
				ws.cell(row=index, column=col).value = val
		
		self.env.cr.execute(self._get_f1_register_sql())
		resx = self.env.cr.fetchall()
		colnames = [
			desc[0] for desc in self.env.cr.description
		]
		resx.insert(0, colnames)
		wsx = wb.create_sheet('HT REG')
		row_position = 1
		col_position = 1
		for index, row in enumerate(resx, row_position):
			for col, val in enumerate(row, col_position):
				wsx.cell(row=index, column=col).value = val
		
		############
		self.env.cr.execute(self._get_sql_10())
		resx = self.env.cr.fetchall()
		colnames = [
			desc[0] for desc in self.env.cr.description
		]
		resx.insert(0, colnames)
		wsx = wb.create_sheet('10')
		row_position = 1
		col_position = 1
		for index, row in enumerate(resx, row_position):
			for col, val in enumerate(row, col_position):
				wsx.cell(row=index, column=col).value = val

		############
		code = ['12','13','14','16','17','18','37','41','42','43','44','45','46','47','48','49']
		for cod in code:
			self.env.cr.execute(self.sql_saldos(cod))
			resx = self.env.cr.fetchall()
			colnames = [
				desc[0] for desc in self.env.cr.description
			]
			resx.insert(0, colnames)
			wsx = wb.create_sheet(cod)
			row_position = 1
			col_position = 1
			for index, row in enumerate(resx, row_position):
				for col, val in enumerate(row, col_position):
					wsx.cell(row=index, column=col).value = val
		##
					
		output = BytesIO()
		wb.save(output)
		output.getvalue()
		output_datas = base64.b64encode(output.getvalue())
		output.close()

		return self.env['popup.it'].get_file('Anexos.xlsx',output_datas)
	
	def _get_sql_10(self):
		sql = """
			select T.* from (
			(select 
			a2.code as cuenta,
			a2.name as nomenclatura,
			(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) - (case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) as saldo
			from get_sumas_mayor_f1('{date_from}','{date_to}',{company_id},TRUE) a1
			left join account_account a2 on a2.id=a1.account_id
			where left(a2.code,2) = '10' and (case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) - (case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) <> 0
			order by a2.code)
			UNION ALL
			(select
			null::character varying as cuenta,
			'SUMAS'::character varying as nomenclatura,
			sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) - sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) as saldo
			from get_sumas_mayor_f1('{date_from}','{date_to}',{company_id},TRUE) a1
			left join account_account a2 on a2.id=a1.account_id
			where left(a2.code,2) = '10')
			)T
		""".format(
				date_from = self.period_from.date_start.strftime('%Y/%m/%d'),
				date_to = self.period_to.date_end.strftime('%Y/%m/%d'),
				company_id = self.company_id.id
			)
		return sql

	def _get_f1_register_sql(self):
		sql = """
			select T.* from (
			(select 
			left(a2.code,2) as mayor,
			a2.code as cuenta,
			a2.name as nomenclatura,
			a1.debe,
			a1.haber,
			case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end as saldo_deudor,
			case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end as saldo_acreedor,
			case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end as activo,
			case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end as pasivo,
			case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end as perdinat,
			case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end as ganannat,
			case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end as perdifun,
			case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end as gananfun,
			ati.name as rubro
			from get_sumas_mayor_f1('{date_from}','{date_to}',{company_id},TRUE) a1
			left join account_account a2 on a2.id=a1.account_id
			left join account_type_it ati on ati.id = a2.account_type_it_id
			order by a2.code)
			UNION ALL
			(select
			null::character varying as mayor,
			null::character varying as cuenta,
			'SUMAS'::character varying as nomenclatura,
			sum(a1.debe) as debe,
			sum(a1.haber) as haber,
			sum(case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end) as saldo_deudor,
			sum(case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end) as saldo_acreedor,
			sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) as activo,
			sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) as pasivo,
			sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) as perdinat,
			sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) as ganannat,
			sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) as perdifun,
			sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) as gananfun,
			null::character varying as rubro
			from get_sumas_mayor_f1('{date_from}','{date_to}',{company_id},TRUE) a1
			left join account_account a2 on a2.id=a1.account_id)
			UNION ALL
			(
				SELECT 
				null::character varying as mayor,
				null::character varying as cuenta,
				'UTILIDAD O PERDIDA'::character varying as nomenclatura,
				case
					when sum(a1.debe) < sum(a1.haber)
					then sum(a1.haber) - sum(a1.debe)
					else 0
				end as debe,
				case
					when sum(a1.debe) > sum(a1.haber)
					then sum(a1.debe) - sum(a1.haber) 
					else 0
				end as haber,
				case
					when sum(case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end) < sum(case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end)
					then sum(case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end) - sum(case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end)
					else 0
				end as saldo_deudor,
				case
					when sum(case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end) > sum(case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end)
					then sum(case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end) - sum(case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end)
					else 0
				end as saldo_acreedor,
				case
					when sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) < sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
					then sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) - sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end)
					else 0
				end as activo,
				case
					when sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) > sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
					then sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) - sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
					else 0
				end as pasivo,
				case
					when sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) < sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
					then sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) - sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end)
					else 0
				end as perdinat,
				case
					when sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) > sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
					then sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) - sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
					else 0
				end as ganannat,
				case
					when sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) < sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
					then sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) - sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end)
					else 0
				end as perdifun,
				case
					when sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) > sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
					then sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) - sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
					else 0
				end as gananfun,
				null::character varying as rubro
				from get_sumas_mayor_f1('{date_from}','{date_to}',{company_id},TRUE) a1
				left join account_account a2 on a2.id=a1.account_id
			)
			)T
		""".format(
				date_from = self.period_from.date_start.strftime('%Y/%m/%d'),
				date_to = self.period_to.date_end.strftime('%Y/%m/%d'),
				company_id = self.company_id.id
			)
		return sql
		
	def _get_f1_balance_sql(self):
		sql = """
			select T.* FROM (
			(select
			a2.code_prefix_start as cuenta,
			a2.name as nomenclatura,
			a1.debe,
			a1.haber,
			case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end as saldo_deudor,
			case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end as saldo_acreedor,
			case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end as activo,
			case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end as pasivo,
			case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end as perdinat,
			case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end as ganannat,
			case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end as perdifun,
			case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end as gananfun
			from get_sumas_balance_f1('{date_from}','{date_to}',{company_id},TRUE) a1
			left join account_group a2 on a2.id=a1.account_id
			order by a2.code_prefix_start,a2.name)
			UNION ALL
			(SELECT 
			null::character varying as cuenta,
			'SUMAS'::character varying as nomenclatura,
			sum(a1.debe) as debe,
			sum(a1.haber) as haber,
			sum(case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end) as saldo_deudor,
			sum(case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end) as saldo_acreedor,
			sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) as activo,
			sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) as pasivo,
			sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) as perdinat,
			sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) as ganannat,
			sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) as perdifun,
			sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) as gananfun
			from get_sumas_balance_f1('{date_from}','{date_to}',{company_id},TRUE) a1
			left join account_group a2 on a2.id=a1.account_id
			)
			UNION ALL
			(
			SELECT 
			null::character varying as cuenta,
			'UTILIDAD O PERDIDA'::character varying as nomenclatura,
			case
				when sum(a1.debe) < sum(a1.haber)
				then sum(a1.haber) - sum(a1.debe)
				else 0
			end as debe,
			case
				when sum(a1.debe) > sum(a1.haber)
				then sum(a1.debe) - sum(a1.haber) 
				else 0
			end as haber,
			case
				when sum(case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end) < sum(case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end)
				then sum(case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end) - sum(case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end)
				else 0
			end as saldo_deudor,
			case
				when sum(case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end) > sum(case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end)
				then sum(case when a1.debe > a1.haber then a1.debe-a1.haber else  0 end) - sum(case when a1.haber> a1.debe then a1.haber-a1.debe else 0 end)
				else 0
			end as saldo_acreedor,
			case
				when sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) < sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
				then sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) - sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end)
				else 0
			end as activo,
			case
				when sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) > sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
				then sum(case when a2.clasification_sheet='0' and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) - sum(case when a2.clasification_sheet='0' and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
				else 0
			end as pasivo,
			case
				when sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) < sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
				then sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) - sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end)
				else 0
			end as perdinat,
			case
				when sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) > sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
				then sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) - sum(case when (a2.clasification_sheet='1' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
				else 0
			end as ganannat,
			case
				when sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) < sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
				then sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end) - sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end)
				else 0
			end as perdifun,
			case
				when sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) > sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
				then sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.debe>a1.haber then a1.debe-a1.haber else 0 end) - sum(case when (a2.clasification_sheet='2' or a2.clasification_sheet='3') and a1.haber>a1.debe then a1.haber-a1.debe else 0 end)
				else 0
			end as gananfun
			from get_sumas_balance_f1('{date_from}','{date_to}',{company_id},TRUE) a1
			left join account_group a2 on a2.id=a1.account_id
			)
			)T
		""".format(
				date_from = self.period_from.date_start.strftime('%Y/%m/%d'),
				date_to = self.period_to.date_end.strftime('%Y/%m/%d'),
				company_id = self.company_id.id
			)
		return sql
	
	def sql_saldos(self,code):
		sql = """
			SELECT periodo, fecha_con, libro,voucher,td_partner,doc_partner,partner,td_sunat,nro_comprobante,
			fecha_doc,fecha_ven,cuenta,moneda,debe,haber,saldo_mn,saldo_me
			FROM get_saldos_anexos('{date_from}','{date_to}',{company},'{code}') 
			WHERE saldo_mn <> 0
		""".format(
				date_from = self.period_from.date_start.strftime('%Y/%m/%d'),
				date_to = self.period_to.date_end.strftime('%Y/%m/%d'),
				company = self.company_id.id,
				code = code
			)
		return sql